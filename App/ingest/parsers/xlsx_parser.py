import openpyxl


def parse_xlsx(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Hoja: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)
