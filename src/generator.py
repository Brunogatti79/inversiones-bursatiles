""" 
src/generator.py
Genera el dashboard HTML dark y las fichas Excel.
Versión actualizada con:
  1. Variación diaria en cards Panorama + refresh cada 10s
  2. Leyenda en gráfico base-100 + eje temporal con fecha más reciente a la derecha
  3. MERVAL usa gráfico de líneas (igual que BOVESPA y S&P 500)
  4. Sección "Radar de Oportunidades Tempranas" en Conclusiones (ranking por score)
  5. Cards Panorama más grandes con mejor jerarquía visual
  6. Sort cronológico real en gráfico comparativo base-100
  7. [NUEVO] Solapa "Oportunidades de Compra" con análisis técnico completo
"""
 
import os
import json
import logging
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Optional
 
logger = logging.getLogger(__name__)
logger.warning("=== GENERATOR VERSION NUEVA 2026-05-29 ===")
 
 
# ─────────────────────────────────────────────
# Helpers técnicos para oportunidades
# ─────────────────────────────────────────────
 
def _rsi(serie, p=14):
    d = serie.diff().dropna()
    g = d.clip(lower=0).rolling(p).mean()
    l = (-d.clip(upper=0)).rolling(p).mean()
    rs = g / l.replace(0, float('nan'))
    r = 100 - 100/(1+rs)
    return float(r.iloc[-1]) if len(r) >= p else 50.0
 
def _momentum(serie, p=21):
    if len(serie) < p+1: return 0.0
    return float((serie.iloc[-1]/serie.iloc[-p]-1)*100)
 
def _find_levels(serie, window=15):
    highs = serie.rolling(window, center=True).max()
    lows  = serie.rolling(window, center=True).min()
    precio = float(serie.iloc[-1])
    res = sorted(set([round(float(r),2) for r in serie[serie==highs].dropna().values if r > precio]))[:3]
    sup = sorted(set([round(float(s),2) for s in serie[serie==lows].dropna().values  if s < precio]), reverse=True)[:3]
    return sup, res
 
def _build_oportunidades(signals, price_data):
    """
    Construye fichas de oportunidades de compra con análisis técnico.
    Si no encuentra la serie de precios en los CSVs, construye la ficha
    con los datos de SIGNALS (sin gráfico de velas).
    price_data: dict {'merval': df, 'bovespa': df, 'sp500': df}
    """
    compras = [s for s in signals if 'COMPRA' in (s.get('signal_v2') or s.get('signal','')) and 'VENTA' not in (s.get('signal_v2') or s.get('signal',''))]

    fichas = []
    for s in compras:
        ticker  = s['ticker']
        empresa = s['empresa']
        market  = s['mercado']
        flag    = '🇦🇷' if market=='MERVAL' else '🇧🇷' if market=='BOVESPA' else '🇺🇸'
        moneda  = 'ARS' if market=='MERVAL' else 'BRL' if market=='BOVESPA' else 'USD'

        df_key = 'merval' if market=='MERVAL' else 'bovespa' if market=='BOVESPA' else 'sp500'
        df     = price_data.get(df_key)

        # Buscar columna — primero por ticker, luego por palabras del nombre
        col = None
        if df is not None and not df.empty:
            ticker_base = ticker.replace('.BA', '')
            for c in df.columns:
                if ticker_base.upper() in c.upper():
                    col = c; break
            if col is None:
                for c in df.columns:
                    if any(w.lower() in c.lower() for w in empresa.split()[:2] if len(w) > 3):
                        col = c; break

        # ── Rama con serie de precios (cálculo técnico completo) ──
        if col is not None:
            serie = df[col].dropna()
            if len(serie) < 20:
                col = None  # caer al fallback

        if col is not None:
            serie    = df[col].dropna()
            precio   = float(serie.iloc[-1])
            max12m   = float(serie.max())
            min12m   = float(serie.min())
            max_dt   = serie.idxmax().strftime('%d/%m/%Y') if hasattr(serie.idxmax(), 'strftime') else ''
            min_dt   = serie.idxmin().strftime('%d/%m/%Y') if hasattr(serie.idxmin(), 'strftime') else ''
            dist_max = round((max12m - precio) / max12m * 100, 1) if max12m > 0 else 0

            ma20  = float(serie.rolling(20).mean().iloc[-1]) if len(serie)>=20 else precio
            ma50  = float(serie.rolling(50).mean().iloc[-1]) if len(serie)>=50 else precio
            ma200 = float(serie.rolling(200).mean().iloc[-1]) if len(serie)>=200 else None

            sup, res = _find_levels(serie)

            entrada = round(min(precio, ma20) * 0.99, 2)
            stop    = round(s.get('atr_stop') or (sup[1]*0.985 if len(sup)>1 else entrada*0.94), 2)
            target  = round(s.get('atr_target') or (res[0]*0.995 if res else precio*1.12), 2)
            riesgo  = round((entrada-stop)/entrada*100, 1) if entrada > 0 else 0
            reward  = round((target-entrada)/entrada*100, 1) if entrada > 0 else 0
            rr      = round(min(5.0, s.get('rr_ratio') or (reward/riesgo if riesgo > 0 else 0)), 2)

            tail60    = serie.tail(60)
            closes60  = [round(float(v), 2) for v in tail60.values]
            dates60   = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d) for d in tail60.index]
            ma20_line, ma50_line = [], []
            if len(serie) >= 80:
                ma20_line = [round(float(v),2) if not pd.isna(v) else None for v in serie.rolling(20).mean().tail(60).values]
            if len(serie) >= 110:
                ma50_line = [round(float(v),2) if not pd.isna(v) else None for v in serie.rolling(50).mean().tail(60).values]

            ma_cross  = bool(serie.rolling(20).mean().iloc[-1] > serie.rolling(50).mean().iloc[-1]) if len(serie)>=50 else False
            momentum  = round(s.get('momentum_21d', _momentum(serie)), 1)
            sin_grafico = False

        else:
            # ── Fallback: construir ficha 100% desde SIGNALS ──
            precio   = float(s.get('precio_actual', 0))
            max12m   = float(s.get('max_12m') or precio)
            min12m   = float(s.get('min_12m') or precio)
            max_dt   = s.get('max_12m_date', '')
            min_dt   = s.get('min_12m_date', '')
            dist_max = round(s.get('dist_max_pct', 0), 1)

            ma20 = ma50 = precio
            ma200    = None
            ma_cross = bool(s.get('ma_cross', False))
            sup, res = [], []

            entrada = precio
            stop    = float(s.get('atr_stop') or precio * 0.94)
            target  = float(s.get('atr_target') or precio * 1.12)
            riesgo  = round((entrada-stop)/entrada*100, 1) if entrada > 0 else 0
            reward  = round((target-entrada)/entrada*100, 1) if entrada > 0 else 0
            rr      = round(min(5.0, float(s.get('rr_ratio') or (reward/riesgo if riesgo > 0 else 0))), 2)

            closes60 = []
            dates60  = []
            ma20_line = []
            ma50_line = []
            momentum  = round(s.get('momentum_21d', 0), 1)
            sin_grafico = True

        # ── Opportunity Score: 40% pred_21d + 35% R/R + 25% confianza ──
        _pred21    = s.get('pred_21d') or 0
        _conf      = s.get('pred_confidence') or 0
        _rr_val    = min(5.0, s.get('rr_ratio') or rr or 0)
        _pred_norm = min(100, max(0, (_pred21 + 15) / 30 * 100))  # -15%..+15% → 0..100
        _rr_norm   = min(100, _rr_val / 5 * 100)                   # 0..5x     → 0..100
        _conf_norm = _conf * 100                                    # 0..1      → 0..100
        opp_score  = round(_pred_norm * 0.40 + _rr_norm * 0.35 + _conf_norm * 0.25, 1)

        fichas.append({
            'ticker': ticker, 'empresa': empresa, 'market': market,
            'flag': flag, 'moneda': moneda,
            'precio': round(precio,2), 'max12m': round(max12m,2), 'min12m': round(min12m,2),
            'max_dt': max_dt, 'min_dt': min_dt, 'dist_max': dist_max,
            'ma20': round(ma20,2), 'ma50': round(ma50,2),
            'ma200': round(ma200,2) if ma200 else None,
            'ma_cross': ma_cross,
            'rsi': round(s.get('rsi',50),1),
            'momentum': momentum,
            'ret_anual': round(s.get('ret_anual',0),1),
            'soportes': sup, 'resistencias': res,
            'entrada': entrada, 'stop': stop, 'target': target,
            'riesgo': riesgo, 'reward': reward, 'rr': rr,
            'atr_stop': s.get('atr_stop'),
            'atr_target': s.get('atr_target'),
            'rr_ratio': s.get('rr_ratio'),
            'adx': s.get('adx'),
            'stress_index': s.get('stress_index'),
            'atr_percentile': s.get('atr_percentile'),
            'score_macro': round(s.get('score_macro',0),1),
            'score_tec':   round(s.get('score_tecnico',0),1),
            'score_fund':  round(s.get('score_fundamental',50),1),
            'score_final': round(s.get('score_final',0),1),
            'signal': s.get('signal_v2') or s.get('signal',''),
            'signal_v2': s.get('signal_v2', s.get('signal','')),
            'closes60': closes60, 'dates60': dates60,
            'ma20_line': ma20_line, 'ma50_line': ma50_line,
            'sin_grafico': sin_grafico,
            # ── Predicciones ARIMA/Ensemble ──
            'pred_5d':          s.get('pred_5d'),
            'pred_10d':         s.get('pred_10d'),
            'pred_21d':         s.get('pred_21d'),
            'pred_target':      s.get('pred_target'),
            'pred_confidence':  s.get('pred_confidence'),
            'pred_signal':      s.get('pred_signal', ''),
            'pred_direction_agree': s.get('pred_direction_agree', False),
            # ── Opportunity Score (40% pred_21d + 35% R/R + 25% confianza) ──
            'opportunity_score': opp_score,
        })
    # Ranking por opportunity_score descendente
    fichas.sort(key=lambda f: -f['opportunity_score'])
    return fichas
 
 
# ─────────────────────────────────────────────
# Dashboard HTML
# ─────────────────────────────────────────────
 
def generate_dashboard(
    signals: list[dict],
    index_stats: dict,
    output_path: str,
    run_date: str = "",
    price_data: dict = None,
    validacion: dict = None,
) -> str:
    """Genera el HTML del dashboard y lo escribe en output_path."""
 
    # Construir fichas de oportunidades
    fichas = []
    if price_data:
        try:
            fichas = _build_oportunidades(signals, price_data)
        except Exception as e:
            logger.warning(f"No se pudieron generar fichas de oportunidades: {e}")
 
    # Banner de validación para el header
    if validacion:
        nivel_g = validacion.get("nivel_global", "OK")
        if nivel_g == "ERROR":
            banner_bg   = "#2b0a0a"
            banner_bord = "#6b1a1a"
            banner_icon = "🔴"
            banner_txt  = "ALERTA — Datos con problemas de calidad"
            banner_col  = "#f87171"
        elif nivel_g == "WARNING":
            banner_bg   = "#2b2000"
            banner_bord = "#6b4a00"
            banner_icon = "🟡"
            banner_txt  = "Advertencia — Revisar frescura de datos"
            banner_col  = "#fbbf24"
        else:
            banner_bg   = "#0a2b0a"
            banner_bord = "#1a4a1a"
            banner_icon = "🟢"
            banner_txt  = "Datos OK"
            banner_col  = "#4ade80"
 
        mercados_html = ""
        for key, label in [("merval","MERVAL"), ("bovespa","BOVESPA"), ("sp500","S&P 500")]:
            res = validacion.get("mercados", {}).get(key, {})
            niv = res.get("nivel", "OK")
            uf  = res.get("ultima_fecha", "—")
            ic  = "🟢" if niv=="OK" else "🟡" if niv=="WARNING" else "🔴"
            mercados_html += f'<span style="margin-right:16px">{ic} <b>{label}</b> <code style="font-size:11px">{uf}</code></span>'
 
        validacion_banner = f'''<div style="background:{banner_bg};border-bottom:1px solid {banner_bord};padding:8px 32px;font-size:12px;color:{banner_col};display:flex;align-items:center;gap:16px;flex-wrap:wrap">
  <span>{banner_icon} <b>{banner_txt}</b></span>
  <span style="color:#666">|</span>
  {mercados_html}
</div>'''
    else:
        validacion_banner = ""
    
    signals_json     = json.dumps(signals,     ensure_ascii=False)
    index_stats_json = json.dumps(index_stats, ensure_ascii=False)
    fichas_json      = json.dumps(fichas,      ensure_ascii=False, default=str)
    railway_url      = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
    if railway_url and not railway_url.startswith("http"):
        railway_url = "https://" + railway_url
    # Fallback a URL conocida si no hay env var
    if not railway_url:
        railway_url = "https://inversiones-bursatiles-production.up.railway.app"
    # Portfolio data para la pestaña Portfolio
    portfolio_json = "{}"
    portfolio_alerts_json = "{}"
    try:
        import os as _os
        _pf = "data/portfolio.json"
        _pa = "data/portfolio_alerts.json"
        if _os.path.exists(_pf):
            with open(_pf) as _f:
                portfolio_json = _f.read()
        if _os.path.exists(_pa):
            with open(_pa) as _f:
                portfolio_alerts_json = _f.read()
    except Exception as e:
        logger.warning(f"Portfolio data no disponible: {e}")
 
    merval_labels  = index_stats.get("merval",  {}).get("monthly_labels", [])
    merval_values  = index_stats.get("merval",  {}).get("monthly_values", [])
    bovespa_labels = index_stats.get("bovespa", {}).get("monthly_labels", [])
    bovespa_values = index_stats.get("bovespa", {}).get("monthly_values", [])
    sp500_labels   = index_stats.get("sp500",   {}).get("monthly_labels", [])
    sp500_values   = index_stats.get("sp500",   {}).get("monthly_values", [])
 
    m_ret  = index_stats.get("merval",  {}).get("ret_anual",  0)
    b_ret  = index_stats.get("bovespa", {}).get("ret_anual",  0)
    s_ret  = index_stats.get("sp500",   {}).get("ret_anual",  0)
    m_act  = index_stats.get("merval",  {}).get("actual",     0)
    b_act  = index_stats.get("bovespa", {}).get("actual",     0)
    s_act  = index_stats.get("sp500",   {}).get("actual",     0)
    m_vol  = index_stats.get("merval",  {}).get("volatilidad", 0)
    b_vol  = index_stats.get("bovespa", {}).get("volatilidad", 0)
    s_vol  = index_stats.get("sp500",   {}).get("volatilidad", 0)
    m_day  = index_stats.get("merval",  {}).get("ret_dia", None)
    b_day  = index_stats.get("bovespa", {}).get("ret_dia", None)
    s_day  = index_stats.get("sp500",   {}).get("ret_dia", None)

    # ── Cross-market (Fase 1/4) ──────────────────────────────────────────
    _cm         = index_stats.get("cross_market", {})
    cm_regime   = _cm.get("regime", "NEUTRAL")
    cm_sp_trend = _cm.get("sp500_trend", "—")
    cm_sp_score = _cm.get("sp500_trend_score", 50)
    cm_adj_mv   = _cm.get("score_adjustments", {}).get("MERVAL", 0)
    cm_adj_bv   = _cm.get("score_adjustments", {}).get("BOVESPA", 0)
    cm_narrative= _cm.get("narrative", "Sin datos de contexto cross-market aún.")
    cm_corr_sp_mv = _cm.get("correlations", {}).get("merval_sp500", 0)
    cm_corr_sp_bv = _cm.get("correlations", {}).get("bovespa_sp500", 0)
    cm_regime_color = {"RISK_ON": "#4ade80", "RISK_OFF": "#f87171"}.get(cm_regime, "#fbbf24")

    # ── Health / Backtest (Fase 3) ────────────────────────────────────────
    import json as _json, os as _os
    _health = {}
    _backtest = {}
    try:
        if _os.path.exists("data/health_metrics.json"):
            with open("data/health_metrics.json") as _hf:
                _health = _json.load(_hf)
        if _os.path.exists("data/backtest_results.json"):
            with open("data/backtest_results.json") as _bf:
                _backtest = _json.load(_bf)
    except Exception:
        pass
    hl_sla       = _health.get("sla_status", "UNKNOWN")
    hl_runs      = _health.get("pipeline_runs_today", 0)
    hl_dur       = _health.get("duration_last_sec", 0)
    hl_buy       = _health.get("buy_signals", 0)
    hl_sla_color = {"OK": "#4ade80", "WARNING": "#fbbf24", "CRITICAL": "#f87171"}.get(hl_sla, "#888")
    bt_ev        = None
    bt_acc       = None
    bt_trades    = _backtest.get("total_trades", 0)
    for _row in _backtest.get("signal_summary", []):
        if "COMPRA" in _row.get("signal", "") and _row.get("expected_value") is not None:
            bt_ev  = _row["expected_value"]
            break
    _pred = _backtest.get("predictor", {})
    bt_acc = _pred.get("directional_accuracy")
    bt_ev_str  = f"{bt_ev:+.1f}%" if bt_ev is not None else "—"
    bt_acc_str = f"{bt_acc:.0%}"  if bt_acc is not None else "—"
 
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Inversiones Bursátiles — {run_date}</title>
<style>
  *{{margin:0;padding:0;box-sizing:border-box}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d0d0f;color:#e8e8ea;font-size:14px}}
  .header{{background:#111115;border-bottom:1px solid #222230;padding:20px 32px;display:flex;justify-content:space-between;align-items:center}}
  .header h1{{font-size:20px;font-weight:600;color:#fff}}
  .badge{{display:inline-block;background:#1e3a5f;color:#5ba3ff;padding:3px 10px;border-radius:10px;font-size:11px;font-weight:600;margin-bottom:4px}}
  .tabs{{display:flex;background:#111115;border-bottom:1px solid #222230;padding:0 32px;gap:4px;position:sticky;top:0;z-index:100;overflow-x:auto}}
  .tab{{padding:12px 18px;cursor:pointer;font-size:13px;color:#888;border-bottom:2px solid transparent;white-space:nowrap}}
  .tab.on{{color:#5ba3ff;border-bottom-color:#5ba3ff;font-weight:500}}
  .page{{display:none;padding:28px 32px;max-width:1200px}}
  .page.on{{display:block}}
  .grid-3{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:24px}}
  .grid-4{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px}}
  .card{{background:#16161e;border:1px solid #222230;border-radius:10px;padding:16px}}
  .card-title{{font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.7px;margin-bottom:6px}}
  .card-value{{font-size:24px;font-weight:700;color:#fff;line-height:1}}
  .card-sub{{font-size:11px;color:#666;margin-top:5px}}
  .pos{{color:#4ade80}}.neg{{color:#f87171}}
  .section-title{{font-size:16px;font-weight:600;color:#fff;margin-bottom:16px;padding-bottom:8px;border-bottom:1px solid #222230}}
  /* ── Glosario tooltips ───────────────────────────────────────────── */
  [data-tip]{{position:relative;cursor:help;border-bottom:1px dotted #555}}
  [data-tip]::after{{content:attr(data-tip);position:absolute;bottom:calc(100% + 6px);left:50%;transform:translateX(-50%);background:#1a1a2e;color:#e8e8ea;font-size:11px;font-weight:400;padding:6px 10px;border-radius:6px;white-space:normal;width:220px;text-align:left;pointer-events:none;opacity:0;transition:opacity .2s;z-index:999;border:1px solid #333;line-height:1.4}}
  [data-tip]:hover::after{{opacity:1}}
  th[data-tip]{{border-bottom:1px dotted #4a4a6a!important;cursor:help}}
  /* ── Chart containers nuevos ─────────────────────────────────────── */
  .chart-scatter{{height:340px;background:#0a0a0f;border-radius:10px;padding:4px;margin-bottom:24px}}
  .chart-heatmap-wrap{{margin-bottom:24px;overflow-x:auto}}
  .hm-grid{{display:grid;gap:3px;margin-top:8px}}
  .hm-cell{{display:flex;align-items:center;justify-content:center;border-radius:4px;font-size:10px;font-weight:700;cursor:default;transition:transform .15s}}
  .hm-cell:hover{{transform:scale(1.08);z-index:2;position:relative}}
  .tbl{{width:100%;border-collapse:collapse;margin-bottom:24px}}
  .tbl th{{text-align:left;padding:9px 12px;font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.4px;border-bottom:1px solid #222230}}
  .tbl td{{padding:10px 12px;border-bottom:1px solid #1a1a22;font-size:13px}}
  .tbl tr:hover td{{background:#1a1a24}}
  .ticker{{font-weight:700;color:#5ba3ff;font-family:monospace;font-size:12px}}
  .chart-wrap{{position:relative;width:100%;height:260px;margin-bottom:24px}}
  .sig-buy{{color:#4ade80}}.sig-neu{{color:#fbbf24}}.sig-sell{{color:#fb923c}}
  .pano-header{{display:flex;flex-direction:column;gap:12px;margin-bottom:24px}}
  .pano-card{{width:100%;background:#16161e;border:1px solid #222230;border-radius:12px;padding:20px 24px;display:grid;grid-template-columns:auto 1fr auto;align-items:center;gap:0 20px}}
  .pano-flag{{font-size:36px;grid-row:1/4}}
  .pano-label{{font-size:11px;color:#555;text-transform:uppercase;letter-spacing:1.5px;font-weight:600}}
  .pano-value{{font-size:clamp(28px,7vw,48px);font-weight:900;color:#fff;line-height:1;letter-spacing:-1px}}
  .pano-anual{{font-size:clamp(24px,6vw,40px);font-weight:900;line-height:1}}
  .pano-day-row{{display:flex;align-items:center;gap:6px;grid-column:2}}
  .pano-day-label{{font-size:10px;color:#444;text-transform:uppercase;letter-spacing:.5px}}
  .pano-day-value{{font-size:clamp(16px,4vw,22px);font-weight:700}}
  .pano-day-dot{{width:8px;height:8px;border-radius:50%;animation:pulse 2s infinite}}
  @keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.3}}}}
  .pano-vol{{font-size:12px;color:#555;grid-column:3;text-align:right;align-self:center;display:flex;flex-direction:column;align-items:flex-end;gap:2px}}
  .pano-vol-label{{font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.8px}}
  .pano-vol-value{{font-size:clamp(16px,3vw,22px);font-weight:700;color:#888}}
  .pano-vol-desc{{font-size:9px;color:#333;max-width:80px;text-align:right;line-height:1.3}}
  .chart-legend{{display:flex;gap:20px;margin-bottom:10px;flex-wrap:wrap}}
  .legend-item{{display:flex;align-items:center;gap:6px;font-size:12px;color:#aaa}}
  .radar-card{{background:#0d1a0d;border:1px solid #1a3320;border-radius:10px;padding:18px;margin-bottom:14px;display:flex;align-items:center;gap:16px;transition:background .2s}}
  .radar-card:hover{{background:#112211}}
  .radar-rank{{font-size:28px;font-weight:900;color:#1e3a22;min-width:44px;text-align:center;line-height:1}}
  .radar-info{{flex:1}}
  .radar-ticker{{font-size:15px;font-weight:700;color:#5ba3ff;font-family:monospace}}
  .radar-metrics{{display:flex;gap:14px;flex-wrap:wrap;font-size:12px;color:#aaa}}
  .radar-score-wrap{{text-align:right;min-width:80px}}
  .radar-score{{font-size:26px;font-weight:900;line-height:1}}
  .radar-score-label{{font-size:10px;color:#555;text-transform:uppercase;letter-spacing:.5px}}
  .radar-bar-wrap{{width:100%;height:4px;background:#1a2a1a;border-radius:2px;margin-top:8px}}
  .radar-bar{{height:4px;border-radius:2px}}
  .radar-signals{{display:flex;gap:8px;margin-top:6px;flex-wrap:wrap}}
  .radar-tag{{font-size:10px;padding:2px 8px;border-radius:10px;font-weight:600}}
  .tag-green{{background:#0d2b1a;color:#4ade80;border:1px solid #1a4a2a}}
  .tag-yellow{{background:#2b2100;color:#fbbf24;border:1px solid #4a3500}}
  .tag-blue{{background:#0d1e3a;color:#5ba3ff;border:1px solid #1a3560}}
  .tag-orange{{background:#2b1500;color:#fb923c;border:1px solid #4a2800}}
  .radar-criteria{{background:#111115;border:1px solid #1a1a2e;border-radius:8px;padding:12px 16px;margin-bottom:20px;font-size:12px;color:#666;display:flex;gap:20px;flex-wrap:wrap}}
 
  /* ── OPORTUNIDADES DE COMPRA ── */
  .op-tabs{{display:flex;gap:4px;margin-bottom:16px;overflow-x:auto;padding-bottom:4px}}
  .op-tab{{padding:7px 14px;cursor:pointer;font-size:11px;color:#888;border:1px solid #222230;border-radius:6px;white-space:nowrap;background:#16161e;font-family:monospace}}
  .op-tab.on{{color:#5ba3ff;border-color:#5ba3ff;background:#0d1e3a;font-weight:600}}
  .op-rank-row{{background:#16161e;border:1px solid #222230;border-radius:8px;padding:12px 16px;margin-bottom:8px;display:flex;align-items:center;gap:12px;cursor:pointer;transition:border-color .15s;flex-wrap:wrap}}
  .op-rank-row:hover{{border-color:#5ba3ff}}
  .op-num{{font-size:18px;font-weight:900;color:#2a2a3a;min-width:28px;font-family:monospace}}
  .op-num.gold{{color:#d29922}}
  .op-main{{flex:1;min-width:140px}}
  .op-ticker{{font-size:14px;font-weight:700;color:#5ba3ff;font-family:monospace}}
  .op-emp{{font-size:11px;color:#666;margin-top:1px}}
  .op-sbar{{width:100%;height:3px;background:#1a1a2e;border-radius:2px;margin-top:5px}}
  .op-sbarf{{height:3px;border-radius:2px}}
  .op-mets{{display:flex;gap:14px;flex-wrap:wrap;align-items:center}}
  .op-m{{display:flex;flex-direction:column;align-items:flex-end;min-width:44px}}
  .op-mv{{font-size:12px;font-weight:600;font-family:monospace}}
  .op-ml{{font-size:9px;color:#555;text-transform:uppercase}}
  .op-sig{{font-size:10px;font-weight:700;padding:3px 8px;border-radius:3px;white-space:nowrap}}
  .op-sig-c{{background:#0d2b1a;color:#4ade80;border:1px solid #1a4a2a}}
  .op-sig-f{{background:#2a1f00;color:#d29922;border:1px solid #4a3500}}
  .op-ficha-hdr{{display:flex;align-items:flex-start;gap:14px;margin-bottom:14px;flex-wrap:wrap}}
  .op-ftick{{font-size:26px;font-weight:900;color:#5ba3ff;font-family:monospace}}
  .op-femp{{font-size:13px;color:#888;margin-top:2px}}
  .op-fprice{{font-size:22px;font-weight:700;margin-left:auto;text-align:right}}
  .op-card{{background:#16161e;border:1px solid #222230;border-radius:8px;padding:12px;margin-bottom:12px}}
  .op-card h3{{font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.6px;margin-bottom:8px}}
  .op-chart-wrap{{position:relative;width:100%;height:260px;margin-bottom:12px}}
  .op-rrbox{{background:#111115;border-radius:8px;padding:12px;display:grid;grid-template-columns:repeat(3,1fr);gap:8px;text-align:center;margin-bottom:12px;border:1px solid #1a1a2e}}
  .op-rrval{{font-size:18px;font-weight:900;font-family:monospace}}
  .op-rrlbl{{font-size:9px;color:#666;text-transform:uppercase;letter-spacing:.4px;margin-top:2px}}
  .op-fgrid{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
  .op-lvl{{display:flex;justify-content:space-between;align-items:center;padding:5px 0;border-bottom:1px solid #1a1a22;font-family:monospace;font-size:12px}}
  .op-ltag{{font-size:9px;padding:2px 5px;border-radius:3px;font-weight:700;margin-right:4px}}
  .op-lt-r{{background:#1f0d0d;color:#f87171}}
  .op-lt-s{{background:#0d1f0d;color:#4ade80}}
  .op-lt-e{{background:#0d1e3a;color:#5ba3ff}}
  .op-lt-st{{background:#2b1500;color:#fb923c}}
  .op-lt-tg{{background:#1a0d2b;color:#bc8cff}}
  .op-sc-row{{display:flex;justify-content:space-between;align-items:center;padding:6px 0;border-bottom:1px solid #1a1a22}}
  .op-sc-name{{font-size:11px;color:#666}}
  .op-sc-val{{font-size:14px;font-weight:700;font-family:monospace}}
  .op-sc-bar{{width:60px;height:3px;background:#1a1a2e;border-radius:2px;margin-top:2px}}
  .op-sc-fill{{height:3px;border-radius:2px}}
  .op-techbox{{background:#111115;border-radius:6px;padding:8px 10px;margin-top:8px;display:flex;gap:12px;flex-wrap:wrap;font-family:monospace;font-size:11px}}
  @media(max-width:768px){{
    .grid-3,.grid-4{{grid-template-columns:repeat(2,1fr)}}
    .page{{padding:16px}}
    .op-fgrid{{grid-template-columns:1fr}}
    .radar-card{{flex-direction:column;align-items:flex-start}}
  }}
/* ── FORMULARIO OPERACIONES ── */
#op-form-grid{{display:grid;grid-template-columns:2fr 1fr 1.2fr 1.2fr auto;gap:12px;align-items:end}}
@media(max-width:900px){{#op-form-grid{{grid-template-columns:1fr 1fr}}}}
@media(max-width:600px){{#op-form-grid{{grid-template-columns:1fr}}}}
.op-ticker-item{{padding:8px 12px;cursor:pointer;font-size:12px;font-family:monospace;border-bottom:1px solid #222230;display:flex;justify-content:space-between;align-items:center}}
.op-ticker-item:hover{{background:#1e2a3a;color:#5ba3ff}}
.op-ticker-item .oti-sig{{font-size:10px;padding:2px 6px;border-radius:4px}}
/* ── CONCLUSIONES EXPANDIBLES ── */
.concl-subtitle{{font-size:12px;color:#666;margin-bottom:14px;letter-spacing:0.3px}}
.concl-card-exp{{border-radius:10px;padding:16px;margin-bottom:12px;cursor:pointer;transition:all .2s}}
.concl-card-exp:hover{{filter:brightness(1.08)}}
.concl-card-exp.buy{{background:#0d2b1a;border:1px solid #1a3a1a}}
.concl-card-exp.radar{{background:#0d1f2b;border:1px solid #1a2a3a}}
.concl-card-exp.sell{{background:#1f0d0d;border:1px solid #3a1a1a}}
.concl-header{{display:flex;align-items:center;gap:10px}}
.concl-rank{{font-family:'DM Mono',monospace;font-size:22px;font-weight:800;color:#555;min-width:36px}}
.concl-main{{flex:1}}
.concl-title-row{{display:flex;align-items:center;gap:8px;flex-wrap:wrap}}
.concl-arrow{{font-size:18px;color:#666;transition:transform .25s;margin-left:auto}}
.concl-card-exp.open .concl-arrow{{transform:rotate(180deg)}}
.concl-metrics-row{{display:flex;gap:16px;flex-wrap:wrap;margin-top:4px;font-size:12px;color:#888}}
.concl-metrics-row b{{color:#ccc}}
.concl-detail{{max-height:0;overflow:hidden;transition:max-height .35s ease;margin-top:0}}
.concl-card-exp.open .concl-detail{{max-height:600px;margin-top:12px}}
.concl-detail-inner{{border-top:1px solid rgba(255,255,255,0.06);padding-top:12px;display:grid;grid-template-columns:1fr 1fr;gap:10px 20px;font-size:12px;color:#999}}
.concl-detail-inner .dl{{display:flex;justify-content:space-between;padding:3px 0}}
.concl-detail-inner .dl b{{color:#ddd}}
</style>
</head>
<body>
<div class="header">
  <div>
    <div class="badge">INVERSIONES BURSÁTILES</div>
    <h1>Informe de Inversiones</h1>
    <div style="font-size:12px;color:#666;margin-top:3px">MERVAL · BOVESPA · S&P 500 · Generado {run_date}</div>
  </div>
  <div style="text-align:right;font-size:12px;color:#666">Pipeline automático<br>Modelo v2.0 — Asset Quality(50%Macro+30%Fundamental+20%Sectorial) · Entry Score(60%Técnico+25%Riesgo/Retorno+15%Dist.Máximo)</div>
</div>
<!-- banner -->
<div class="tabs">
  <div class="tab on" onclick="sw('panorama',this)">Panorama</div>
  <div class="tab"    onclick="sw('merval',this)">MERVAL</div>
  <div class="tab"    onclick="sw('bovespa',this)">BOVESPA</div>
  <div class="tab"    onclick="sw('sp500',this)">S&amp;P 500</div>
  <div class="tab"    onclick="sw('oportunidades',this)">🎯 Oportunidades</div>
  <div class="tab"    onclick="sw('portfolio',this)">💼 Portfolio</div>
</div>
 
<!-- PANORAMA -->
<div id="panorama" class="page on">
  <div class="pano-header">
    <div class="pano-card">
      <div class="pano-flag">🇦🇷</div>
      <div style="display:flex;flex-direction:column;gap:4px">
        <div class="pano-label">MERVAL</div>
        <div class="pano-value" id="pano-m-val">{m_act:,.0f}</div>
        <div class="pano-anual" id="pano-m-anual" style="color:{'#4ade80' if m_ret>=0 else '#f87171'}">{'+ ' if m_ret>=0 else ''}{m_ret:.2f}%</div>
        <div class="pano-day-row">
          <div class="pano-day-dot" id="dot-m" style="background:#4ade80"></div>
          <span class="pano-day-label" title="Variación respecto al cierre anterior">ÚLTIMO CIERRE</span>
          <span class="pano-day-value" id="pano-m-day">{'—' if m_day is None else ('+' if m_day>=0 else '')+f'{m_day:.2f}%'}</span>
        </div>
        <div id="pano-m-fecha" style="font-size:9px;color:#333;margin-top:2px"></div>
      </div>
      <div class="pano-vol">
        <span class="pano-vol-label">Volatilidad</span>
        <span class="pano-vol-value">{m_vol:.1f}%</span>
        <span class="pano-vol-desc">anualizada<br>últimos 12m</span>
      </div>
    </div>
    <div class="pano-card">
      <div class="pano-flag">🇧🇷</div>
      <div style="display:flex;flex-direction:column;gap:4px">
        <div class="pano-label">BOVESPA</div>
        <div class="pano-value" id="pano-b-val">{b_act:,.0f}</div>
        <div class="pano-anual" id="pano-b-anual" style="color:{'#4ade80' if b_ret>=0 else '#f87171'}">{'+ ' if b_ret>=0 else ''}{b_ret:.2f}%</div>
        <div class="pano-day-row">
          <div class="pano-day-dot" id="dot-b" style="background:#4ade80"></div>
          <span class="pano-day-label" title="Variación respecto al cierre anterior">ÚLTIMO CIERRE</span>
          <span class="pano-day-value" id="pano-b-day">{'—' if b_day is None else ('+' if b_day>=0 else '')+f'{b_day:.2f}%'}</span>
        </div>
        <div id="pano-b-fecha" style="font-size:9px;color:#333;margin-top:2px"></div>
      </div>
      <div class="pano-vol">
        <span class="pano-vol-label">Volatilidad</span>
        <span class="pano-vol-value">{b_vol:.1f}%</span>
        <span class="pano-vol-desc">anualizada<br>últimos 12m</span>
      </div>
    </div>
    <div class="pano-card">
      <div class="pano-flag">🇺🇸</div>
      <div style="display:flex;flex-direction:column;gap:4px">
        <div class="pano-label">S&amp;P 500</div>
        <div class="pano-value" id="pano-s-val">{s_act:,.0f}</div>
        <div class="pano-anual" id="pano-s-anual" style="color:{'#4ade80' if s_ret>=0 else '#f87171'}">{'+ ' if s_ret>=0 else ''}{s_ret:.2f}%</div>
        <div class="pano-day-row">
          <div class="pano-day-dot" id="dot-s" style="background:#4ade80"></div>
          <span class="pano-day-label" title="Variación respecto al cierre anterior">ÚLTIMO CIERRE</span>
          <span class="pano-day-value" id="pano-s-day">{'—' if s_day is None else ('+' if s_day>=0 else '')+f'{s_day:.2f}%'}</span>
        </div>
        <div id="pano-s-fecha" style="font-size:9px;color:#333;margin-top:2px"></div>
      </div>
      <div class="pano-vol">
        <span class="pano-vol-label">Volatilidad</span>
        <span class="pano-vol-value">{s_vol:.1f}%</span>
        <span class="pano-vol-desc">anualizada<br>últimos 12m</span>
      </div>
    </div>
  </div>
  <!-- ── Health & Backtest metrics ── -->
  <div style="display:flex;gap:10px;flex-wrap:wrap;margin:12px 0 18px;align-items:center">
    <div style="background:#0d0d14;border:1px solid #222;border-radius:8px;padding:8px 14px;display:flex;gap:8px;align-items:center">
      <span style="font-size:11px;color:#666;font-weight:600;text-transform:uppercase;letter-spacing:.5px">Sistema</span>
      <span style="font-size:12px;font-weight:700;color:{hl_sla_color};background:{hl_sla_color}22;padding:2px 8px;border-radius:4px">{hl_sla}</span>
      <span style="font-size:11px;color:#555">|</span>
      <span style="font-size:11px;color:#666">Runs hoy: <b style="color:#aaa">{hl_runs}</b></span>
      <span style="font-size:11px;color:#555">|</span>
      <span style="font-size:11px;color:#666">Dur: <b style="color:#aaa">{hl_dur:.0f}s</b></span>
      <span style="font-size:11px;color:#555">|</span>
      <span style="font-size:11px;color:#666">Compras: <b style="color:#4ade80">{hl_buy}</b></span>
    </div>
    <div style="background:#0d0d14;border:1px solid #222;border-radius:8px;padding:8px 14px;display:flex;gap:8px;align-items:center">
      <span style="font-size:11px;color:#666;font-weight:600;text-transform:uppercase;letter-spacing:.5px">Backtest</span>
      <span style="font-size:11px;color:#666">EV compras 21d: <b style="color:{'#4ade80' if bt_ev and bt_ev>0 else '#f87171' if bt_ev and bt_ev<0 else '#888'}">{bt_ev_str}</b></span>
      <span style="font-size:11px;color:#555">|</span>
      <span style="font-size:11px;color:#666">Predictor acc: <b style="color:#a78bfa">{bt_acc_str}</b></span>
      <span style="font-size:11px;color:#555">|</span>
      <span style="font-size:11px;color:#666">n={bt_trades}</span>
    </div>
  </div>

  <!-- ── Cross-Market Context ── -->
  <div style="background:#0d0d14;border:1px solid #1a1a2e;border-radius:10px;padding:14px 18px;margin:0 0 18px">
    <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:10px">
      <span style="font-size:12px;font-weight:700;color:#666;text-transform:uppercase;letter-spacing:1px">Cross-Market</span>
      <span style="font-size:13px;font-weight:800;color:{cm_regime_color};background:{cm_regime_color}22;padding:3px 10px;border-radius:5px">{cm_regime}</span>
      <span style="font-size:12px;color:#666">SP500: <b style="color:{'#4ade80' if cm_sp_trend=='ALCISTA' else '#f87171' if cm_sp_trend=='BAJISTA' else '#fbbf24'}">{cm_sp_trend}</b> ({cm_sp_score:.0f}/100)</span>
      <span style="font-size:12px;color:#666">Corr MV↔SP: <b style="color:#a78bfa">{cm_corr_sp_mv:.2f}</b></span>
      <span style="font-size:12px;color:#666">Corr BV↔SP: <b style="color:#a78bfa">{cm_corr_sp_bv:.2f}</b></span>
      <span style="font-size:12px;color:#666">Adj MERVAL: <b style="color:{'#4ade80' if cm_adj_mv>0 else '#f87171' if cm_adj_mv<0 else '#888'}">{cm_adj_mv:+.1f}pt</b></span>
      <span style="font-size:12px;color:#666">Adj BOVESPA: <b style="color:{'#4ade80' if cm_adj_bv>0 else '#f87171' if cm_adj_bv<0 else '#888'}">{cm_adj_bv:+.1f}pt</b></span>
    </div>
    <div style="font-size:12px;color:#777;line-height:1.6;font-style:italic">{cm_narrative}</div>
  </div>

  <div class="section-title">Evolución comparativa — base 100</div>
  <div class="chart-legend">
    <div class="legend-item"><div style="background:#5ba3ff;height:3px;width:24px;border-radius:2px"></div><span style="color:#5ba3ff;font-weight:600">MERVAL 🇦🇷</span></div>
    <div class="legend-item"><div style="background:#4ade80;height:3px;width:24px;border-radius:2px"></div><span style="color:#4ade80;font-weight:600">BOVESPA 🇧🇷</span></div>
    <div class="legend-item"><div style="background:#fbbf24;height:3px;width:24px;border-radius:2px"></div><span style="color:#fbbf24;font-weight:600">S&amp;P 500 🇺🇸</span></div>
  </div>
  <div class="chart-wrap"><canvas id="chartPano"></canvas></div>
  <div class="section-title">Señales por mercado — Ranking global</div>
  <div style="margin-bottom:6px;font-size:11px;color:#555">Ordenado por Score V2 · Señal descendente</div>
  <div style="margin:14px 0 4px;font-size:13px;font-weight:700;color:#5ba3ff;border-bottom:1px solid #5ba3ff;padding-bottom:4px">🇦🇷 MERVAL</div>
  <table class="tbl" id="tbl-pano-merval"></table>
  <div style="margin:18px 0 4px;font-size:13px;font-weight:700;color:#4ade80;border-bottom:1px solid #4ade80;padding-bottom:4px">🇧🇷 BOVESPA</div>
  <table class="tbl" id="tbl-pano-bovespa"></table>
  <div style="margin:18px 0 4px;font-size:13px;font-weight:700;color:#fbbf24;border-bottom:1px solid #fbbf24;padding-bottom:4px">🇺🇸 S&amp;P 500</div>
  <table class="tbl" id="tbl-pano-sp500"></table>
</div>
 
<!-- MERVAL -->
<div id="merval" class="page">
  <div class="section-title">MERVAL — Estadísticas 12 meses</div>
  <div class="grid-4" id="merval-stats"></div>
  <div class="chart-wrap"><canvas id="chartMerval"></canvas></div>
  <div class="section-title">Señales del modelo</div>
  <table class="tbl" id="tbl-merval"></table>
</div>
 
<!-- BOVESPA -->
<div id="bovespa" class="page">
  <div class="section-title">BOVESPA — Estadísticas 12 meses</div>
  <div class="grid-4" id="bovespa-stats"></div>
  <div class="chart-wrap"><canvas id="chartBovespa"></canvas></div>
  <div class="section-title">Señales del modelo</div>
  <table class="tbl" id="tbl-bovespa"></table>
</div>
 
<!-- SP500 -->
<div id="sp500" class="page">
  <div class="section-title">S&amp;P 500 — Estadísticas 12 meses</div>
  <div class="grid-4" id="sp500-stats"></div>
  <div class="chart-wrap"><canvas id="chartSP500"></canvas></div>
  <div class="section-title">Señales del modelo</div>
  <table class="tbl" id="tbl-sp500"></table>
</div>
 
<!-- OPORTUNIDADES DE COMPRA -->
<div id="oportunidades" class="page">
  <div class="section-title" style="color:#4ade80">🎯 Oportunidades — Ranking por Convicción</div>
  <div style="font-size:12px;color:#666;margin-bottom:16px;background:#111115;border:1px solid #1a1a2e;border-radius:8px;padding:10px 14px;display:flex;gap:20px;flex-wrap:wrap">
    <span>📈 <b style="color:#4ade80">40% Predicción 21d</b> · retorno esperado ensemble</span>
    <span>⚖️ <b style="color:#fbbf24">35% R/R Ratio</b> · recompensa por unidad de riesgo</span>
    <span>🎯 <b style="color:#a78bfa">25% Confianza</b> · certeza del predictor</span>
    <span style="border-left:1px solid #333;padding-left:16px">Filtro: señal COMPRA · score ≥ 50 · top 25% universo</span>
  </div>
  <div id="op-rank-page">
    <div style="font-size:13px;font-weight:600;color:#aaa;margin-bottom:12px;padding-bottom:7px;border-bottom:1px solid #222230">
      Oportunidades activas — ordenadas por Opportunity Score descendente
    </div>
    <div style="display:flex;align-items:center;gap:12px;padding:4px 16px 8px;flex-wrap:wrap">
      <div style="flex:1;min-width:140px"></div>
      <div style="display:flex;gap:14px;flex-wrap:wrap;align-items:center;padding-right:4px">
        <div style="min-width:44px;text-align:right;font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.5px">Precio</div>
        <div style="min-width:44px;text-align:right;font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.5px">12 meses</div>
        <div style="min-width:44px;text-align:right;font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.5px">RSI</div>
        <div style="min-width:44px;text-align:right;font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.5px">vs Máx</div>
        <div style="min-width:44px;text-align:right;font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.5px">Score</div>
        <div style="min-width:44px;text-align:right;font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.5px">R/R</div>
        <div style="min-width:70px;text-align:right;font-size:9px;color:#444;text-transform:uppercase;letter-spacing:.5px">Señal</div>
      </div>
    </div>
    <div id="op-rg"></div>
  </div>
  <div id="op-ficha-page" style="display:none">
    <button onclick="showOpRank()" style="background:#16161e;border:1px solid #222230;color:#5ba3ff;padding:6px 14px;border-radius:6px;cursor:pointer;font-size:12px;margin-bottom:16px">← Volver al ranking</button>
    <div id="op-fi"></div>
  </div>
</div>
 
<!-- CONCLUSIONES -->
<div id="conclusiones" class="page">

  <!-- ── 0) PRÓXIMO DÓLAR — PRIMERO ── -->
  <div class="section-title" style="color:#fbbf24;margin-bottom:6px">💵 ¿Dónde poner el próximo dólar?</div>
  <div class="concl-subtitle" style="margin-bottom:16px">Asignación óptima de capital nuevo · Criterio: mayor retorno esperado (Score V2 × R/R) · Primero posiciones existentes con señal positiva, luego nuevas</div>
  <div id="capital-block"></div>

  <!-- ── A) SCATTER Score V2 vs Predicción 21d ── -->
  <div class="section-title" style="color:#a78bfa;margin-bottom:4px">📊 Mapa de Señales — Score V2 vs Predicción 21d</div>
  <div style="font-size:12px;color:#555;margin-bottom:10px">Cuadrante superior derecho = consenso modelo + predictor. Cada punto es un ticker. Hover para detalles.</div>
  <div class="chart-scatter"><canvas id="chartScatter"></canvas></div>

  <!-- ── B) HEATMAP sectores × mercados ── -->
  <div class="section-title" style="color:#5ba3ff;margin-bottom:4px">🌡️ Heatmap — Score Promedio por Sector × Mercado</div>
  <div style="font-size:12px;color:#555;margin-bottom:10px">Verde intenso = sector fuerte en ese mercado. Rojo = sector débil. Sirve para detectar rotación sectorial.</div>
  <div class="chart-heatmap-wrap" id="heatmap-container"></div>

  <!-- 1) COMPRAS CONFIRMADAS — PRIMERO -->
  <div class="section-title" style="color:#4ade80;margin-bottom:6px;margin-top:28px">✅ Oportunidades de Compra Confirmadas</div>
  <div class="concl-subtitle">Señales activas del modelo macro × técnico × sectorial × fundamental · Ordenadas por ranking accionable</div>
  <div id="compras-block"></div>
 
  <!-- 2) OPORTUNIDADES RANKEADAS — SEGUNDO -->
  <div class="section-title" style="color:#86efac;margin-top:32px;margin-bottom:6px">🎯 Oportunidades Rankeadas por Convicción</div>
  <div class="concl-subtitle">Solo las que cumplen los tres criterios simultáneamente · Click en cualquier card para ir directamente a la ficha completa</div>
  <div class="radar-criteria">
    <span>📈 <b style="color:#4ade80">40% Predicción 21d</b> — retorno esperado ensemble</span>
    <span>⚖️ <b style="color:#fbbf24">35% R/R Ratio</b> — recompensa por unidad de riesgo</span>
    <span>🎯 <b style="color:#a78bfa">25% Confianza</b> — certeza del predictor</span>
    <span style="color:#555">Filtro: señal COMPRA · score ≥ 50 · top 25% universo</span>
  </div>
  <div id="radar-block"></div>
 
  <!-- 3) SEÑALES DE REDUCCIÓN -->
  <div class="section-title" style="color:#fb923c;margin-top:32px;margin-bottom:6px">🔴 Señales de Reducción de Portfolio</div>
  <div class="concl-subtitle">Activos con señal de venta o venta parcial · Ordenados por urgencia de reducción</div>
  <div id="ventas-block"></div>


</div>
 
<!-- PORTFOLIO -->
<div id="portfolio" class="page">

  <!-- ── Header con refresh real ── -->
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px">
    <div class="section-title" style="margin-bottom:0;border-bottom:none">💼 Mi Portfolio — Posiciones Activas</div>
    <div style="display:flex;align-items:center;gap:10px">
      <div id="port-refresh-indicator" style="display:none;width:8px;height:8px;border-radius:50%;background:#fbbf24;animation:pulse 1s infinite"></div>
      <span id="port-update-ts" style="font-size:11px;color:#555">Cargando precios...</span>
      <button id="btn-refresh-port" onclick="portRefreshNow()"
        style="background:#0d1f3c;border:1px solid #1e3a5f;border-radius:6px;padding:5px 14px;color:#5ba3ff;font-size:12px;cursor:pointer;font-family:inherit;display:flex;align-items:center;gap:5px;transition:background .2s">
        <span id="refresh-icon">↻</span> Actualizar
      </button>
      <span style="font-size:10px;color:#333" id="port-countdown">Auto en 60s</span>
    </div>
  </div>

  <!-- ── KPI cards ── -->
  <div id="portfolio-summary" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:10px;margin-bottom:18px"></div>



  <!-- ── Tabla de posiciones ── -->
  <div class="section-title" style="color:#e2e8f0;margin-top:18px;margin-bottom:6px;font-size:13px">📋 Detalle de Posiciones</div>
  <div style="overflow-x:auto">
    <table class="tbl" id="portfolio-table" style="min-width:750px"></table>
  </div>

  <!-- ── Mini chart de evolución P&L ── -->
  <div class="section-title" style="color:#5ba3ff;margin-top:18px;margin-bottom:4px;font-size:13px">📈 Distribución P&amp;L por Posición</div>
  <div style="position:relative;height:180px;margin-bottom:16px"><canvas id="chartPortBar"></canvas></div>

  <!-- ── Alertas ── -->
  <div class="section-title" style="margin-top:8px;color:#fbbf24;font-size:13px">⚠️ Alertas Activas</div>
  <div id="portfolio-alerts" style="font-size:13px;margin-bottom:24px"></div>

  <!-- ── FORMULARIO DE OPERACIONES ── -->
  <div class="section-title" style="margin-top:8px;color:#5ba3ff">📝 Registrar Operación</div>
  <div style="background:#16161e;border:1px solid #222230;border-radius:10px;padding:20px;margin-bottom:20px">
    <!-- Fila 1: Ticker + Operación + Instrumento -->
    <div style="display:grid;grid-template-columns:1.4fr 0.8fr 1.2fr;gap:12px;align-items:end;margin-bottom:12px">
      <!-- Ticker con autocomplete -->
      <div>
        <label style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.6px;display:block;margin-bottom:4px">Ticker</label>
        <div style="position:relative">
          <input id="op-ticker-input" type="text" placeholder="Ej: CEPU.BA, MELI, HAPV3.SA…" autocomplete="off"
            style="width:100%;background:#0d0d0f;border:1px solid #333;border-radius:6px;padding:8px 10px;color:#e8e8ea;font-size:13px;font-family:monospace"
            oninput="tickerAutocomplete(this.value)" onblur="setTimeout(function(){{document.getElementById('op-ticker-drop').style.display='none';}},400)">
          <div id="op-ticker-drop" style="display:none;position:absolute;top:100%;left:0;right:0;background:#1a1a2a;border:1px solid #333;border-radius:6px;z-index:200;max-height:180px;overflow-y:auto"></div>
        </div>
        <div id="op-ticker-status" style="font-size:10px;margin-top:3px;color:#555">Ingresá un ticker del modelo</div>
      </div>
      <!-- Tipo operación -->
      <div>
        <label style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.6px;display:block;margin-bottom:4px">Operación</label>
        <select id="op-tipo" style="width:100%;background:#0d0d0f;border:1px solid #333;border-radius:6px;padding:8px 10px;color:#e8e8ea;font-size:13px">
          <option value="COMPRA">🟢 COMPRA</option>
          <option value="VENTA">🔴 VENTA</option>
        </select>
      </div>
      <!-- Tipo instrumento -->
      <div>
        <label style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.6px;display:block;margin-bottom:4px">
          Instrumento
          <span style="color:#f59e0b;font-size:9px;margin-left:4px">⚠️ define el cálculo de P&L</span>
        </label>
        <select id="op-instrumento" onchange="onInstrumentoChange()"
          style="width:100%;background:#0d0d0f;border:1px solid #333;border-radius:6px;padding:8px 10px;color:#e8e8ea;font-size:13px">
          <option value="MERVAL_CSV">🇦🇷 Acción local MERVAL (.BA)</option>
          <option value="BOVESPA_CSV">🇧🇷 Acción local BOVESPA (.SA)</option>
          <option value="SP500_CSV">🇺🇸 CEDEAR / ETF / ADR (BYMA)</option>
        </select>
      </div>
    </div>
    <!-- Fila 2: Precio + Cantidad + Ratio CEDEAR (condicional) + Botón -->
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr auto;gap:12px;align-items:end">
      <!-- Precio USD/acción -->
      <div>
        <label style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.6px;display:block;margin-bottom:4px">
          Precio USD/acción <span id="op-precio-hint" style="color:#4ade80;font-size:9px"></span>
        </label>
        <input id="op-precio" type="number" step="0.0001" min="0" placeholder="Precio de compra en USD"
          style="width:100%;background:#0d0d0f;border:1px solid #333;border-radius:6px;padding:8px 10px;color:#e8e8ea;font-size:13px">
      </div>
      <!-- Cantidad nominal -->
      <div>
        <label style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.6px;display:block;margin-bottom:4px">Cantidad (VN)</label>
        <input id="op-cantidad" type="number" step="1" min="1" placeholder="Unidades"
          style="width:100%;background:#0d0d0f;border:1px solid #333;border-radius:6px;padding:8px 10px;color:#e8e8ea;font-size:13px">
      </div>
      <!-- Ratio CEDEAR — oculto, se usa 1.0 siempre (precio ingresado ya es USD broker) -->
      <input type="hidden" id="op-ratio" value="1.0">
      <div id="op-ratio-wrap" style="display:none"></div>
      <div id="op-ratio-placeholder"></div>
      <!-- Botón -->
      <div>
        <button id="op-submit-btn" onclick="registrarOperacion()" disabled
          style="opacity:0.4;cursor:not-allowed;background:#1e3a5f;border:1px solid #5ba3ff;color:#5ba3ff;padding:9px 20px;border-radius:6px;font-size:13px;font-weight:600;white-space:nowrap">
          + Registrar
        </button>
      </div>
    </div>
    <div id="op-form-msg" style="margin-top:10px;font-size:12px;display:none"></div>
  </div>

  <!-- ── HISTORIAL DE OPERACIONES ── -->
  <div class="section-title" style="color:#aaa">📋 Historial de Operaciones</div>
  <div id="op-historial-wrap">
    <table class="tbl" id="op-historial-table">
      <tr><th>Fecha</th><th>Tipo</th><th>Ticker</th><th>Mercado</th><th>Señal actual</th><th>Precio unit.</th><th>Cantidad</th><th>Total</th><th></th></tr>
    </table>
    <div id="op-historial-empty" style="color:#555;font-size:12px;padding:12px">Sin operaciones registradas.</div>
  </div>
</div>
<script>
function sw(id,el){{
  document.querySelectorAll('.tab').forEach(function(t){{t.classList.remove('on');}});
  document.querySelectorAll('.page').forEach(function(p){{p.classList.remove('on');}});
  el.classList.add('on'); document.getElementById(id).classList.add('on');
  window.scrollTo(0,0);
  setTimeout(function(){{window.dispatchEvent(new Event('resize'));}},50);
}}
</script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js" onerror="document.getElementById('cdnError').style.display='block'"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js" onerror=""></script>
<div id="cdnError" style="display:none;background:#b91c1c;color:white;padding:14px 24px;font-size:14px;position:sticky;top:42px;z-index:200">⚠️ Chart.js no pudo cargar (CDN bloqueado por firewall). Tablas y datos se muestran sin gráficos.</div>
<script>
try {{
var SIGNALS = {signals_json};
var PORTFOLIO = {portfolio_json};
var PORTFOLIO_ALERTS = {portfolio_alerts_json};
var RAILWAY_API_URL  = 'https://inversiones-bursatiles-production.up.railway.app';
var IDX     = {index_stats_json};
var FICHAS  = {fichas_json};
var mL = {json.dumps(merval_labels)};
var mV = {json.dumps(merval_values)};
var bL = {json.dumps(bovespa_labels)};
var bV = {json.dumps(bovespa_values)};
var sL = {json.dumps(sp500_labels)};
var sV = {json.dumps(sp500_values)};
var opChartInst = null;
 
function fn(v,d){{ if(v==null||isNaN(v))return'—'; return Number(v).toFixed(d!=null?d:2); }}
function fp(v){{ return (v>=0?'+':'')+Number(v).toFixed(1)+'%'; }}
function rc(v){{ return v>=0?'#4ade80':'#f87171'; }}
 
// Day refresh
function refreshDay(){{
  [['pano-m-day','dot-m','merval','pano-m-fecha'],
   ['pano-b-day','dot-b','bovespa','pano-b-fecha'],
   ['pano-s-day','dot-s','sp500','pano-s-fecha']].forEach(function(x){{
    var val=IDX[x[2]]&&IDX[x[2]].ret_dia!==undefined?IDX[x[2]].ret_dia:null;
    var fecha=IDX[x[2]]&&IDX[x[2]].fecha?IDX[x[2]].fecha:'';
    var e=document.getElementById(x[0]),d=document.getElementById(x[1]),f=document.getElementById(x[3]);
    var c=val===null?'#888':val>=0?'#4ade80':'#f87171';
    if(e){{e.textContent=val===null?'—':(val>=0?'+':'')+val.toFixed(2)+'%';e.style.color=c;}}
    if(d) d.style.background=c;
    if(f&&fecha){{f.textContent='Cierre: '+fecha;}}
  }});
}}
refreshDay(); setInterval(refreshDay,10000);
 
function sw(id,el){{
  document.querySelectorAll('.tab').forEach(function(t){{t.classList.remove('on');}});
  document.querySelectorAll('.page').forEach(function(p){{p.classList.remove('on');}});
  el.classList.add('on'); document.getElementById(id).classList.add('on');
  setTimeout(function(){{window.dispatchEvent(new Event('resize'));}},50);
}}
 
function sigColor(s){{
  if(s.indexOf('COMPRA FUERTE')>=0) return '#ffd700';
  if(s.indexOf('COMPRA')>=0)        return '#4ade80';
  if(s.indexOf('NEUTRAL')>=0)       return '#fbbf24';
  if(s.indexOf('VENTA P')>=0)       return '#fb923c';
  return '#f87171';
}}
 
function buildTable(tbId,market){{
var signalOrder={{'⭐ COMPRA FUERTE':0,'🟢 COMPRA':1,'🟡 NEUTRAL/ESPERAR':2,'🟠 VENTA PARCIAL':3,'🔴 VENTA':4}};
var mktOrder={{'MERVAL':1,'BOVESPA':2,'SP500':3}};
var rows=market?SIGNALS.filter(function(s){{return s.mercado===market;}}):SIGNALS.slice();
rows.sort(function(a,b){{
  if(!market){{var ma=mktOrder[a.mercado]||9,mb=mktOrder[b.mercado]||9;if(ma!==mb) return ma-mb;}}
  var sa=signalOrder[a.signal_v2||a.signal],sb=signalOrder[b.signal_v2||b.signal];
  if(sa==null)sa=2;if(sb==null)sb=2;if(sa!==sb)return sa-sb;
  return (b.ranking_accionable||b.score_final)-(a.ranking_accionable||a.score_final);
}});
if(!market) rows=rows.slice(0,50);
var tb=document.getElementById(tbId); if(!tb) return;
var lastMkt='';
tb.innerHTML='<tr>'+
'<th>Ticker</th>'+
'<th>Empresa</th>'+
'<th>Precio</th>'+
'<th data-tip=\"Retorno últimos 7 días hábiles\">Sem%</th>'+
'<th data-tip=\"Retorno últimos 30 días hábiles\">Mes%</th>'+
'<th data-tip=\"RSI(14): mide sobrecompra/sobreventa. <30=sobrevendido (oportunidad), >70=sobrecomprado (precaución). El modelo lo usa INVERSO: RSI bajo = score alto.\">RSI</th>'+
'<th data-tip=\"Asset Quality (0-100): calidad estructural del activo. Macro×0.45 + Fundamental×0.35 + Sectorial×0.20. No considera timing.\">AQ</th>'+
'<th data-tip=\"Entry Score (0-100): calidad del timing de entrada. Técnico×0.55 + dist.máximo×0.25 + dist.soporte×0.20. Boosters por divergencia RSI (+10) y squeeze (+15).\">ES</th>'+
'<th data-tip=\"Risk/Reward: (target-precio)/(precio-stop). >2x es favorable. Target = primera resistencia ≥3% del precio. Stop = ATR dinámico por volatilidad y régimen.\">R/R</th>'+
'<th data-tip=\"Score final V2 (0-100): blend de AQ y ES ponderado por mercado. Incluye penalizaciones por tendencia semanal y mensual bajista.\">Score V2</th>'+
'<th data-tip=\"Ranking accionable: Score V2×0.50 + AQ×0.30 + vol_score×0.20. Ordena la tabla. Es la prioridad real de acción entre los 68 tickers.\">Rank</th>'+
'<th data-tip=\"Señal final: ⭐≥70 / 🟢58-69 / 🟡45-57 / 🟠35-44 / 🔴<35\">Señal V2</th>'+
'<th style=\"color:#a78bfa\" data-tip=\"Predicción GBR Ensemble a 21 días (%). Si negativo + señal COMPRA → override automático degrada la señal.\">📈21d</th>'+
'<th style=\"color:#a78bfa\" data-tip=\"Confianza del predictor (0-100%). Cross-validation del GBR. <50% = predicción poco confiable.\">🎯</th>'+
'<th style=\"color:#c084fc\" data-tip=\"Alineación de 3 timeframes: 3✓=Triple confirmación (+5pts) / 2✓=Doble (+2pts) / 1✗=Conflicto parcial (-5pts) / 2✗=Conflicto total (-8pts)\">Align</th>'+
'<th data-tip=\"Tendencia Mensual: MA6M vs MA12M + momentum 6m. ▲=alcista / ▼=bajista(×0.93 al score) / ●=lateral\">📅M</th>'+
'</tr>'+
rows.map(function(s){{
var aq=s.asset_quality||0, es=s.entry_score||0, rr=s.rr_ratio||0, sv2=s.score_final_v2||s.score_final, ra=s.ranking_accionable||sv2, sig2=s.signal_v2||s.signal;
var mktSep='';
if(!market&&s.mercado!==lastMkt){{lastMkt=s.mercado;var fl=s.mercado==='MERVAL'?'🇦🇷':s.mercado==='BOVESPA'?'🇧🇷':'🇺🇸';mktSep='<tr><td colspan="12" style="background:#111118;padding:8px 12px;font-weight:700;color:#5ba3ff;font-size:13px;border-bottom:2px solid #5ba3ff">'+fl+' '+s.mercado+'</td></tr>';}}
return mktSep+'<tr><td class="ticker">'+s.ticker+'</td><td style="color:#ccc">'+s.empresa.substring(0,22)+'</td><td>'+s.precio_actual.toLocaleString('es-AR')+'</td><td style="color:'+rc(s.ret_sem)+';font-weight:600">'+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(1)+'%</td><td style="color:'+rc(s.ret_mes)+';font-weight:600">'+(s.ret_mes>=0?'+':'')+s.ret_mes.toFixed(1)+'%</td><td>'+s.rsi.toFixed(0)+'</td><td style="color:#bc8cff;font-weight:600">'+aq.toFixed(1)+'</td><td style="color:#5ba3ff;font-weight:600">'+es.toFixed(1)+'</td><td style="color:#fbbf24;font-weight:600">'+rr.toFixed(1)+'x</td><td style="color:'+sigColor(sig2)+';font-weight:700">'+sv2.toFixed(1)+'</td><td style="font-weight:900;color:#fff">'+ra.toFixed(1)+'</td><td style="color:'+sigColor(sig2)+';font-weight:600">'+sig2+'</td>'+(s.pred_21d!=null?'<td style="color:'+(s.pred_21d>=0?'#4ade80':'#f87171')+';font-size:11px;font-weight:700">'+(s.pred_21d>=0?'+':'')+s.pred_21d.toFixed(1)+'%</td>':'<td style="color:#444">—</td>')+(s.pred_confidence?'<td style="color:#a78bfa;font-size:11px">'+Math.round(s.pred_confidence*100)+'%</td>':'<td style="color:#444">—</td>')+(s.alignment_label?'<td style="font-size:10px;font-weight:700;color:'+(s.alignment_label.indexOf('TRIPLE')>=0?'#4ade80':s.alignment_label.indexOf('DOBLE')>=0?'#a78bfa':s.alignment_label.indexOf('CONFLICTO')>=0?'#f87171':'#666')+'">'+(s.alignment_label.indexOf('TRIPLE')>=0?'3✓':s.alignment_label.indexOf('DOBLE')>=0?'2✓':'?')+'</td>':'<td style="color:#444">—</td>')+(s.monthly_trend&&s.monthly_trend!=='SIN DATOS'?'<td style="color:'+(s.monthly_trend==='ALCISTA'?'#4ade80':s.monthly_trend==='BAJISTA'?'#f87171':'#fbbf24')+';font-size:12px;font-weight:700">'+(s.monthly_trend==='ALCISTA'?'▲':s.monthly_trend==='BAJISTA'?'▼':'●')+'</td>':'<td style="color:#444">—</td>')+'</tr>';}}).join('');}}
function buildStats(divId,marketKey){{
  var st=IDX[marketKey]||{{}};
  var d=document.getElementById(divId); if(!d) return;
  d.innerHTML=[['Cierre actual',st.actual?st.actual.toLocaleString('es-AR'):'—',''],
    ['Variación 12m',st.ret_anual!=null?(st.ret_anual>=0?'+':'')+st.ret_anual.toFixed(2)+'%':'—',st.ret_anual>=0?'#4ade80':'#f87171'],
    ['Máximo 12m',st.max_12m?st.max_12m.toLocaleString('es-AR'):'—','#fbbf24'],
    ['Mínimo 12m',st.min_12m?st.min_12m.toLocaleString('es-AR'):'—','#f87171']]
    .map(function(x){{return '<div class="card"><div class="card-title">'+x[0]+'</div><div class="card-value" style="color:'+(x[2]||'#fff')+'">'+x[1]+'</div></div>';}}).join('');
}}
 
function normalize(arr){{ var b=arr[0]||1; return arr.map(function(v){{return +(v/b*100).toFixed(1);}}); }}
var scaleOpts={{x:{{ticks:{{color:'#666',font:{{size:11}},autoSkip:true,maxTicksLimit:12,maxRotation:45}},grid:{{color:'rgba(255,255,255,.05)'}}}},y:{{ticks:{{color:'#666',font:{{size:11}}}},grid:{{color:'rgba(255,255,255,.05)'}}}}}} ;
 
var monthMap={{Jan:0,Feb:1,Mar:2,Apr:3,May:4,Jun:5,Jul:6,Aug:7,Sep:8,Oct:9,Nov:10,Dec:11}};
function labelToDate(l){{
  var p=l.split('-'),yr=parseInt(p[1])+(parseInt(p[1])<50?2000:1900);
  return new Date(yr,monthMap[p[0]]||0,1);
}}
 
// ── VISIBILITY TRICK: show all pages so canvas has real dimensions ──
var allPages=document.querySelectorAll('.page');
allPages.forEach(function(p){{p.style.display='block';p.style.visibility='hidden';p.style.position='absolute';}});
 
try {{
if(typeof Chart!=='undefined'){{
if(mL.length&&bL.length&&sL.length){{
  var allLabels=[].concat(mL,bL,sL).filter(function(v,i,a){{return a.indexOf(v)===i;}}).sort(function(a,b){{return labelToDate(a)-labelToDate(b);}});
  function pick(labels,vals,all){{return all.map(function(l){{var i=labels.indexOf(l);return i>=0?vals[i]:null;}});}}
  var mN=normalize(pick(mL,mV,allLabels).filter(function(v){{return v!==null;}}));
  var bN=normalize(pick(bL,bV,allLabels).filter(function(v){{return v!==null;}}));
  var sN=normalize(pick(sL,sV,allLabels).filter(function(v){{return v!==null;}}));
  new Chart(document.getElementById('chartPano'),{{type:'line',data:{{labels:allLabels,datasets:[
    {{label:'MERVAL',data:mN,borderColor:'#5ba3ff',borderWidth:2.5,pointRadius:3,tension:.3,fill:false}},
    {{label:'BOVESPA',data:bN,borderColor:'#4ade80',borderWidth:2,pointRadius:3,tension:.3,fill:false,borderDash:[5,4]}},
    {{label:'S&P 500',data:sN,borderColor:'#fbbf24',borderWidth:2,pointRadius:3,tension:.3,fill:false,borderDash:[2,3]}},
  ]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{mode:'index',intersect:false}}}},scales:scaleOpts}}}});
}}
 
if(mL.length) new Chart(document.getElementById('chartMerval'),{{type:'line',data:{{labels:mL,datasets:[{{data:mV,borderColor:'#5ba3ff',borderWidth:2.5,pointRadius:3,fill:true,backgroundColor:'rgba(91,163,255,.07)',tension:.3}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:scaleOpts}}}});
if(bL.length) new Chart(document.getElementById('chartBovespa'),{{type:'line',data:{{labels:bL,datasets:[{{data:bV,borderColor:'#4ade80',borderWidth:2,pointRadius:3,fill:true,backgroundColor:'rgba(74,222,128,.07)',tension:.3}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:scaleOpts}}}});
if(sL.length) new Chart(document.getElementById('chartSP500'),{{type:'line',data:{{labels:sL,datasets:[{{data:sV,borderColor:'#fbbf24',borderWidth:2,pointRadius:3,fill:true,backgroundColor:'rgba(251,191,36,.07)',tension:.3}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:scaleOpts}}}});
}} else {{ document.getElementById('cdnError').style.display='block'; }}
}} catch(chartErr){{ document.getElementById('cdnError').style.display='block'; }}

// ══════════════════════════════════════════════════════════════════════════
// A) SCATTER: Score V2 vs Predicción 21d
// ══════════════════════════════════════════════════════════════════════════
(function(){{
  var el=document.getElementById('chartScatter'); if(!el) return;
  var mkColors={{'MERVAL':'rgba(91,163,255,0.85)','BOVESPA':'rgba(74,222,128,0.85)','SP500':'rgba(251,191,36,0.85)'}};
  var datasets={{}};
  SIGNALS.forEach(function(s){{
    if(s.pred_21d==null||s.score_final_v2==null) return;
    var mk=s.mercado||'SP500';
    if(!datasets[mk]) datasets[mk]={{label:mk,data:[],backgroundColor:mkColors[mk]||'#888',pointRadius:7,pointHoverRadius:10}};
    datasets[mk].data.push({{x:s.score_final_v2,y:s.pred_21d,ticker:s.ticker,signal:s.signal_v2||s.signal,empresa:s.empresa||''}});
  }});
  var ds=Object.values(datasets);
  if(!ds.length) return;
  new Chart(el,{{
    type:'scatter',
    data:{{datasets:ds}},
    options:{{
      responsive:true,maintainAspectRatio:false,
      plugins:{{
        legend:{{display:true,position:'top',labels:{{color:'#aaa',font:{{size:12}}}}}},
        tooltip:{{callbacks:{{
          label:function(ctx){{
            var d=ctx.raw;
            return [ctx.dataset.label+': '+d.ticker,(d.empresa||'').substring(0,25),'Score V2: '+ctx.parsed.x.toFixed(1),'Pred 21d: '+(d.y>=0?'+':'')+d.y.toFixed(1)+'%','Señal: '+d.signal];
          }}
        }}}},
        annotation:{{}}
      }},
      scales:{{
        x:{{title:{{display:true,text:'Score V2',color:'#888',font:{{size:12}}}},min:0,max:100,ticks:{{color:'#666'}},grid:{{color:'rgba(255,255,255,.05)'}},
          afterDraw:function(ax){{
            var ctx2=ax.chart.ctx;
            var y58=ax.getPixelForValue(58);
            ctx2.save();ctx2.strokeStyle='rgba(74,222,128,0.3)';ctx2.lineWidth=1;ctx2.setLineDash([4,4]);
            ctx2.beginPath();ctx2.moveTo(y58,ax.chart.chartArea.top);ctx2.lineTo(y58,ax.chart.chartArea.bottom);ctx2.stroke();
            ctx2.restore();
          }}
        }},
        y:{{title:{{display:true,text:'Predicción 21d (%)',color:'#888',font:{{size:12}}}},ticks:{{color:'#666',callback:function(v){{return(v>=0?'+':'')+v+'%';}}}},grid:{{color:'rgba(255,255,255,.05)'}},
          afterDraw:function(ax){{
            var ctx2=ax.chart.ctx;
            var y0=ax.getPixelForValue(0);
            ctx2.save();ctx2.strokeStyle='rgba(255,255,255,0.15)';ctx2.lineWidth=1;ctx2.setLineDash([4,4]);
            ctx2.beginPath();ctx2.moveTo(ax.chart.chartArea.left,y0);ctx2.lineTo(ax.chart.chartArea.right,y0);ctx2.stroke();
            ctx2.restore();
          }}
        }}
      }}
    }}
  }});
}})();

// ══════════════════════════════════════════════════════════════════════════
// B) HEATMAP: Sectores × Mercados
// ══════════════════════════════════════════════════════════════════════════
(function(){{
  var container=document.getElementById('heatmap-container'); if(!container) return;
  var markets=['MERVAL','BOVESPA','SP500'];
  var sectorScores={{}};
  SIGNALS.forEach(function(s){{
    var sec=(s.sector||'GENERAL').toUpperCase().substring(0,10);
    var mk=s.mercado||'SP500';
    var v=s.score_final_v2||s.score_final||0;
    var key=sec+'|'+mk;
    if(!sectorScores[key]) sectorScores[key]={{sum:0,n:0}};
    sectorScores[key].sum+=v; sectorScores[key].n++;
  }});
  var sectors=[...new Set(SIGNALS.map(function(s){{return (s.sector||'GENERAL').toUpperCase().substring(0,10);}}))]
    .filter(function(s){{return s&&s!=='ÍNDICE'&&s!=='INDICE';}}).sort();
  if(!sectors.length) return;

  var html='<div class="hm-grid" style="grid-template-columns:120px repeat('+markets.length+',1fr)">';
  // Header
  html+='<div style="font-size:11px;color:#555;padding:4px">Sector</div>';
  markets.forEach(function(mk){{
    var flag=mk==='MERVAL'?'🇦🇷':mk==='BOVESPA'?'🇧🇷':'🇺🇸';
    html+='<div style="font-size:12px;font-weight:700;color:#aaa;text-align:center;padding:4px">'+flag+' '+mk+'</div>';
  }});
  // Rows
  sectors.forEach(function(sec){{
    html+='<div style="font-size:11px;color:#888;padding:4px 6px;display:flex;align-items:center">'+sec+'</div>';
    markets.forEach(function(mk){{
      var key=sec+'|'+mk;
      var d=sectorScores[key];
      if(!d||d.n===0){{html+='<div class="hm-cell" style="background:#111;color:#333;height:36px" title="Sin datos">—</div>';return;}}
      var avg=d.sum/d.n;
      // Color: 0-45 rojo, 45-58 amarillo, 58-70 verde, 70-100 verde intenso
      var r,g,b;
      if(avg>=70){{r=22;g=163;b=74;}}
      else if(avg>=58){{r=45;g=212;b=191;}}
      else if(avg>=45){{r=234;g=179;b=8;}}
      else{{r=220;g=38;b=38;}}
      var alpha=0.25+((avg/100)*0.65);
      html+='<div class="hm-cell" style="background:rgba('+r+','+g+','+b+','+alpha.toFixed(2)+');color:#fff;height:36px;min-width:60px" title="'+sec+' en '+mk+' — Score promedio: '+avg.toFixed(1)+' ('+d.n+' tickers)">'+avg.toFixed(0)+'<br><span style=\"font-size:9px;opacity:.7\">n='+d.n+'</span></div>';
    }});
  }});
  html+='</div>';
  container.innerHTML=html;
}})();

// ══════════════════════════════════════════════════════════════════════════

// ── RESTORE: hide pages, show only Panorama, then force resize ──
allPages.forEach(function(p){{p.style.display='';p.style.visibility='';p.style.position='';}});
setTimeout(function(){{
  window.dispatchEvent(new Event('resize'));
}},100);
 
var globalSorted=SIGNALS.slice().sort(function(a,b){{return (b.ranking_accionable||b.score_final)-(a.ranking_accionable||a.score_final);}});
SIGNALS=globalSorted;
buildTable('tbl-pano-merval','MERVAL'); buildTable('tbl-pano-bovespa','BOVESPA'); buildTable('tbl-pano-sp500','SP500');
buildTable('tbl-merval','MERVAL'); buildTable('tbl-bovespa','BOVESPA'); buildTable('tbl-sp500','SP500');
buildStats('merval-stats','merval'); buildStats('bovespa-stats','bovespa'); buildStats('sp500-stats','sp500');
 
// ── OPPORTUNITY SCORE ──────────────────────────────────────────────────────
// Fórmula institucional: 40% pred_21d + 35% R/R + 25% confianza
function computeOpportunityScore(s){{
  var pred21   = s.pred_21d!=null ? s.pred_21d : 0;
  var conf     = s.pred_confidence!=null ? s.pred_confidence*100 : 0;
  var rr       = s.rr_ratio!=null ? s.rr_ratio : 0;
  var predNorm = Math.min(100, Math.max(0, (pred21+15)/30*100));
  var rrNorm   = Math.min(100, rr/5*100);
  return Math.round(predNorm*0.40 + rrNorm*0.35 + conf*0.25);
}}
function flagOf(m){{ return m==='MERVAL'?'🇦🇷':m==='BOVESPA'?'🇧🇷':'🇺🇸'; }}

// ── FILTRO TRIPLE ────────────────────────────────────────────────────────────
// 1) señal V2 = COMPRA o COMPRA FUERTE
// 2) opportunity_score >= 60
// 3) top 25% del universo (score >= p75)
var universe = SIGNALS.map(function(s){{
  return Object.assign({{}},s,{{opp_score:computeOpportunityScore(s)}});
}});
var buyUniverse = universe.filter(function(s){{
  var sig = s.signal_v2||s.signal||'';
  return sig.indexOf('COMPRA')>=0 && sig.indexOf('VENTA')<0;
}});
var allScores = universe.map(function(s){{return s.opp_score;}}).sort(function(a,b){{return a-b;}});
var p75idx = Math.floor(allScores.length*0.75);
var p75 = allScores[p75idx]||0;
var ranked = buyUniverse
  .filter(function(s){{ return s.opp_score>=50 && s.opp_score>=p75; }})
  .sort(function(a,b){{return b.opp_score-a.opp_score;}});
var radarHtml=ranked.length===0
  ? '<div style="text-align:center;padding:48px 20px;color:#555;font-size:14px">'+
    '<div style="font-size:36px;margin-bottom:12px">🔍</div>'+
    '<div style="font-weight:700;color:#777;margin-bottom:8px">No hay oportunidades con convicción suficiente en este momento</div>'+
    '<div style="font-size:12px;color:#444">El modelo requiere: señal COMPRA + Opportunity Score ≥ 60 + top 25% del universo</div>'+
    '<div style="font-size:11px;color:#333;margin-top:8px">Umbral activo: score ≥ 50 y ≥ p75 ('+p75+')</div></div>'
  : ranked.map(function(s,i){{
  var pfm=s.max_12m>0?((s.max_12m-s.precio_actual)/s.max_12m*100).toFixed(1):'—';
  var tags=[];
  if(s.rsi>=28&&s.rsi<=45) tags.push('<span class="radar-tag tag-green">RSI sobreventa</span>');
  if(s.rsi>45&&s.rsi<=55) tags.push('<span class="radar-tag tag-blue">RSI neutro-pos</span>');
  if(s.ret_sem>0&&s.ret_mes<5) tags.push('<span class="radar-tag tag-green">Arranque temprano</span>');
  if(s.ret_sem>0&&s.ret_mes>=5) tags.push('<span class="radar-tag tag-blue">Momentum activo</span>');
  if(s.ma_cross) tags.push('<span class="radar-tag tag-yellow">Cruce MA</span>');
  if(parseFloat(pfm)>30) tags.push('<span class="radar-tag tag-orange">-'+pfm+'% vs máx</span>');
  var sig=s.signal_v2||s.signal||'';
  if(sig.indexOf('COMPRA FUERTE')>=0) tags.push('<span class="radar-tag tag-green">⭐ Compra Fuerte</span>');
  else tags.push('<span class="radar-tag tag-green">🟢 Compra</span>');
  var sc=s.opp_score>=80?'#22c55e':s.opp_score>=65?'#86efac':'#fbbf24';
  // Desglose del opportunity score
  var pred21=s.pred_21d!=null?s.pred_21d:0;
  var predNorm=Math.min(100,Math.max(0,(pred21+15)/30*100));
  var rrNorm=Math.min(100,(s.rr_ratio||0)/5*100);
  var confNorm=(s.pred_confidence||0)*100;
  var desglose='pred:'+(predNorm*0.40).toFixed(0)+' rr:'+(rrNorm*0.35).toFixed(0)+' conf:'+(confNorm*0.25).toFixed(0);
  return '<div class="radar-card" style="cursor:pointer" onclick="sw(&apos;oportunidades&apos;,document.querySelector(&apos;.tab[onclick*=oportunidades]&apos;));showOpFicha(&apos;'+s.ticker+'&apos;)">'+
    '<div class="radar-rank" style="font-size:'+(i<3?'32px':'24px')+'">#'+(i+1)+'</div>'+
    '<div class="radar-info"><div style="display:flex;align-items:center;gap:8px;margin-bottom:2px">'+
    '<span class="radar-ticker">'+flagOf(s.mercado)+' '+s.ticker+'</span>'+
    '<span style="font-size:12px;color:#666">'+s.empresa.substring(0,28)+'</span></div>'+
    '<div class="radar-metrics">'+
    '<span>💰 '+s.precio_actual.toLocaleString('es-AR')+'</span>'+
    '<span style="color:'+rc(s.ret_sem)+'">Sem: '+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(1)+'%</span>'+
    '<span style="color:'+rc(s.ret_mes)+'">Mes: '+(s.ret_mes>=0?'+':'')+s.ret_mes.toFixed(1)+'%</span>'+
    '<span>RSI: '+s.rsi.toFixed(0)+'</span>'+
    (s.rr_ratio!=null?'<span style="color:#fbbf24">R/R: '+s.rr_ratio.toFixed(1)+'x</span>':'')+
    (s.pred_21d!=null?'<span style="color:'+(s.pred_21d>=0?'#4ade80':'#f87171')+';font-weight:700">21d:'+(s.pred_21d>=0?'+':'')+s.pred_21d.toFixed(1)+'%</span>':'')+
    (s.pred_confidence?'<span style="color:#a78bfa">🎯'+Math.round(s.pred_confidence*100)+'%</span>':'')+'</div>'+
    '<div class="radar-signals">'+tags.join('')+'</div>'+
    '<div class="radar-bar-wrap"><div class="radar-bar" style="width:'+s.opp_score+'%;background:'+sc+'"></div></div>'+
    '<div style="font-size:9px;color:#444;margin-top:2px">'+desglose+'</div></div>'+
    '<div class="radar-score-wrap">'+
    '<div class="radar-score" style="color:'+sc+'">'+s.opp_score+'</div>'+
    '<div class="radar-score-label">Opp<br>Score</div>'+
    '<div style="font-size:10px;color:#4ade80;margin-top:6px;font-weight:700">📈 Ver ficha →</div>'+
    '</div></div>';
}}).join('');
var rb=document.getElementById('radar-block');
if(rb) rb.innerHTML=radarHtml;
 
// ── CONCLUSIONES: Compras confirmadas (PRIMERO) ──────────────────
var compras=SIGNALS.filter(function(s){{var sig=s.signal_v2||s.signal;return sig.indexOf('COMPRA')>=0;}}).sort(function(a,b){{var mo={{'MERVAL':0,'BOVESPA':1,'SP500':2}};var ma=mo[a.mercado]||9,mb=mo[b.mercado]||9;if(ma!==mb)return ma-mb;var aF=(a.signal_v2||a.signal).indexOf('FUERTE')>=0?0:1,bF=(b.signal_v2||b.signal).indexOf('FUERTE')>=0?0:1;if(aF!==bF)return aF-bF;return(b.ranking_accionable||b.score_final)-(a.ranking_accionable||a.score_final);}});
var lastCMkt='';
document.getElementById('compras-block').innerHTML=compras.length?compras.map(function(s,i){{
  var sig=s.signal_v2||s.signal;
  var isFuerte=sig.indexOf('FUERTE')>=0;
  var icon=isFuerte?'⭐':'🟢';
  var upside=s.upside_graham!=null?(s.upside_graham>=0?'+':'')+s.upside_graham.toFixed(1)+'%':'—';
  var mktSep='';
  if(s.mercado!==lastCMkt){{lastCMkt=s.mercado;var fl=flagOf(s.mercado);mktSep='<div style="padding:10px 0 6px;font-weight:700;color:#5ba3ff;font-size:14px;border-bottom:2px solid #5ba3ff;margin-bottom:8px">'+fl+' '+s.mercado+'</div>';}}
  var fichaBtn='';
  for(var fi=0;fi<FICHAS.length;fi++){{if(FICHAS[fi].ticker===s.ticker){{fichaBtn='<div class="dl" style="grid-column:1/-1"><button onclick="event.stopPropagation();sw(&#39;oportunidades&#39;,document.querySelectorAll(&#39;.tab&#39;)[5]);showOpFicha(&#39;'+s.ticker+'&#39;)" style="background:#0d2b1a;border:1px solid #1a3a1a;color:#4ade80;padding:6px 14px;border-radius:6px;cursor:pointer;font-size:12px;width:100%">📈 Ver ficha técnica con gráfico</button></div>';break;}}}}
  return mktSep+'<div class="concl-card-exp buy" onclick="this.classList.toggle(&#39;open&#39;)">'+
    '<div class="concl-header">'+
    '<div class="concl-rank">#'+(i+1)+'</div>'+
    '<div class="concl-main">'+
    '<div class="concl-title-row">'+
    '<span style="font-size:15px;font-weight:700;color:#4ade80">'+flagOf(s.mercado)+' '+icon+' '+sig+'</span>'+
    '<span style="font-size:15px;font-weight:700;color:#fff">'+s.ticker+'</span>'+
    '<span style="font-size:12px;color:#888">'+s.empresa.substring(0,30)+'</span></div>'+
    '<div class="concl-metrics-row">'+
    '<span>Score: <b>'+s.score_final.toFixed(0)+'</b></span>'+
    '<span>RSI: <b>'+s.rsi.toFixed(0)+'</b></span>'+
    '<span style="color:'+rc(s.ret_sem)+'">Sem: <b>'+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(1)+'%</b></span>'+
    '<span style="color:'+rc(s.ret_mes)+'">Mes: <b>'+(s.ret_mes>=0?'+':'')+s.ret_mes.toFixed(1)+'%</b></span>'+
    '<span style="color:'+rc(s.ret_anual)+'">Anual: <b>'+(s.ret_anual>=0?'+':'')+s.ret_anual.toFixed(1)+'%</b></span>'+
    '<span>💰 '+s.precio_actual.toLocaleString('es-AR')+'</span>'+(s.signal_override?'<span style="background:#92400e;color:#fef3c7;font-size:9px;padding:1px 5px;border-radius:3px;margin-left:4px">⚡ '+s.signal_override+'</span>':'')+
    ((s.pred_5d!=null||s.pred_21d!=null)?'<span style="color:#a78bfa;font-size:10px;font-weight:700">📈 PRED:</span>'+
      (s.pred_5d!=null?'<span style="color:'+(s.pred_5d>=0?'#4ade80':'#f87171')+';font-size:11px">5d:'+(s.pred_5d>=0?'+':'')+s.pred_5d.toFixed(1)+'%</span>':'')+
      (s.pred_10d!=null?'<span style="color:'+(s.pred_10d>=0?'#4ade80':'#f87171')+';font-size:11px">10d:'+(s.pred_10d>=0?'+':'')+s.pred_10d.toFixed(1)+'%</span>':'')+
      (s.pred_21d!=null?'<span style="color:'+(s.pred_21d>=0?'#4ade80':'#f87171')+';font-size:11px">21d:'+(s.pred_21d>=0?'+':'')+s.pred_21d.toFixed(1)+'%</span>':'')+
      (s.pred_confidence?'<span style="color:#a78bfa;font-size:11px">🎯'+(s.pred_confidence*100).toFixed(0)+'%</span>':'')
    :'')+
    '</div></div>'+
    '<div class="concl-arrow">▼</div></div>'+
    '<div class="concl-detail"><div class="concl-detail-inner">'+
    '<div class="dl">Ranking accionable <b>'+(s.ranking_accionable||s.score_final).toFixed(1)+'</b></div>'+
    '<div class="dl">Score macro <b>'+(s.score_macro!=null?s.score_macro.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">Score técnico <b>'+(s.score_tecnico!=null?s.score_tecnico.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">Score sectorial <b>'+(s.score_sectorial!=null?s.score_sectorial.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">Momentum 21d <b>'+(s.momentum!=null?s.momentum.toFixed(2):'—')+'</b></div>'+
    '<div class="dl">RS vs Índice <b style="color:'+(s.relative_strength>1.05?'#4ade80':s.relative_strength<0.95?'#f87171':'#aaa')+'">'+(s.relative_strength!=null?s.relative_strength.toFixed(3):'—')+'</b></div>'+
    '<div class="dl">RSI(14) <b>'+s.rsi.toFixed(1)+'</b></div>'+
    '<div class="dl">Upside Graham <b>'+upside+'</b></div>'+
    '<div class="dl">Precio actual <b>'+s.precio_actual.toLocaleString('es-AR')+'</b></div>'+
    '<div class="dl">Máximo 12M <b>'+(s.max_12m!=null?s.max_12m.toLocaleString('es-AR'):'—')+'</b></div>'+
    '<div class="dl">Mínimo 12M <b>'+(s.min_12m!=null?s.min_12m.toLocaleString('es-AR'):'—')+'</b></div>'+
    '<div class="dl">Ret. semanal <b style="color:'+rc(s.ret_sem)+'">'+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(2)+'%</b></div>'+
    '<div class="dl">Ret. mensual <b style="color:'+rc(s.ret_mes)+'">'+(s.ret_mes>=0?'+':'')+s.ret_mes.toFixed(2)+'%</b></div>'+
    '<div class="dl">Ret. anual <b style="color:'+rc(s.ret_anual)+'">'+(s.ret_anual>=0?'+':'')+s.ret_anual.toFixed(2)+'%</b></div>'+
    '<div class="dl">Score cuantitativo <b>'+(s.score_cuant!=null?s.score_cuant.toFixed(1):'—')+'</b></div>'+
    '<div class="dl" style="grid-column:1/-1;border-top:1px solid rgba(255,255,255,0.06);padding-top:6px;margin-top:4px"><b style="color:#60a5fa">📊 FASE 3</b></div>'+
    '<div class="dl">RS vs Índice <b style="color:'+(s.relative_strength>1.05?'#4ade80':s.relative_strength<0.95?'#f87171':'#aaa')+'">'+(s.relative_strength!=null?s.relative_strength.toFixed(3):'—')+'</b></div>'+
    '<div class="dl">ATR Percentil <b>'+(s.atr_percentile!=null?s.atr_percentile.toFixed(0)+'%':'—')+'</b></div>'+
    '<div class="dl">ADX <b style="color:'+(s.adx>25?'#4ade80':'#f87171')+'">'+(s.adx!=null?s.adx.toFixed(1):'—')+'</b></div>'+
    (s.stress_index!=null?'<div class="dl">Stress Index <b style="color:'+(s.stress_index>70?'#f87171':s.stress_index<40?'#4ade80':'#fbbf24')+'">'+s.stress_index.toFixed(0)+'</b></div>':'')+
    (s.rr_ratio!=null?'<div class="dl">R/R Ratio <b style="color:#fbbf24">'+s.rr_ratio.toFixed(2)+'x</b></div>':'')+
    (s.atr_stop!=null?'<div class="dl">ATR Stop <b style="color:#fb923c">'+s.atr_stop.toLocaleString(\'es-AR\')+'</b></div>':'')+
    (s.atr_target!=null?'<div class="dl">ATR Target <b style="color:#bc8cff">'+s.atr_target.toLocaleString(\'es-AR\')+'</b></div>':'')+
    ((s.pred_5d!=null||s.pred_21d!=null)?'<div class="dl" style="grid-column:1/-1;border-top:1px solid rgba(255,255,255,0.06);padding-top:6px;margin-top:4px"><b style="color:#a78bfa">🔭 PREDICCIÓN ENSEMBLE</b></div>'+
      (s.pred_5d!=null?'<div class="dl">Pred. 5d <b style="color:'+(s.pred_5d>=0?'#4ade80':'#f87171')+'">'+(s.pred_5d>=0?'+':'')+s.pred_5d.toFixed(1)+'%</b></div>':'')+
      (s.pred_21d!=null?'<div class="dl">Pred. 21d <b style="color:'+(s.pred_21d>=0?'#4ade80':'#f87171')+'">'+(s.pred_21d>=0?'+':'')+s.pred_21d.toFixed(1)+'%</b></div>':'')+
      (s.pred_target?'<div class="dl">Target pred. <b style="color:#bc8cff">'+s.pred_target.toLocaleString(\'es-AR\')+'</b></div>':'')+
      (s.pred_confidence?'<div class="dl">Confianza <b style="color:#a78bfa">'+fn(s.pred_confidence*100,0)+'%</b></div>':'')+
      '<div class="dl">Señal pred. <b style="color:#a78bfa">'+s.pred_signal+'</b></div>'+
      (s.pred_direction_agree!==undefined?'<div class="dl">vs Modelo <b style="color:'+(s.pred_direction_agree?'#4ade80':'#fbbf24')+'">'+(s.pred_direction_agree?'✅ Confirma':'⚠️ Diverge')+'</b></div>':'')
    :'')+
    '</div></div></div>';
}}).join(''):'<div style="color:#666;padding:16px">Sin señales de compra activas.</div>';
 
// ── CONCLUSIONES: Radar de Oportunidades Tempranas (SEGUNDO) ─────
// radarItems ahora usa ranked (ya calculado con filtro triple y opportunity_score)
var radarItems=ranked.slice();
var radarHtml='';var lastRMkt='';
radarItems.forEach(function(s,i){{
  var pfm=s.max_12m?((1-s.precio_actual/s.max_12m)*100).toFixed(1):'?';
  var tags=[];
  if(s.signal.indexOf('COMPRA FUERTE')>=0) tags.push('<span class="radar-tag tag-green">⭐ Compra Fuerte</span>');
  else if(s.signal.indexOf('COMPRA')>=0) tags.push('<span class="radar-tag tag-green">🟢 Compra</span>');
  else tags.push('<span class="radar-tag tag-yellow">🟡 Monitorear</span>');
  var oppSc2=s.opp_score!=null?s.opp_score:computeOpportunityScore(s);
  var sc=oppSc2>=80?'#22c55e':oppSc2>=65?'#86efac':'#fbbf24';
  if(s.mercado!==lastRMkt){{lastRMkt=s.mercado;radarHtml+='<div style="padding:10px 0 6px;font-weight:700;color:#5ba3ff;font-size:14px;border-bottom:2px solid #5ba3ff;margin-bottom:8px">'+flagOf(s.mercado)+' '+s.mercado+'</div>';}}
  radarHtml+='<div class="concl-card-exp radar" onclick="this.classList.toggle(&#39;open&#39;)">'+
    '<div class="concl-header">'+
    '<div class="concl-rank" style="color:'+sc+'">#'+(i+1)+'</div>'+
    '<div class="concl-main">'+
    '<div class="concl-title-row">'+
    '<span class="radar-ticker">'+flagOf(s.mercado)+' '+s.ticker+'</span>'+
    '<span style="font-size:12px;color:#888">'+s.empresa.substring(0,28)+'</span>'+
    tags.join('')+
    '<span style="font-size:13px;font-weight:700;color:'+sc+'">Opp: '+oppSc2+'</span></div>'+
    '<div class="concl-metrics-row">'+
    '<span>💰 '+s.precio_actual.toLocaleString('es-AR')+'</span>'+(s.signal_override?'<span style="background:#92400e;color:#fef3c7;font-size:9px;padding:1px 5px;border-radius:3px;margin-left:4px">⚡ '+s.signal_override+'</span>':'')+
    '<span style="color:'+rc(s.ret_sem)+'">Sem: <b>'+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(1)+'%</b></span>'+
    '<span style="color:'+rc(s.ret_mes)+'">Mes: <b>'+(s.ret_mes>=0?'+':'')+s.ret_mes.toFixed(1)+'%</b></span>'+
    '<span>RSI: <b>'+s.rsi.toFixed(0)+'</b></span>'+
    '<span style="color:#fb923c">-'+pfm+'% vs máx</span>'+
    '</div></div>'+
    '<div class="concl-arrow">▼</div></div>'+
    '<div class="concl-detail"><div class="concl-detail-inner">'+
    '<div class="dl">Score radar <b style="color:'+sc+'">'+oppSc2+'</b></div>'+
    '<div class="dl">Score final modelo <b>'+s.score_final.toFixed(0)+'</b></div>'+
    '<div class="dl">Score macro <b>'+(s.score_macro!=null?s.score_macro.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">Score técnico <b>'+(s.score_tecnico!=null?s.score_tecnico.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">RSI(14) <b>'+s.rsi.toFixed(1)+'</b></div>'+
    '<div class="dl">Momentum 21d <b>'+(s.momentum!=null?s.momentum.toFixed(2):'—')+'</b></div>'+
    '<div class="dl">RS vs Índice <b style="color:'+(s.relative_strength>1.05?'#4ade80':s.relative_strength<0.95?'#f87171':'#aaa')+'">'+(s.relative_strength!=null?s.relative_strength.toFixed(3):'—')+'</b></div>'+
    '<div class="dl">Precio actual <b>'+s.precio_actual.toLocaleString('es-AR')+'</b></div>'+
    '<div class="dl">Máximo 12M <b>'+(s.max_12m!=null?s.max_12m.toLocaleString('es-AR'):'—')+'</b></div>'+
    '<div class="dl">Dist. vs máximo <b style="color:#fb923c">-'+pfm+'%</b></div>'+
    '<div class="dl">Ret. semanal <b style="color:'+rc(s.ret_sem)+'">'+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(2)+'%</b></div>'+
    '<div class="dl">Ret. mensual <b style="color:'+rc(s.ret_mes)+'">'+(s.ret_mes>=0?'+':'')+s.ret_mes.toFixed(2)+'%</b></div>'+
    '<div class="dl">Ret. anual <b style="color:'+rc(s.ret_anual)+'">'+(s.ret_anual>=0?'+':'')+s.ret_anual.toFixed(2)+'%</b></div>'+
    (s.rr_ratio!=null?'<div class="dl">R/R Ratio <b style="color:#fbbf24">'+s.rr_ratio.toFixed(2)+'x</b></div>':'')+
    (s.atr_stop!=null?'<div class="dl">ATR Stop <b style="color:#fb923c">'+s.atr_stop.toLocaleString(\'es-AR\')+'</b></div>':'')+
    (s.atr_target!=null?'<div class="dl">ATR Target <b style="color:#bc8cff">'+s.atr_target.toLocaleString(\'es-AR\')+'</b></div>':'')+
    '<div class="radar-bar-wrap" style="grid-column:1/-1"><div class="radar-bar" style="width:'+oppSc2+'%;background:'+sc+'"></div></div>'+
    '</div></div></div>';
}});
var rb=document.getElementById('radar-block');
if(rb) rb.innerHTML=radarHtml||'<div style="color:#666;padding:20px">Sin datos de radar.</div>';
 
// ── CONCLUSIONES: Señales de Reducción (TERCERO) ─────────────────
var ventas=SIGNALS.filter(function(s){{var sig=s.signal_v2||s.signal;return sig.indexOf('VENTA')>=0;}}).sort(function(a,b){{var mo={{"MERVAL":0,"BOVESPA":1,"SP500":2}};var ma=mo[a.mercado]||9,mb=mo[b.mercado]||9;if(ma!==mb)return ma-mb;var sO={{"🔴 VENTA":0,"🟠 VENTA PARCIAL":1}};var sa=sO[a.signal_v2||a.signal]||1,sb=sO[b.signal_v2||b.signal]||1;if(sa!==sb)return sa-sb;return (a.ranking_accionable||a.score_final)-(b.ranking_accionable||b.score_final);}});
var ventasHtml='';var lastVMkt='';
ventas.forEach(function(s,i){{
  var sig=s.signal_v2||s.signal;
  var isVenta=sig.indexOf('VENTA PARCIAL')<0;
  var icon=isVenta?'🔴':'🟠';
  if(s.mercado!==lastVMkt){{lastVMkt=s.mercado;ventasHtml+='<div style="padding:10px 0 6px;font-weight:700;color:#f87171;font-size:14px;border-bottom:2px solid #f87171;margin-bottom:8px">'+flagOf(s.mercado)+' '+s.mercado+'</div>';}}
  ventasHtml+='<div class="concl-card-exp sell" onclick="this.classList.toggle(&#39;open&#39;)">'+
    '<div class="concl-header">'+
    '<div class="concl-rank">#'+(i+1)+'</div>'+
    '<div class="concl-main">'+
    '<div class="concl-title-row">'+
    '<span style="font-size:15px;font-weight:700;color:'+(isVenta?'#f04d5a':'#fb923c')+'">'+flagOf(s.mercado)+' '+icon+' '+sig+'</span>'+
    '<span style="font-size:15px;font-weight:700;color:#fff">'+s.ticker+'</span>'+
    '<span style="font-size:12px;color:#888">'+s.empresa.substring(0,30)+'</span></div>'+
    '<div class="concl-metrics-row">'+
    '<span>Score: <b>'+s.score_final.toFixed(0)+'</b></span>'+
    '<span>RSI: <b>'+s.rsi.toFixed(0)+'</b></span>'+
    '<span style="color:'+rc(s.ret_sem)+'">Sem: <b>'+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(1)+'%</b></span>'+
    '<span style="color:'+rc(s.ret_anual)+'">Anual: <b>'+(s.ret_anual>=0?'+':'')+s.ret_anual.toFixed(1)+'%</b></span>'+
    '<span>💰 '+s.precio_actual.toLocaleString('es-AR')+'</span>'+(s.signal_override?'<span style="background:#92400e;color:#fef3c7;font-size:9px;padding:1px 5px;border-radius:3px;margin-left:4px">⚡ '+s.signal_override+'</span>':'')+
    ((s.pred_5d!=null||s.pred_21d!=null)?'<span style="color:#a78bfa;font-size:10px;font-weight:700">📈 PRED:</span>'+
      (s.pred_5d!=null?'<span style="color:'+(s.pred_5d>=0?'#4ade80':'#f87171')+';font-size:11px">5d:'+(s.pred_5d>=0?'+':'')+s.pred_5d.toFixed(1)+'%</span>':'')+
      (s.pred_10d!=null?'<span style="color:'+(s.pred_10d>=0?'#4ade80':'#f87171')+';font-size:11px">10d:'+(s.pred_10d>=0?'+':'')+s.pred_10d.toFixed(1)+'%</span>':'')+
      (s.pred_21d!=null?'<span style="color:'+(s.pred_21d>=0?'#4ade80':'#f87171')+';font-size:11px">21d:'+(s.pred_21d>=0?'+':'')+s.pred_21d.toFixed(1)+'%</span>':'')+
      (s.pred_confidence?'<span style="color:#a78bfa;font-size:11px">🎯'+(s.pred_confidence*100).toFixed(0)+'%</span>':'')
    :'')+
    '</div></div>'+
    '<div class="concl-arrow">▼</div></div>'+
    '<div class="concl-detail"><div class="concl-detail-inner">'+
    '<div class="dl">Ranking accionable <b>'+(s.ranking_accionable||s.score_final).toFixed(1)+'</b></div>'+
    '<div class="dl">Score macro <b>'+(s.score_macro!=null?s.score_macro.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">Score técnico <b>'+(s.score_tecnico!=null?s.score_tecnico.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">Score sectorial <b>'+(s.score_sectorial!=null?s.score_sectorial.toFixed(0):'—')+'</b></div>'+
    '<div class="dl">RSI(14) <b>'+s.rsi.toFixed(1)+'</b></div>'+
    '<div class="dl">Momentum 21d <b>'+(s.momentum!=null?s.momentum.toFixed(2):'—')+'</b></div>'+
    '<div class="dl">RS vs Índice <b style="color:'+(s.relative_strength>1.05?'#4ade80':s.relative_strength<0.95?'#f87171':'#aaa')+'">'+(s.relative_strength!=null?s.relative_strength.toFixed(3):'—')+'</b></div>'+
    '<div class="dl">Precio actual <b>'+s.precio_actual.toLocaleString('es-AR')+'</b></div>'+
    '<div class="dl">Máximo 12M <b>'+(s.max_12m!=null?s.max_12m.toLocaleString('es-AR'):'—')+'</b></div>'+
    '<div class="dl">Ret. semanal <b style="color:'+rc(s.ret_sem)+'">'+(s.ret_sem>=0?'+':'')+s.ret_sem.toFixed(2)+'%</b></div>'+
    '<div class="dl">Ret. mensual <b style="color:'+rc(s.ret_mes)+'">'+(s.ret_mes>=0?'+':'')+s.ret_mes.toFixed(2)+'%</b></div>'+
    '<div class="dl">Ret. anual <b style="color:'+rc(s.ret_anual)+'">'+(s.ret_anual>=0?'+':'')+s.ret_anual.toFixed(2)+'%</b></div>'+
    '<div class="dl">Score cuantitativo <b>'+(s.score_cuant!=null?s.score_cuant.toFixed(1):'—')+'</b></div>'+
    '</div></div></div>';
}});
document.getElementById('ventas-block').innerHTML=ventasHtml||'<div style="color:#666;padding:16px">Sin señales de venta activas.</div>';

// ── CONCLUSIONES: ¿Dónde poner el próximo dólar? ─────────────────────────
(function(){{
  var cb=document.getElementById('capital-block');
  if(!cb) return;

  // Tickers ya en cartera
  var portMap={{}};
  if(PORTFOLIO&&PORTFOLIO.positions){{
    PORTFOLIO.positions.forEach(function(p){{
      portMap[p.ticker]={{
        pnl_pct: p.precio_compra_usd>0&&p.precio_actual_usd>0
                  ? ((p.precio_actual_usd/p.precio_compra_usd-1)*100) : null,
        pnl_usd: p.rend_usd||null,
        cantidad: p.cantidad
      }};
    }});
  }}

  // Candidatos: señales de COMPRA
  var cands=SIGNALS.filter(function(s){{
    var sig=s.signal_v2||s.signal||'';
    return sig.indexOf('COMPRA')>=0;
  }}).map(function(s){{
    var sv2  = s.score_final_v2||s.score_final||0;
    var rr   = s.rr_ratio||0;
    var pred = s.pred_21d||0;
    var aq   = s.asset_quality||0;
    var es   = s.entry_score||0;
    var rank = s.ranking_accionable||sv2;
    // Retorno esperado = V2 × max(R/R,0.5) + predictor_bonus
    var ev   = sv2 * Math.max(0.5, rr) + (pred>0? pred*3 : pred*1);
    var enC  = !!portMap[s.ticker];
    return {{
      ticker:    s.ticker,
      empresa:   (s.empresa||s.ticker).substring(0,32),
      mercado:   s.mercado,
      signal:    s.signal_v2||s.signal||'',
      sv2:sv2, rr:rr, pred_5d:s.pred_5d, pred_10d:s.pred_10d,
      pred_21d:s.pred_21d, pred_conf:s.pred_confidence,
      pred_signal:s.pred_signal||'',
      pred_agree:s.pred_direction_agree,
      aq:aq, es:es, rank:rank,
      score_mac: s.score_macro, score_tec: s.score_tecnico,
      score_fund:s.score_fund||s.score_fundamental,
      score_sect:s.score_sectorial,
      atr_stop:  s.atr_stop,  atr_target: s.atr_target,
      horizonte: s.horizonte||'—',
      upside_g:  s.upside_graham,
      score_cuant: s.score_cuant,
      ret_sem:   s.ret_sem, ret_mes: s.ret_mes, ret_anual: s.ret_anual,
      precio:    s.precio_actual,
      rs:        s.relative_strength,
      weekly:    s.weekly_trend||'',
      adx:       s.adx,
      ev: ev, enC: enC,
      portData: enC?portMap[s.ticker]:null
    }};
  }});

  // Ordenar: en-cartera primero, luego por ev desc
  cands.sort(function(a,b){{
    if(a.enC!==b.enC) return a.enC?-1:1;
    return b.ev-a.ev;
  }});

  if(!cands.length){{
    cb.innerHTML='<div style="color:#666;padding:20px;text-align:center">Sin señales de compra activas.</div>';
    return;
  }}

  var maxEV=Math.max.apply(null,cands.map(function(c){{return c.ev;}}));
  var lastGrp=null;
  var html='';

  cands.forEach(function(c,i){{
    // Separador de grupo
    var grp=c.enC?'cartera':'nueva';
    if(grp!==lastGrp){{
      lastGrp=grp;
      html+='<div style="display:flex;align-items:center;gap:10px;margin:16px 0 10px">'+
        '<div style="flex:1;height:1px;background:rgba(251,191,36,.25)"></div>'+
        '<span style="font-size:11px;font-weight:700;color:#fbbf24;white-space:nowrap;padding:2px 10px;background:rgba(251,191,36,.1);border-radius:12px">'+
          (c.enC?'📂 POSICIONES EN CARTERA — Agregar':'🆕 NUEVAS POSICIONES SUGERIDAS')+
        '</span>'+
        '<div style="flex:1;height:1px;background:rgba(251,191,36,.25)"></div>'+
      '</div>';
    }}

    var isFuerte=c.signal.indexOf('FUERTE')>=0;
    var barPct  =Math.round(Math.min(100,(c.ev/maxEV)*100));
    var barCol  =i===0?'#f59e0b':i<=2?'#fbbf24':'#a38540';
    var fl      =c.mercado==='MERVAL'?'🇦🇷':c.mercado==='BOVESPA'?'🇧🇷':'🇺🇸';
    var predCol =c.pred_21d!=null?(c.pred_21d>=0?'#4ade80':'#f87171'):'#666';
    var rrCol   =c.rr>=2.5?'#4ade80':c.rr>=1.5?'#fbbf24':'#f87171';
    var weekBadge=c.weekly&&c.weekly.indexOf('ALCISTA')>=0
      ? '<span style="background:rgba(74,222,128,.15);color:#4ade80;font-size:9px;padding:1px 6px;border-radius:10px">📈 Semanal Alcista</span>'
      : c.weekly&&c.weekly.indexOf('BAJISTA')>=0
      ? '<span style="background:rgba(248,113,113,.15);color:#f87171;font-size:9px;padding:1px 6px;border-radius:10px">📉 Semanal Bajista</span>'
      : '';
    var agreeBadge=c.pred_agree===true
      ? '<span style="background:rgba(74,222,128,.15);color:#4ade80;font-size:9px;padding:1px 6px;border-radius:10px">✅ Predictor confirma</span>'
      : c.pred_agree===false
      ? '<span style="background:rgba(251,191,36,.15);color:#fbbf24;font-size:9px;padding:1px 6px;border-radius:10px">⚠️ Predictor diverge</span>'
      : '';

    html+=
      '<div style="background:#16161e;border:1px solid #2a2a1a;border-left:3px solid '+barCol+';border-radius:10px;padding:14px 16px;margin-bottom:10px">'+

        // ── fila 1: rank + ticker + badges
        '<div style="display:flex;align-items:flex-start;gap:10px;margin-bottom:10px">'+
          '<div style="font-size:28px;font-weight:900;color:'+barCol+';min-width:36px;text-align:center;line-height:1">#'+(i+1)+'</div>'+
          '<div style="flex:1">'+
            '<div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:4px">'+
              '<span style="font-size:15px;font-weight:700;color:#fff">'+fl+' '+c.ticker+'</span>'+
              '<span style="font-size:11px;color:#777">'+c.empresa+'</span>'+
              (isFuerte
                ?'<span style="background:#14532d;color:#4ade80;font-size:9px;padding:1px 7px;border-radius:12px;font-weight:700">⭐ COMPRA FUERTE</span>'
                :'<span style="background:#052e16;color:#86efac;font-size:9px;padding:1px 7px;border-radius:12px">🟢 COMPRA</span>')+
              (c.enC?'<span style="background:rgba(251,191,36,.15);color:#fbbf24;font-size:9px;padding:1px 7px;border-radius:12px">📂 En cartera</span>':'')+
              weekBadge+agreeBadge+
            '</div>'+
            // barra de potencial
            '<div style="display:flex;align-items:center;gap:6px">'+
              '<span style="font-size:9px;color:#555;width:54px">Potencial</span>'+
              '<div style="flex:1;background:#1a1a1a;border-radius:3px;height:5px">'+
                '<div style="width:'+barPct+'%;background:'+barCol+';border-radius:3px;height:5px;transition:width .4s"></div>'+
              '</div>'+
              '<span style="font-size:10px;color:'+barCol+';font-weight:700;width:34px;text-align:right">'+barPct+'%</span>'+
            '</div>'+
          '</div>'+
        '</div>'+

        // ── fila 2: métricas en grid
        '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(90px,1fr));gap:6px;margin-bottom:10px">'+
          _cap_metric('Score V2',   c.sv2!=null?c.sv2.toFixed(0):'—',       '#fbbf24')+
          _cap_metric('R/R',        c.rr>0?c.rr.toFixed(2)+'x':'—',         rrCol)+
          (c.pred_21d!=null?_cap_metric('Pred.21d',(c.pred_21d>=0?'+':'')+c.pred_21d.toFixed(1)+'%', predCol):'')+
          (c.pred_5d!=null?_cap_metric('Pred.5d',(c.pred_5d>=0?'+':'')+c.pred_5d.toFixed(1)+'%',c.pred_5d>=0?'#4ade80':'#f87171'):'')+
          (c.pred_conf!=null?_cap_metric('Confianza',Math.round(c.pred_conf*100)+'%','#a78bfa'):'')+
          _cap_metric('Score Macro',c.score_mac!=null?c.score_mac.toFixed(0):'—','#5ba3ff')+
          _cap_metric('Score Téc.',c.score_tec!=null?c.score_tec.toFixed(0):'—','#22d3ee')+
          (c.score_fund!=null?_cap_metric('Fundamental',c.score_fund.toFixed(0),'#fb923c'):'')+
          (c.aq>0?_cap_metric('Asset Q.',c.aq.toFixed(0),'#e879f9'):'')+
          (c.es>0?_cap_metric('Entry Sc.',c.es.toFixed(0),'#34d399'):'')+
          (c.adx!=null?_cap_metric('ADX',c.adx.toFixed(1),c.adx>25?'#4ade80':'#888'):'')+
          (c.upside_g!=null?_cap_metric('Graham',(c.upside_g>=0?'+':'')+c.upside_g.toFixed(1)+'%',c.upside_g>0?'#4ade80':'#f87171'):'')+
        '</div>'+

        // ── fila 3: stop/target + horizonte
        '<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;font-size:11px">'+
          (c.atr_stop!=null?'<span style="background:rgba(251,146,60,.12);color:#fb923c;padding:2px 8px;border-radius:6px">🛡 Stop '+c.atr_stop.toLocaleString('es-AR')+'</span>':'')+
          (c.atr_target!=null?'<span style="background:rgba(74,222,128,.12);color:#4ade80;padding:2px 8px;border-radius:6px">🎯 Target '+c.atr_target.toLocaleString('es-AR')+'</span>':'')+
          (c.horizonte!=='—'?'<span style="background:rgba(91,163,255,.12);color:#5ba3ff;padding:2px 8px;border-radius:6px">⏱ '+c.horizonte+'</span>':'')+
          (c.pred_signal?'<span style="background:rgba(167,139,250,.12);color:#a78bfa;padding:2px 8px;border-radius:6px">'+c.pred_signal+'</span>':'')+
        '</div>'+

        // ── fila 4 (si en cartera): P&L actual
        (c.enC&&c.portData?
          '<div style="margin-top:8px;padding:8px 10px;background:rgba(251,191,36,.07);border-radius:6px;font-size:11px;display:flex;gap:14px;flex-wrap:wrap">'+
            '<span style="color:#888">📂 Posición actual:</span>'+
            '<span style="color:#e2e8f0">Cant: <b>'+c.portData.cantidad+'</b></span>'+
            (c.portData.pnl_pct!=null?'<span style="color:'+(c.portData.pnl_pct>=0?'#4ade80':'#f87171')+'">P&L: <b>'+(c.portData.pnl_pct>=0?'+':'')+c.portData.pnl_pct.toFixed(1)+'%</b></span>':'')+
            (c.portData.pnl_usd!=null?'<span style="color:'+(c.portData.pnl_usd>=0?'#4ade80':'#f87171')+'">→ <b>'+(c.portData.pnl_usd>=0?'+':'')+c.portData.pnl_usd.toFixed(0)+' USD</b></span>':'')+
          '</div>'
        :'')+

      '</div>';
  }});

  cb.innerHTML=html;

  function _cap_metric(lbl,val,col){{
    return '<div style="background:#0d0d0f;border-radius:6px;padding:5px 8px">'+
      '<div style="font-size:8px;color:#555;text-transform:uppercase;letter-spacing:.4px">'+lbl+'</div>'+
      '<div style="font-size:13px;font-weight:700;color:'+col+'">'+val+'</div>'+
    '</div>';
  }}
}})();

// ── OPORTUNIDADES DE COMPRA ────────────────────────────────────────────────
function showOpRank(){{
  document.getElementById('op-rank-page').style.display='block';
  document.getElementById('op-ficha-page').style.display='none';
  if(opChartInst){{try{{opChartInst.destroy();}}catch(e){{}} opChartInst=null;}}
}}
 
// Aplicar filtro triple sobre FICHAS usando opportunity_score
var fichasScored=FICHAS.map(function(f){{
  var opp=f.opportunity_score!=null?f.opportunity_score:computeOpportunityScore(f);
  return Object.assign({{}},f,{{opp_score:opp}});
}});
var allFichaScores=fichasScored.map(function(f){{return f.opp_score;}}).sort(function(a,b){{return a-b;}});
var fp75idx=Math.floor(allFichaScores.length*0.75);
var fp75=allFichaScores[fp75idx]||0;
var fichasFiltradas=fichasScored
  .filter(function(f){{
    var sig=f.signal_v2||f.signal||'';
    return sig.indexOf('COMPRA')>=0 && sig.indexOf('VENTA')<0 && f.opp_score>=50 && f.opp_score>=fp75;
  }})
  .sort(function(a,b){{return b.opp_score-a.opp_score;}});
var opRg=document.getElementById('op-rg');
if(fichasFiltradas.length===0){{
  opRg.innerHTML='<div style="text-align:center;padding:48px 20px;color:#555;font-size:14px">'+
    '<div style="font-size:36px;margin-bottom:12px">🔍</div>'+
    '<div style="font-weight:700;color:#777;margin-bottom:8px">No hay oportunidades con convicción suficiente en este momento</div>'+
    '<div style="font-size:12px;color:#444">El mercado no presenta señales que cumplan los tres criterios simultáneamente</div></div>';
}} else {{
  for(var i=0;i<fichasFiltradas.length;i++){{
    (function(f,idx){{
      var oppSc=f.opp_score||0;
      var sc2=oppSc>=80?'#22c55e':oppSc>=65?'#86efac':'#fbbf24';
      var row=document.createElement('div');
      row.className='op-rank-row';
      var num=document.createElement('div');
      num.className='op-num'+(idx<3?' gold':'');
      num.textContent='#'+(idx+1);
      row.appendChild(num);
      var main=document.createElement('div');
      main.className='op-main';
      main.innerHTML='<div class="op-ticker">'+f.flag+' '+f.ticker+'</div>'+
        '<div class="op-emp">'+f.empresa+' · '+f.market+'</div>'+
        '<div style="display:flex;align-items:center;gap:8px;margin-top:3px">'+
        '<div class="op-sbar" style="flex:1"><div class="op-sbarf" style="width:'+oppSc+'%;background:'+sc2+'"></div></div>'+
        '<span style="font-size:11px;font-weight:700;color:'+sc2+'">'+oppSc+'</span>'+
        '<span style="font-size:9px;color:#555">Opp.Score</span></div>';
      row.appendChild(main);
      var mets=document.createElement('div');
      mets.className='op-mets';
      mets.innerHTML=
        '<div class="op-m"><span class="op-mv">'+fn(f.precio)+'</span><span class="op-ml">'+f.moneda+'</span></div>'+
        '<div class="op-m"><span class="op-mv" style="color:'+rc(f.ret_anual)+'">'+fp(f.ret_anual)+'</span><span class="op-ml">12m</span></div>'+
        '<div class="op-m"><span class="op-mv" style="color:'+(f.rsi<40?'#4ade80':f.rsi>65?'#f87171':'#fbbf24')+'">'+fn(f.rsi,1)+'</span><span class="op-ml">RSI</span></div>'+
        '<div class="op-m"><span class="op-mv" style="color:#f87171">-'+fn(f.dist_max,1)+'%</span><span class="op-ml">vs Máx</span></div>'+
        '<div class="op-m"><span class="op-mv" style="color:#bc8cff">'+fn(f.rr,1)+'x</span><span class="op-ml">R/R</span></div>'+
        (f.pred_21d!=null?'<div class="op-m"><span class="op-mv" style="color:'+(f.pred_21d>=0?'#4ade80':'#f87171')+';font-weight:700">'+(f.pred_21d>=0?'+':'')+fn(f.pred_21d,1)+'%</span><span class="op-ml">📈 21d</span></div>':'')+
        (f.pred_confidence?'<div class="op-m"><span class="op-mv" style="color:#a78bfa">'+Math.round(f.pred_confidence*100)+'%</span><span class="op-ml">🎯 conf.</span></div>':'')+
        (f.suggested_pct!=null&&f.suggested_pct>0?'<div class="op-m"><span class="op-mv" style="color:#fbbf24;font-weight:800">'+f.suggested_pct.toFixed(1)+'%</span><span class="op-ml">💰 alloc.</span></div>':'')+
        (f.exit_score!=null?'<div class="op-m"><span class="op-mv" style="color:'+(f.exit_score>=56?'#f87171':f.exit_score>=31?'#fbbf24':'#4ade80')+'">'+f.exit_score.toFixed(0)+'</span><span class="op-ml">exit⚡</span></div>':'')+
        '<span class="op-sig '+(f.signal.indexOf('FUERTE')>=0?'op-sig-f':'op-sig-c')+'">'+(f.signal_v2||f.signal)+'</span>';
      row.appendChild(mets);
      row.onclick=function(){{showOpFicha(f.ticker);}};
      opRg.appendChild(row);
    }})(fichasFiltradas[i],i);
  }}
}}
 
function showOpFicha(ticker){{
  var f=null;
  for(var i=0;i<FICHAS.length;i++){{if(FICHAS[i].ticker===ticker){{f=FICHAS[i];break;}}}}
  if(!f) return;
  document.getElementById('op-rank-page').style.display='none';
  document.getElementById('op-ficha-page').style.display='block';
  if(opChartInst){{try{{opChartInst.destroy();}}catch(e){{}} opChartInst=null;}}
 
  var sc2=f.score_final>=65?'#d29922':'#4ade80';
  var res=f.resistencias||[], sup=f.soportes||[];
  var lvls='';
  for(var i=res.length-1;i>=0;i--){{
    var d=fn((res[i]-f.precio)/f.precio*100,1);
    lvls+='<div class="op-lvl"><span><span class="op-ltag op-lt-r">R</span>'+fn(res[i])+'</span><span style="color:#666;font-size:10px">+'+d+'%</span></div>';
  }}
  lvls+='<div class="op-lvl" style="background:#111115"><span><span class="op-ltag op-lt-e">ENTRADA</span>'+fn(f.entrada)+'</span><span style="color:#5ba3ff;font-size:10px">actual '+fn(f.precio)+'</span></div>';
  for(var i=0;i<sup.length;i++){{
    var d2=fn((f.precio-sup[i])/f.precio*100,1);
    lvls+='<div class="op-lvl"><span><span class="op-ltag op-lt-s">S</span>'+fn(sup[i])+'</span><span style="color:#666;font-size:10px">-'+d2+'%</span></div>';
  }}
  lvls+='<div class="op-lvl"><span><span class="op-ltag op-lt-st">STOP</span>'+fn(f.stop)+'</span><span style="color:#fb923c;font-size:10px">-'+fn(f.riesgo,1)+'%</span></div>';
  lvls+='<div class="op-lvl"><span><span class="op-ltag op-lt-tg">TARGET</span>'+fn(f.target)+'</span><span style="color:#bc8cff;font-size:10px">+'+fn(f.reward,1)+'%</span></div>';
 
  document.getElementById('op-fi').innerHTML=
    '<div class="op-ficha-hdr">'+
      '<div><div class="op-ftick">'+f.flag+' '+f.ticker+'</div><div class="op-femp">'+f.empresa+' · '+f.market+'</div></div>'+
      '<div class="op-fprice">'+fn(f.precio)+' <span style="font-size:11px;color:#666">'+f.moneda+'</span><br>'+
        '<span style="font-size:10px;color:#666">Máx12m: '+fn(f.max12m)+' ('+f.max_dt+') · Mín: '+fn(f.min12m)+' ('+f.min_dt+')</span></div>'+
    '</div>'+
    '<div class="op-card">'+
      '<h3>📈 Precio + MA20 + MA50 · 60 sesiones · Soportes/Resistencias</h3>'+
      (f.sin_grafico?
        '<div style="padding:20px;text-align:center;color:#555;font-size:12px;border:1px dashed #222230;border-radius:6px">📉 Serie de precios no disponible en datos locales — datos del modelo activos</div>'
        :'<div class="op-chart-wrap"><canvas id="opChartF"></canvas></div>')+
    '</div>'+
    '<div class="op-rrbox">'+
      '<div><div class="op-rrval" style="color:#5ba3ff">'+fn(f.entrada)+'</div><div class="op-rrlbl">Entrada</div></div>'+
      '<div><div class="op-rrval" style="color:#bc8cff">'+fn(f.rr_ratio!=null?f.rr_ratio:f.rr,2)+'x</div><div class="op-rrlbl">R/R</div></div>'+
      '<div><div class="op-rrval" style="color:#4ade80">+'+fn(f.reward,1)+'%</div><div class="op-rrlbl">Upside</div></div>'+
      (f.atr_stop!=null?'<div><div class="op-rrval" style="color:#fb923c">'+fn(f.atr_stop)+'</div><div class="op-rrlbl">ATR Stop</div></div>':'')+
      (f.atr_target!=null?'<div><div class="op-rrval" style="color:#a78bfa">'+fn(f.atr_target)+'</div><div class="op-rrlbl">ATR Target</div></div>':'')+
    '</div>'+
    '<div class="op-fgrid">'+
      '<div class="op-card"><h3>📐 Niveles operativos</h3>'+lvls+
        '<div style="margin-top:10px">'+
          '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:6px">'+
            '<div style="background:#111115;border:0.5px solid #1a1a2e;border-radius:6px;padding:6px 8px">'+
              '<div style="display:flex;align-items:center;gap:5px;margin-bottom:2px">'+
                '<div style="width:14px;height:2px;background:#fb923c;border-radius:1px"></div>'+
                '<span style="font-size:11px;font-weight:600;color:#e8e8ea;font-family:monospace">MA20: '+fn(f.ma20)+'</span>'+
              '</div>'+
              '<div style="font-size:9px;color:#555">Corto plazo · 4 sem</div>'+
              '<div style="font-size:10px;margin-top:2px;color:'+(f.ma20>f.ma50?'#4ade80':'#f87171')+'">'+
                (f.ma20>f.ma50?'↗ sobre MA50':'↘ bajo MA50')+
              '</div>'+
            '</div>'+
            '<div style="background:#111115;border:0.5px solid #1a1a2e;border-radius:6px;padding:6px 8px">'+
              '<div style="display:flex;align-items:center;gap:5px;margin-bottom:2px">'+
                '<div style="width:14px;height:2px;background:#bc8cff;border-radius:1px"></div>'+
                '<span style="font-size:11px;font-weight:600;color:#e8e8ea;font-family:monospace">MA50: '+fn(f.ma50)+'</span>'+
              '</div>'+
              '<div style="font-size:9px;color:#555">Mediano plazo · 2.5m</div>'+
              '<div style="font-size:10px;margin-top:2px;color:'+(f.precio>f.ma50?'#4ade80':'#f87171')+'">'+
                (f.precio>f.ma50?'↗ precio sobre MA50':'↘ precio bajo MA50')+
              '</div>'+
            '</div>'+
            (f.ma200?'<div style="background:#111115;border:0.5px solid #1a1a2e;border-radius:6px;padding:6px 8px">'+
              '<div style="display:flex;align-items:center;gap:5px;margin-bottom:2px">'+
                '<div style="width:14px;height:2px;background:#fbbf24;border-radius:1px"></div>'+
                '<span style="font-size:11px;font-weight:600;color:#e8e8ea;font-family:monospace">MA200: '+fn(f.ma200)+'</span>'+
              '</div>'+
              '<div style="font-size:9px;color:#555">Largo plazo · 10 meses</div>'+
              '<div style="font-size:10px;margin-top:2px;color:'+(f.precio>f.ma200?'#4ade80':'#f87171')+'">'+
                (f.precio>f.ma200?'↗ precio sobre MA200':'↘ precio bajo MA200')+
              '</div>'+
            '</div>':'<div></div>')+
          '</div>'+
          '<div style="padding:5px 8px;background:#111115;border-radius:4px;border-left:2px solid '+(f.ma_cross?'#4ade80':'#f87171')+'">'+
            '<span style="font-size:10px;color:#888">'+(f.ma_cross?'✅ MA20 > MA50 — momentum corto plazo positivo. ':'⚠️ MA20 < MA50 — presión bajista de corto plazo. ')+
            (f.ma200?(f.precio>f.ma200?'Precio sobre MA200 — tendencia larga alcista.':'Precio bajo MA200 — tendencia larga bajista.'):'MA200 no disponible.')+
            '</span>'+
          '</div>'+
        '</div></div>'+
      '<div class="op-card"><h3>🧮 Scoring Fase 2 (35M+35T+10S+20F)</h3>'+
        '<div class="op-sc-row"><span class="op-sc-name">Macro (35%)</span>'+
          '<div><span class="op-sc-val" style="color:#fbbf24">'+fn(f.score_macro,1)+'</span>'+
          '<div class="op-sc-bar"><div class="op-sc-fill" style="width:'+f.score_macro+'%;background:#fbbf24"></div></div></div></div>'+
        '<div class="op-sc-row"><span class="op-sc-name">Técnico (35%)</span>'+
          '<div><span class="op-sc-val" style="color:#5ba3ff">'+fn(f.score_tec,1)+'</span>'+
          '<div class="op-sc-bar"><div class="op-sc-fill" style="width:'+f.score_tec+'%;background:#5ba3ff"></div></div></div></div>'+
        '<div class="op-sc-row"><span class="op-sc-name">Fundamental (20%)</span>'+
          '<div><span class="op-sc-val" style="color:#bc8cff">'+fn(f.score_fund,1)+'</span>'+
          '<div class="op-sc-bar"><div class="op-sc-fill" style="width:'+f.score_fund+'%;background:#bc8cff"></div></div></div></div>'+
        '<div class="op-sc-row" style="border-bottom:none">'+
          '<span class="op-sc-name" style="font-weight:700;color:#fff">SCORE FINAL</span>'+
          '<div><span class="op-sc-val" style="font-size:18px;color:'+sc2+'">'+fn(f.score_final,1)+'</span>'+
          '<div class="op-sc-bar"><div class="op-sc-fill" style="width:'+f.score_final+'%;background:'+sc2+'"></div></div></div></div>'+
        '<div class="op-techbox">'+
          '<span>RSI <b style="color:'+(f.rsi<40?'#4ade80':f.rsi>65?'#f87171':'#fbbf24')+'">'+fn(f.rsi,1)+'</b></span>'+
          '<span>Mom <b style="color:'+rc(f.momentum)+'">'+fp(f.momentum)+'</b></span>'+
          '<span style="color:#666">-'+fn(f.dist_max,1)+'% del máx</span>'+
          '<span>12m <b style="color:'+rc(f.ret_anual)+'">'+fp(f.ret_anual)+'</b></span>'+
          (f.adx!=null?'<span>ADX <b style="color:'+(f.adx>25?'#4ade80':'#f87171')+'">'+fn(f.adx,1)+'</b></span>':'')+
          (f.stress_index!=null?'<span>Stress <b style="color:'+(f.stress_index>70?'#f87171':f.stress_index<40?'#4ade80':'#fbbf24')+'">'+fn(f.stress_index,0)+'</b></span>':'')+
          (f.atr_percentile!=null?'<span>ATR% <b>'+fn(f.atr_percentile,0)+'</b></span>':'')+
        '</div>'+
      '</div>'+
      ((f.pred_5d!=null||f.pred_21d!=null)?'<div class="op-card"><h3>🔭 Predicción Ensemble (5d / 21d)</h3>'+
        '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:8px">'+
          '<div style="background:#0d1117;border:1px solid #1a2030;border-radius:6px;padding:8px;text-align:center">'+
            '<div style="font-size:10px;color:#666;margin-bottom:3px">5 días</div>'+
            '<div style="font-size:16px;font-weight:700;color:'+(f.pred_5d!=null&&f.pred_5d>=0?'#4ade80':'#f87171')+'">'+
              (f.pred_5d!=null?(f.pred_5d>=0?'+':'')+fn(f.pred_5d,1)+'%':'—')+
            '</div>'+
          '</div>'+
          '<div style="background:#0d1117;border:1px solid #1a2030;border-radius:6px;padding:8px;text-align:center">'+
            '<div style="font-size:10px;color:#666;margin-bottom:3px">21 días</div>'+
            '<div style="font-size:16px;font-weight:700;color:'+(f.pred_21d!=null&&f.pred_21d>=0?'#4ade80':'#f87171')+'">'+
              (f.pred_21d!=null?(f.pred_21d>=0?'+':'')+fn(f.pred_21d,1)+'%':'—')+
            '</div>'+
          '</div>'+
          '<div style="background:#0d1117;border:1px solid #1a2030;border-radius:6px;padding:8px;text-align:center">'+
            '<div style="font-size:10px;color:#666;margin-bottom:3px">Confianza</div>'+
            '<div style="font-size:16px;font-weight:700;color:#a78bfa">'+
              (f.pred_confidence!=null?fn(f.pred_confidence*100,0)+'%':'—')+
            '</div>'+
          '</div>'+
        '</div>'+
        '<div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:6px">'+
          '<div style="font-size:13px;font-weight:700;color:#a78bfa">'+f.pred_signal+'</div>'+
          (f.pred_target?'<div style="font-size:12px;color:#666">Target: <b style="color:#bc8cff">'+fn(f.pred_target)+'</b></div>':'')+
          (f.pred_direction_agree?'<div style="font-size:11px;color:#4ade80">✅ Confirma señal del modelo</div>':'<div style="font-size:11px;color:#fbbf24">⚠️ Diverge del modelo</div>')+
        '</div>'+
      '</div>':'')+ 
    '</div>';

 
  setTimeout(function(){{
    if(typeof Chart==='undefined') return;
    var canvas=document.getElementById('opChartF');
    if(!canvas) return;
    var ctx=canvas.getContext('2d');
    var closes=f.closes60||[], labels=(f.dates60||[]).map(function(d){{return d.slice(5);}});
    var datasets=[{{
      label:f.ticker, data:closes,
      borderColor:'#5ba3ff', borderWidth:2, pointRadius:0, tension:.3,
      fill:true, backgroundColor:'rgba(91,163,255,0.06)',
    }}];
    if(f.ma20_line&&f.ma20_line.length) datasets.push({{
      label:'MA20', data:f.ma20_line,
      borderColor:'#fb923c', borderWidth:1.5, pointRadius:0, tension:.3, fill:false, borderDash:[4,3],
    }});
    if(f.ma50_line&&f.ma50_line.length) datasets.push({{
      label:'MA50', data:f.ma50_line,
      borderColor:'#bc8cff', borderWidth:1.5, pointRadius:0, tension:.3, fill:false, borderDash:[8,4],
    }});
    var lp={{id:'lp',afterDraw:function(ch){{
      var c2=ch.ctx,ya=ch.scales.y;
      var x0=ch.chartArea.left,x1=ch.chartArea.right,top=ch.chartArea.top,bot=ch.chartArea.bottom;
      function dl(val,color,lbl){{
        if(!val||isNaN(val))return;
        var y=ya.getPixelForValue(val);
        if(y<top||y>bot)return;
        c2.save();c2.strokeStyle=color;c2.lineWidth=1;c2.setLineDash([4,4]);
        c2.beginPath();c2.moveTo(x0,y);c2.lineTo(x1,y);c2.stroke();
        c2.fillStyle=color;c2.font='9px monospace';
        c2.fillText(lbl+' '+val.toFixed(2),x1-90,y-2);c2.restore();
      }}
      (f.resistencias||[]).forEach(function(r){{dl(r,'rgba(248,81,73,0.7)','R');}});
      (f.soportes||[]).forEach(function(s){{dl(s,'rgba(74,222,128,0.7)','S');}});
      dl(f.entrada,'rgba(91,163,255,0.9)','ENTRADA');
      dl(f.stop,'rgba(251,146,60,0.8)','STOP');
      dl(f.target,'rgba(188,140,255,0.8)','TARGET');
    }}}};
    opChartInst=new Chart(ctx,{{
      type:'line',
      data:{{labels:labels,datasets:datasets}},
      options:{{
        responsive:true,maintainAspectRatio:false,
        plugins:{{legend:{{display:true,labels:{{color:'#666',font:{{size:10}},boxWidth:16}}}},tooltip:{{mode:'index',intersect:false}}}},
        scales:{{
          x:{{ticks:{{color:'#666',font:{{size:9}},maxTicksLimit:10}},grid:{{color:'rgba(255,255,255,.04)'}}}},
          y:{{ticks:{{color:'#666',font:{{size:9}}}},grid:{{color:'rgba(255,255,255,.04)'}}}},
        }},
      }},
      plugins:[lp],
    }});
  }},80);
}}
// ── PORTFOLIO TAB ──
(function(){{
// ─── estado local ───────────────────────────────────────────────
var _portBarInst = null;
var _countdownVal = 60;
var _countdownTimer = null;

// ─── helpers ────────────────────────────────────────────────────
function _setTs(ok){{
  var el=document.getElementById('port-update-ts');
  if(!el) return;
  var now=new Date();
  var s=now.toLocaleDateString('es-AR',{{day:'2-digit',month:'2-digit',year:'numeric'}})+
        ' '+now.toLocaleTimeString('es-AR',{{hour:'2-digit',minute:'2-digit',second:'2-digit'}});
  el.textContent=s+(ok?' ✅':' ⚠️ fallback');
  el.style.color=ok?'#4ade80':'#fbbf24';
}}
function _setRefreshing(on){{
  var ind=document.getElementById('port-refresh-indicator');
  var btn=document.getElementById('btn-refresh-port');
  var ico=document.getElementById('refresh-icon');
  if(ind) ind.style.display=on?'block':'none';
  if(ico) ico.textContent=on?'⟳':'↻';
  if(btn) btn.style.background=on?'#0d2b1a':'#0d1f3c';
}}
function _resetCountdown(){{
  _countdownVal=60;
  var el=document.getElementById('port-countdown');
  if(el) el.textContent='Auto en 60s';
}}
function _tickCountdown(){{
  _countdownVal--;
  var el=document.getElementById('port-countdown');
  if(el) el.textContent=_countdownVal>0?'Auto en '+_countdownVal+'s':'Actualizando...';
}}

// ─── renderPortfolio ────────────────────────────────────────────
function renderPortfolio(portfolio, pAlerts, fromApi){{
  if(!portfolio || !portfolio.positions) return;
  pAlerts = pAlerts||{{}};
  var positions = portfolio.positions;
  var alertsList=(pAlerts&&pAlerts.alerts)?pAlerts.alerts:[];
  var pnlItems=alertsList.filter(function(a){{return a.tipo==='📊 P&L';}});

  // ── Totales ──
  var totalActualUsd=0, totalInicialUsd=0, totalRendUsd=0;
  positions.forEach(function(p){{
    totalActualUsd +=(p.valor_actual_usd||0);
    totalInicialUsd+=(p.valor_inicial_usd||0);
    totalRendUsd   +=(p.rend_usd||0);
  }});
  var pnlPct=totalInicialUsd>0?((totalRendUsd/totalInicialUsd)*100):0;
  var bestPos=positions.reduce(function(mx,p){{return(p.rend_usd||0)>(mx.rend_usd||0)?p:mx;}},positions[0]||{{}});
  var worstPos=positions.reduce(function(mn,p){{return(p.rend_usd||0)<(mn.rend_usd||0)?p:mn;}},positions[0]||{{}});
  var comprasAct=positions.filter(function(p){{
    var sd=SIGNALS.find(function(s){{return s.ticker===p.ticker||s.ticker===p.ticker.replace('.BA','').replace('.SA','');}});
    return sd&&(sd.signal_v2||sd.signal||'').indexOf('COMPRA')>=0;
  }}).length;

  // ── KPI Cards ──
  var sm=document.getElementById('portfolio-summary');
  if(sm) sm.innerHTML=
    _kpi('💵 Invertido','USD '+_fmt(totalInicialUsd,0),'Costo de compra','#5ba3ff')+
    _kpi('📈 Valor actual','USD '+_fmt(totalActualUsd,0),'Precio de mercado','#e2e8f0')+
    _kpi('💰 P&L Total',(totalRendUsd>=0?'+':'')+_fmt(totalRendUsd,0)+' USD','('+(pnlPct>=0?'+':'')+pnlPct.toFixed(1)+'%)',totalRendUsd>=0?'#4ade80':'#f87171')+
    _kpi('📋 Posiciones',positions.length,''+comprasAct+' con señal compra','#e2e8f0')+
    _kpi('🏆 Mejor pos.',bestPos.ticker||'—',(bestPos.rend_usd>=0?'+':'')+_fmt(bestPos.rend_usd||0,0)+' USD','#4ade80')+
    _kpi('⚠️ Revisar',worstPos.ticker||'—',(worstPos.rend_usd>=0?'+':'')+_fmt(worstPos.rend_usd||0,0)+' USD','#f87171');

  // ── Ordenar ──
  var _acOrd={{'🔴 VENDER':0,'🔴 STOP (señal positiva, evaluar recompra)':1,'🟡 REDUCIR':2,'⭐ AGREGAR':3,'🟢 HOLD':4,'⚠️ Sin precio':5,'⏰ TIME STOP':4}};
  var _sgOrd={{'⭐ COMPRA FUERTE':0,'🟢 COMPRA':1,'🟡 NEUTRAL/ESPERAR':2,'🟠 VENTA PARCIAL':3,'🔴 VENTA':4}};
  var posSorted=positions.slice().sort(function(a,b){{
    var pnlA=pnlItems.find(function(x){{return x.ticker===a.ticker;}}),pnlB=pnlItems.find(function(x){{return x.ticker===b.ticker;}});
    var accA=pnlA?(pnlA.accion||'🟢 HOLD'):'🟢 HOLD',accB=pnlB?(pnlB.accion||'🟢 HOLD'):'🟢 HOLD';
    var oA=_acOrd[accA]!=null?_acOrd[accA]:4,oB=_acOrd[accB]!=null?_acOrd[accB]:4;
    if(oA!==oB) return oA-oB;
    var sdA=SIGNALS.find(function(s){{return s.ticker===a.ticker||s.ticker===a.ticker.replace('.BA','').replace('.SA','');}}),
        sdB=SIGNALS.find(function(s){{return s.ticker===b.ticker||s.ticker===b.ticker.replace('.BA','').replace('.SA','');}});
    var sig2A=sdA?(sdA.signal_v2||sdA.signal||''):'',sig2B=sdB?(sdB.signal_v2||sdB.signal||''):'';
    var sA=_sgOrd[sig2A]!=null?_sgOrd[sig2A]:3,sB=_sgOrd[sig2B]!=null?_sgOrd[sig2B]:3;
    if(sA!==sB) return sA-sB;
    return (b.rend_usd||0)-(a.rend_usd||0);
  }});

  // ── Tabla ──
  var tb=document.getElementById('portfolio-table');
  if(tb){{
    tb.innerHTML=
      '<tr>'+
        '<th>Ticker</th><th>Mercado</th>'+
        '<th>Compra USD</th><th>Actual USD</th>'+
        '<th>Cant</th>'+
        '<th>P&L %</th><th>P&L USD</th>'+
        '<th>Señal V2</th><th>Pred.21d</th><th>R/R</th><th>Acción</th>'+
      '</tr>'+
      posSorted.map(function(p){{
        var pnl=pnlItems.find(function(a){{return a.ticker===p.ticker;}});
        var sd=SIGNALS.find(function(s){{return s.ticker===p.ticker||s.ticker===p.ticker.replace('.BA','').replace('.SA','')||s.ticker===p.ticker+'.SA';}});
        var pCompUsd=p.precio_compra_usd||0,pActUsd=p.precio_actual_usd||0;
        var pnlPctPos=pCompUsd>0&&pActUsd>0?((pActUsd/pCompUsd-1)*100):null;
        var pnlUsdPos=p.rend_usd!=null?p.rend_usd:null;
        var sig2=pnl?(pnl.signal_v2||'—'):(sd?(sd.signal_v2||sd.signal||'—'):'⚠️ Sin datos');
        var pred21=sd&&sd.pred_21d!=null?sd.pred_21d:null;
        var rr=pnl?(pnl.rr_ratio||0):(sd?(sd.rr_ratio||0):0);
        var acc=pnl?(pnl.accion||'🟢 HOLD'):(sd?'🟢 HOLD':'⚠️ Sin precio');
        var accColor=acc.indexOf('VENDER')>=0?'#f87171':acc.indexOf('REDUCIR')>=0?'#fbbf24':acc.indexOf('PARCIAL')>=0?'#c084fc':acc.indexOf('AGREGAR')>=0||acc.indexOf('HOLD')>=0?'#4ade80':'#aaa';
        var fl=p.mercado==='MERVAL'?'🇦🇷':p.mercado==='BOVESPA'?'🇧🇷':'🇺🇸';
        var rowBg=acc.indexOf('VENDER')>=0?'background:rgba(248,113,113,.05);':'';
        return '<tr style="'+rowBg+'">'+
          '<td class="ticker" style="font-weight:700">'+p.ticker+'</td>'+
          '<td style="color:#888;font-size:11px">'+fl+' '+p.mercado+'</td>'+
          '<td style="color:#888">'+( pCompUsd>0?'$'+pCompUsd.toFixed(4):'—')+'</td>'+
          '<td style="color:'+(pActUsd>0?'#e2e8f0':'#555')+';font-weight:600">'+(pActUsd>0?'$'+pActUsd.toFixed(4):'—')+'</td>'+
          '<td>'+p.cantidad+'</td>'+
          '<td style="color:'+(pnlPctPos===null?'#555':pnlPctPos>=0?'#4ade80':'#f87171')+';font-weight:600">'+(pnlPctPos!==null?(pnlPctPos>=0?'+':'')+pnlPctPos.toFixed(1)+'%':'—')+'</td>'+
          '<td style="color:'+(pnlUsdPos===null?'#555':pnlUsdPos>=0?'#4ade80':'#f87171')+';font-weight:700">'+(pnlUsdPos!==null?(pnlUsdPos>=0?'+':'')+pnlUsdPos.toFixed(0)+' USD':'—')+'</td>'+
          '<td style="color:'+sigColor(sig2)+';font-size:11px">'+sig2+'</td>'+
          '<td style="color:'+(pred21===null?'#555':pred21>=0?'#4ade80':'#f87171')+';font-weight:600">'+(pred21!==null?(pred21>=0?'+':'')+pred21.toFixed(1)+'%':'—')+'</td>'+
          '<td style="color:#bc8cff;font-weight:600">'+(rr>0?rr.toFixed(2)+'x':'—')+'</td>'+
          '<td style="color:'+accColor+';font-weight:700;font-size:11px">'+acc+'</td>'+
        '</tr>';
      }}).join('');
  }}

  // ── Bar chart P&L por posición ──
  var barCtx=document.getElementById('chartPortBar');
  if(barCtx){{
    if(_portBarInst){{try{{_portBarInst.destroy();}}catch(e){{}}}}
    var labelsBar=posSorted.map(function(p){{return p.ticker;}});
    var dataBar  =posSorted.map(function(p){{return p.rend_usd!=null?parseFloat(p.rend_usd.toFixed(0)):0;}});
    var colorsBar=dataBar.map(function(v){{return v>=0?'rgba(74,222,128,.7)':'rgba(248,113,113,.7)';}});
    _portBarInst=new Chart(barCtx,{{
      type:'bar',
      data:{{labels:labelsBar,datasets:[{{label:'P&L USD',data:dataBar,backgroundColor:colorsBar,borderRadius:4}}]}},
      options:{{
        responsive:true,maintainAspectRatio:false,
        plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:function(ctx){{return(ctx.raw>=0?'+':'')+ctx.raw.toFixed(0)+' USD';}}}}}}}},
        scales:{{
          x:{{ticks:{{color:'#888',font:{{size:10}}}},grid:{{color:'rgba(255,255,255,.04)'}}}},
          y:{{ticks:{{color:'#888',font:{{size:10}},callback:function(v){{return(v>=0?'+':'')+v+' USD';}}}},grid:{{color:'rgba(255,255,255,.06)'}}}}
        }}
      }}
    }});
  }}

  // ── Timestamp ──
  _setTs(fromApi!==false);

  // ── Alertas ──
  var criticas=alertsList.filter(function(a){{return a.tipo!=='📊 P&L';}});
  var al=document.getElementById('portfolio-alerts');
  if(al){{
    if(criticas.length===0) al.innerHTML='<div style="color:#4ade80;padding:8px">✅ Sin alertas activas</div>';
    else al.innerHTML=criticas.map(function(a){{
      return '<div style="padding:6px 12px;border-left:3px solid #fbbf24;margin-bottom:4px;background:rgba(251,191,36,.05);border-radius:0 6px 6px 0">'+
             '<b>'+a.tipo+'</b> '+a.mensaje+'</div>';
    }}).join('');
  }}
}}  // end renderPortfolio

// ─── helper KPI card ────────────────────────────────────────────
function _kpi(title, value, sub, color){{
  return '<div class="card" style="min-width:0">'+
    '<div class="card-title" style="font-size:10px">'+title+'</div>'+
    '<div class="card-value" style="color:'+color+';font-size:16px;font-weight:800">'+value+'</div>'+
    (sub?'<div class="card-sub" style="font-size:10px">'+sub+'</div>':'')+
  '</div>';
}}
function _fmt(v,d){{ return Number(v).toLocaleString('es-AR',{{minimumFractionDigits:d,maximumFractionDigits:d}}); }}

// ─── fetch + render ─────────────────────────────────────────────
function _doFetch(){{
  _setRefreshing(true);
  if(!RAILWAY_API_URL){{ renderPortfolio(PORTFOLIO,PORTFOLIO_ALERTS,false); _setRefreshing(false); return; }}
  fetch(RAILWAY_API_URL+'/api/portfolio',{{method:'GET',cache:'no-store'}})
    .then(function(r){{return r.ok?r.json():null;}})
    .then(function(d){{
      _setRefreshing(false);
      if(d&&d.positions) renderPortfolio(d,PORTFOLIO_ALERTS,true);
      else renderPortfolio(PORTFOLIO,PORTFOLIO_ALERTS,false);
    }})
    .catch(function(){{
      _setRefreshing(false);
      renderPortfolio(PORTFOLIO,PORTFOLIO_ALERTS,false);
    }});
}}

// ─── exponer globalmente (botón HTML la llama) ───────────────────
window.portRefreshNow = function(){{
  _resetCountdown();
  _doFetch();
}};

// ─── arranque + contador regresivo ──────────────────────────────
_doFetch();
_countdownTimer=setInterval(function(){{
  var pg=document.getElementById('portfolio');
  if(!pg||!pg.classList.contains('on')) return;
  _tickCountdown();
  if(_countdownVal<=0){{ _resetCountdown(); _doFetch(); }}
}},1000);
}})();
// ── SISTEMA DE OPERACIONES ────────────────────────────────────────────────────
var OP_KEY = 'inv_operaciones_v1';
var opTickerValido = null;

// Índice de tickers válidos: solo los que están en SIGNALS
var tickersModelo = {{}};
SIGNALS.forEach(function(s){{
  tickersModelo[s.ticker] = {{
    empresa: s.empresa,
    mercado: s.mercado,
    signal: s.signal_v2 || s.signal,
    precio_actual: s.precio_actual,
  }};
}});

function tickerAutocomplete(val){{
  var drop = document.getElementById('op-ticker-drop');
  var status = document.getElementById('op-ticker-status');
  opTickerValido = null;
  val = val.trim().toUpperCase();
  if(!val){{ drop.style.display='none'; status.textContent='Ingresá un ticker del modelo'; status.style.color='#555'; return; }}
  var matches = Object.keys(tickersModelo).filter(function(t){{
    return t.toUpperCase().indexOf(val)>=0 || tickersModelo[t].empresa.toUpperCase().indexOf(val)>=0;
  }}).slice(0,8);
  if(matches.length===0){{
    drop.style.display='none';
    status.textContent='⚠️ Ticker no encontrado en el modelo'; status.style.color='#f87171';
    return;
  }}
  var sigColor2=function(sg){{return sg.indexOf('COMPRA FUERTE')>=0?'#ffd700':sg.indexOf('COMPRA')>=0?'#4ade80':sg.indexOf('NEUTRAL')>=0?'#fbbf24':sg.indexOf('VENTA')>=0?'#fb923c':'#888';}};
  drop.innerHTML = matches.map(function(t){{
    var d=tickersModelo[t];
    var sc=sigColor2(d.signal);
    return '<div class="op-ticker-item" onclick="seleccionarTicker(&#39;'+t+'&#39;)">'
      +'<div><b style="color:#5ba3ff">'+t+'</b> <span style="color:#666;font-size:10px">'+d.empresa.substring(0,28)+'</span></div>'
      +'<span class="oti-sig" style="color:'+sc+';border:1px solid '+sc+'20;background:'+sc+'10">'+d.signal+'</span>'
      +'</div>';
  }}).join('');
  drop.style.display='block';
  // Check exact match
  if(tickersModelo[val]){{
    opTickerValido=val;
    var d=tickersModelo[val];
    var mf=d.mercado==='MERVAL'?'🇦🇷':d.mercado==='BOVESPA'?'🇧🇷':'🇺🇸';
    status.textContent='✅ '+d.empresa+' · '+mf+' '+d.mercado+' · Precio actual: '+d.precio_actual.toLocaleString('es-AR');
    status.style.color='#4ade80';
    document.getElementById('op-submit-btn').disabled=false;
    document.getElementById('op-submit-btn').style.opacity='1';
  }} else{{
    // También habilitar si el valor escrito coincide exactamente (mayúsculas)
    var valUpper = val.toUpperCase();
    var exactMatch = null;
    Object.keys(tickersModelo).forEach(function(tk){{
      if(tk.toUpperCase() === valUpper) exactMatch = tk;
    }});
    if(exactMatch){{
      seleccionarTicker(exactMatch);
      drop.style.display='none';
    }} else {{
      opTickerValido=null;
      status.textContent='⚠️ Seleccioná un ticker válido del modelo';
      status.style.color='#fbbf24';
      document.getElementById('op-submit-btn').disabled=true;
      document.getElementById('op-submit-btn').style.opacity='0.4';
    }}
  }}
}}

function seleccionarTicker(t){{
  document.getElementById('op-ticker-input').value=t;
  document.getElementById('op-ticker-drop').style.display='none';
  opTickerValido=t;
  var d=tickersModelo[t];
  // Auto-sugerir precio actual en USD desde pipeline
  var precioSug = d.precio_actual_usd || 0;
  if(precioSug > 0){{
    document.getElementById('op-precio').value = precioSug.toFixed(4);
    document.getElementById('op-precio').style.borderColor='#4ade80';
    setTimeout(function(){{document.getElementById('op-precio').style.borderColor='#333';}}, 2000);
  }}
  // Mostrar hint de precio sugerido
  var hint = document.getElementById('op-precio-hint');
  if(hint && precioSug > 0) hint.textContent = '(sugerido: USD '+precioSug.toFixed(4)+')';
  else if(hint) hint.textContent = '';
  var status=document.getElementById('op-ticker-status');
  status.textContent='✅ '+d.empresa+' · '+d.mercado+' · Precio actual: '+d.precio_actual.toLocaleString('es-AR');
  status.style.color='#4ade80';
}}

function registrarOperacion(){{
  var msg=document.getElementById('op-form-msg');
  msg.style.display='block';
  if(!opTickerValido){{ msg.textContent='⚠️ Seleccioná un ticker válido del modelo.'; msg.style.color='#f87171'; return; }}
  var precio=parseFloat(document.getElementById('op-precio').value);
  var cantidad=parseInt(document.getElementById('op-cantidad').value);
  var tipo=document.getElementById('op-tipo').value;
  if(!precio||precio<=0){{ msg.textContent='⚠️ Ingresá un precio unitario válido.'; msg.style.color='#f87171'; return; }}
  if(!cantidad||cantidad<=0){{ msg.textContent='⚠️ Ingresá una cantidad válida.'; msg.style.color='#f87171'; return; }}
  var d=tickersModelo[opTickerValido];
  var op={{
    id: Date.now(),
    fecha: new Date().toISOString().split('T')[0],
    tipo: tipo,
    ticker: opTickerValido,
    empresa: d.empresa,
    mercado: d.mercado,
    precio: precio,
    cantidad: cantidad,
    total: Math.round(precio*cantidad*100)/100,
    signal: d.signal,
  }};
  // ── POST al servidor Railway ──
  // Inferir precio_fuente desde el ticker (no depender del select op-instrumento)
  var instrumento = opTickerValido.endsWith('.BA') ? 'MERVAL_CSV'
                  : opTickerValido.endsWith('.SA') ? 'BOVESPA_CSV'
                  : 'SP500_CSV';
  var ratioCedear = 1.0;  // precio ingresado ya es USD broker
  var endpoint = RAILWAY_API_URL + (tipo === 'COMPRA' ? '/api/compra' : '/api/venta');
  msg.textContent='⏳ Enviando al servidor…'; msg.style.color='#fbbf24';
  document.getElementById('op-submit-btn').disabled=true;
  fetch(endpoint, {{
    method:'POST',
    headers:{{'Content-Type':'application/json'}},
    body: JSON.stringify({{
      ticker: op.ticker,
      nombre: op.empresa,
      mercado: op.mercado,
      precio: op.precio,
      cantidad: op.cantidad,
      total_usd: op.total,
      signal: op.signal,
      precio_fuente: instrumento,
      ratio_cedear: ratioCedear
    }})
  }})
  .then(function(r){{ return r.json(); }})
  .then(function(res){{
    if(res.error){{
      msg.textContent='❌ Error: '+res.error; msg.style.color='#f87171';
    }} else {{
      // Guardar en historial local también
      var ops=cargarOperaciones();
      ops.unshift(op);
      localStorage.setItem(OP_KEY, JSON.stringify(ops));
      var okMsg = res.msg || res.message || 'Operación registrada';
      if(res.push_warning){{
        // Push a GitHub falló — mostrar advertencia clara
        msg.innerHTML = '✅ '+okMsg+'<br><span style="color:#fbbf24;font-size:11px">⚠️ No se sincronizó con GitHub: '+res.push_warning+'<br>Los datos están guardados en Railway pero pueden perderse en un redeploy. Ejecutá /run desde Telegram para forzar sincronización.</span>';
        msg.style.color='#e2e8f0';
        setTimeout(function(){{ msg.style.display='none'; if(typeof portRefreshNow==='function') portRefreshNow(); }}, 8000);
      }} else {{
        msg.textContent='✅ '+okMsg+' — sincronizado ✓'; msg.style.color='#4ade80';
        setTimeout(function(){{ msg.style.display='none'; if(typeof portRefreshNow==='function') portRefreshNow(); }}, 1200);
      }}
    }}
  }})
  .catch(function(err){{
    msg.textContent='⚠️ Sin conexión al servidor — guardado local.'; msg.style.color='#fbbf24';
    // Fallback: guardar localmente si el servidor no responde
    var ops=cargarOperaciones();
    ops.unshift(op);
    localStorage.setItem(OP_KEY, JSON.stringify(ops));
    setTimeout(function(){{msg.style.display='none';}},3000);
  }})
  .finally(function(){{
    // Reset form
    document.getElementById('op-ticker-input').value='';
    document.getElementById('op-precio').value='';
    document.getElementById('op-cantidad').value='';
    document.getElementById('op-ticker-status').textContent='Ingresá un ticker del modelo';
    document.getElementById('op-ticker-status').style.color='#555';
    opTickerValido=null;
    document.getElementById('op-submit-btn').disabled=false;
    document.getElementById('op-submit-btn').style.opacity='1';
    renderHistorial();
  }});
}}

function onInstrumentoChange(){{
  // No-op: ratio CEDEAR eliminado del formulario
}}

function cargarOperaciones(){{
  try{{ return JSON.parse(localStorage.getItem(OP_KEY)||'[]'); }}catch(e){{return[];}}
}}

function eliminarOperacion(id){{
  var ops=cargarOperaciones().filter(function(o){{return o.id!==id;}});
  localStorage.setItem(OP_KEY,JSON.stringify(ops));
  renderHistorial();
}}

function renderHistorial(){{
  var ops=cargarOperaciones();
  var empty=document.getElementById('op-historial-empty');
  var tb=document.getElementById('op-historial-table');
  if(!ops.length){{ if(empty)empty.style.display='block'; if(tb)tb.style.display='none'; return; }}
  if(empty)empty.style.display='none';
  if(tb){{
    tb.style.display='';
    var sigColor2=function(sg){{return sg.indexOf('COMPRA FUERTE')>=0?'#ffd700':sg.indexOf('COMPRA')>=0?'#4ade80':sg.indexOf('NEUTRAL')>=0?'#fbbf24':sg.indexOf('VENTA')>=0?'#fb923c':'#888';}};
    var current = tickersModelo;
    tb.innerHTML='<tr><th>Fecha</th><th>Tipo</th><th>Ticker</th><th>Mercado</th><th>Señal actual</th><th>Precio unit.</th><th>Cantidad</th><th>Total</th><th></th></tr>'
      +ops.map(function(o){{
        var sigNow=(current[o.ticker]&&current[o.ticker].signal)||o.signal;
        var sc=sigColor2(sigNow);
        var tipoColor=o.tipo==='COMPRA'?'#4ade80':'#f87171';
        return '<tr>'
          +'<td style="color:#666">'+o.fecha+'</td>'
          +'<td style="color:'+tipoColor+';font-weight:700">'+o.tipo+'</td>'
          +'<td class="ticker">'+o.ticker+'</td>'
          +'<td style="color:#666">'+o.mercado+'</td>'
          +'<td style="color:'+sc+'">'+sigNow+'</td>'
          +'<td>'+o.precio.toLocaleString('es-AR')+'</td>'
          +'<td>'+o.cantidad.toLocaleString('es-AR')+'</td>'
          +'<td style="font-weight:600">'+o.total.toLocaleString('es-AR')+'</td>'
          +'<td><button onclick="eliminarOperacion('+o.id+')" style="background:none;border:1px solid #333;color:#666;padding:2px 8px;border-radius:4px;cursor:pointer;font-size:10px">✕</button></td>'
          +'</tr>';
      }}).join('');
  }}
}}

// Inicializar historial al cargar
renderHistorial();

}} catch(err) {{ document.body.insertAdjacentHTML('afterbegin','<div style="background:red;color:white;padding:20px;font-size:16px;z-index:9999;position:fixed;top:0;left:0;right:0">ERROR JS: '+err.message+'</div>'); }}
</script>
</body>
</html>"""
 
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path
 
 
# ─────────────────────────────────────────────
# Excel de fichas
# ─────────────────────────────────────────────
 
def generate_excel(signals: list[dict], index_stats: dict, output_path: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        logger.warning("openpyxl no disponible, saltando Excel")
        return ""
 
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
 
    HDR_FILL  = PatternFill("solid", fgColor="1e2a3a")
    BUY_FILL  = PatternFill("solid", fgColor="0d2b1a")
    SELL_FILL = PatternFill("solid", fgColor="2b1010")
    NEU_FILL  = PatternFill("solid", fgColor="1a1a2e")
    HDR_FONT  = Font(bold=True, color="5ba3ff", size=11)
    WHITE     = Font(color="e8e8ea", size=10)
    GREEN     = Font(color="4ade80", bold=True, size=10)
    RED       = Font(color="f87171", bold=True, size=10)
    ORANGE    = Font(color="fb923c", bold=True, size=10)
 
    headers = ["Ticker","Empresa","Sector","Precio","Sem%","Mes%","Anual%",
               "RSI","Macro","Técnico","Score","Señal","MA>50","Máx 12m","Mín 12m"]
 
    def _write_sheet(ws_name, market):
        ws = wb.create_sheet(ws_name)
        ws.freeze_panes = "A2"
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
        rows = [s for s in signals if s["mercado"] == market]
        for r, s in enumerate(rows, 2):
            row_data = [s["ticker"],s["empresa"],s["sector"],s["precio_actual"],
                        s["ret_sem"],s["ret_mes"],s["ret_anual"],s["rsi"],
                        s["score_macro"],s["score_tecnico"],s["score_final"],s["signal"],
                        "Sí" if s.get("ma_cross") else "No",s["max_12m"],s["min_12m"]]
            sig  = s["signal"]
            fill = BUY_FILL if "COMPRA" in sig else SELL_FILL if "VENTA" in sig else NEU_FILL
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill; cell.font = WHITE
                cell.alignment = Alignment(horizontal="center" if col > 3 else "left")
                if col in (5,6,7) and isinstance(val,(int,float)):
                    cell.font = GREEN if val >= 0 else RED
                if col == 12:
                    if "COMPRA" in str(val): cell.font = GREEN
                    elif "VENTA" in str(val): cell.font = ORANGE
        widths = [12,28,14,14,8,8,8,7,8,8,8,22,7,14,14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i)].width = w
 
    _write_sheet("MERVAL","MERVAL")
    _write_sheet("BOVESPA","BOVESPA")
    _write_sheet("SP500","SP500")
 
    ws_rank = wb.create_sheet("Ranking Global")
    ws_rank.freeze_panes = "A2"
    headers_ext = ["Mercado"] + headers
    from openpyxl.styles import Font as F2, Alignment as A2
    for col, h in enumerate(headers_ext, 1):
        cell = ws_rank.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = A2(horizontal="center")
    top = sorted(signals, key=lambda x: x["score_final"], reverse=True)[:30]
    for r, s in enumerate(top, 2):
        sig  = s["signal"]
        fill = BUY_FILL if "COMPRA" in sig else SELL_FILL if "VENTA" in sig else NEU_FILL
        vals = [s["mercado"],s["ticker"],s["empresa"],s["sector"],
                s["precio_actual"],s["ret_sem"],s["ret_mes"],s["ret_anual"],
                s["rsi"],s["score_macro"],s["score_tecnico"],s["score_final"],
                s["signal"],"Sí" if s.get("ma_cross") else "No",s["max_12m"],s["min_12m"]]
        for col, val in enumerate(vals, 1):
            cell = ws_rank.cell(row=r, column=col, value=val)
            cell.fill = fill; cell.font = WHITE
            cell.alignment = A2(horizontal="center" if col > 4 else "left")
            if col in (6,7,8) and isinstance(val,(int,float)):
                cell.font = GREEN if val >= 0 else RED
 
    wb.save(output_path)
    logger.info(f"Excel guardado: {output_path}")
    return output_path
 
 
