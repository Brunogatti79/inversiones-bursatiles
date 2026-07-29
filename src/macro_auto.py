"""
src/macro_auto.py — Carga automática de datos macro desde APIs públicas
 
APIs utilizadas:
- FRED (USA): Fed Funds, CPI, desempleo, GDP, confianza, PCE Core, ISM
- BCRA (Argentina): tasa, reservas, tipo de cambio
- BCB/SGS (Brasil): SELIC, IPCA, desempleo, reservas, BRL/USD
- Yahoo Finance: DXY, HY spread (via tickers)
- Ámbito/web: riesgo país ARG/BRA (EMBI)
 
Retorna dict compatible con macro_loader.py y analyzer.py
"""
 
import os
import logging
import json
import base64
import requests
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
 
logger = logging.getLogger(__name__)
 
FRED_API_KEY = os.getenv("FRED_API_KEY", "")
CACHE_PATH = "data/macro_auto_cache.json"
CCL_CACHE_PATH = "data/ccl_cache.json"
LAST_KNOWN_PATH = "data/macro_last_known.json"
DATA_QUALITY_HISTORY_PATH = "data/data_quality_history.json"
 
# ─────────────────────────────────────────────
# FRED (USA) — Federal Reserve Economic Data
# ─────────────────────────────────────────────
 
 
def _fred_latest(series_id, api_key=None):
    """Obtiene el último valor de una serie FRED."""
    key = api_key or FRED_API_KEY
    if not key:
        return None, None
    try:
        url = f"https://api.stlouisfed.org/fred/series/observations"
        params = {
            "series_id": series_id,
            "api_key": key,
            "file_type": "json",
            "sort_order": "desc",
            "limit": 5,
        }
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        obs = r.json().get("observations", [])
        for o in obs:
            val = o.get("value", ".")
            if val != ".":
                return float(val), o.get("date", "")
        return None, None
    except Exception as e:
        logger.warning(f"FRED error [{series_id}]: {e}")
        return None, None
 
 
def _fred_yoy(series_id, api_key=None):
    """Calcula YoY % change para series como CPI, PCE."""
    key = api_key or FRED_API_KEY
    if not key:
        return None
    try:
        url = f"https://api.stlouisfed.org/fred/series/observations"
        params = {
            "series_id": series_id,
            "api_key": key,
            "file_type": "json",
            "sort_order": "desc",
            "limit": 15,
            "units": "pc1",  # percent change from year ago
        }
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        obs = r.json().get("observations", [])
        for o in obs:
            val = o.get("value", ".")
            if val != ".":
                return round(float(val), 2)
        return None
    except Exception as e:
        logger.warning(f"FRED YoY error [{series_id}]: {e}")
        return None
 
 
def fetch_usa_macro():
    """Obtiene todas las variables macro de USA desde FRED."""
    data = {}
 
    # Fed Funds Rate
    val, dt = _fred_latest("FEDFUNDS")
    data["fed_funds"] = {"valor": val, "fecha": dt}
 
    # CPI YoY
    cpi_yoy = _fred_yoy("CPIAUCSL")
    data["cpi"] = {"valor": cpi_yoy, "fecha": ""}
 
    # Unemployment
    val, dt = _fred_latest("UNRATE")
    data["unemployment"] = {"valor": val, "fecha": dt}
 
    # GDP Growth
    val, dt = _fred_latest("A191RL1Q225SBEA")
    data["gdp_growth"] = {"valor": val, "fecha": dt}
 
    # Consumer Confidence (U. Michigan)
    val, dt = _fred_latest("UMCSENT")
    data["consumer_conf"] = {"valor": val, "fecha": dt}
 
    # PCE Core YoY
    pce_yoy = _fred_yoy("PCEPILFE")
    data["pce_core"] = {"valor": pce_yoy, "fecha": ""}
 
    # ISM Manufacturing — si NAPMPI no está disponible se deja en None.
    # (Se eliminó el fallback a MANEMP: es nivel de empleo manufacturero en
    # miles de personas, unidad incompatible con el rango de normalización
    # 40-62 calibrado para un índice PMI 0-100; usarlo distorsionaba el score.)
    val, dt = _fred_latest("NAPMPI")
    data["ism_mfg"] = {"valor": val, "fecha": dt}
 
    # DXY
    val, dt = _fred_latest("DTWEXBGS")
    data["dxy"] = {"valor": val, "fecha": dt}
 
    # HY Spread
    val, dt = _fred_latest("BAMLH0A0HYM2")
    data["hy_spread"] = {"valor": val, "fecha": dt}
 
    logger.info(f"USA macro: {sum(1 for v in data.values() if v['valor'] is not None)}/{len(data)} variables obtenidas")
    return data
 
 
# ─────────────────────────────────────────────
# BCRA (Argentina) — API pública
# ─────────────────────────────────────────────
 
# API alternativa estadisticasbcra.com (no requiere auth, API oficial v2 deprecada)
BCRA_ENDPOINTS = {
    "tasa_plazo_fijo": "https://api.estadisticasbcra.com/tasa_depositos_30_dias",
    "reservas":        "https://api.estadisticasbcra.com/reservas",
    "tipo_cambio":     "https://api.estadisticasbcra.com/usd_of",
}
 
 
def _bcra_latest(url):
    """Obtiene el último valor desde api.estadisticasbcra.com."""
    try:
        headers = {"Authorization": "BEARER eyJhbGciOiJIUzUxMiIsInR5cCI6IkpXVCJ9"}
        r = requests.get(url, timeout=15, headers=headers)
        r.raise_for_status()
        data = r.json()
        if data and len(data) > 0:
            latest = data[-1]
            return float(latest.get("v", 0)), latest.get("d", "")
        return None, None
    except Exception as e:
        logger.warning(f"BCRA error: {e}")
        return None, None
 
 
def fetch_argentina_macro():
    """Obtiene variables macro de Argentina via Yahoo Finance + fallback."""
    data = {}
 
    # Tipo de cambio USD/ARS desde Yahoo Finance
    try:
        import yfinance as yf
        ars = yf.download("ARS=X", period="5d", progress=False)
        if not ars.empty:
            if isinstance(ars.columns, pd.MultiIndex):
                ars.columns = ars.columns.get_level_values(0)
            tc = float(ars['Close'].iloc[-1])
            data["tipo_cambio"] = {"valor": tc, "fecha": datetime.now().strftime("%Y-%m-%d")}
        else:
            data["tipo_cambio"] = {"valor": None, "fecha": ""}
    except Exception as e:
        logger.warning(f"Yahoo ARS error: {e}")
        data["tipo_cambio"] = {"valor": None, "fecha": ""}
 
    # Tasa plazo fijo — BCRA estadisticasbcra.com en vivo, fallback a último conocido
    try:
        val, dt = _bcra_latest(BCRA_ENDPOINTS["tasa_plazo_fijo"])
        if val is not None:
            data["tasa_tamar"] = {"valor": val, "fecha": dt or datetime.now().strftime("%Y-%m-%d")}
        else:
            data["tasa_tamar"] = {"valor": 29.0, "fecha": "2026-05-01"}  # fallback último conocido
    except Exception as e:
        logger.warning(f"TAMAR error: {e}")
        data["tasa_tamar"] = {"valor": 29.0, "fecha": "2026-05-01"}
 
    # Reservas — intentar BCRA, fallback
    try:
        r = requests.get("https://api.estadisticasbcra.com/reservas", timeout=10, headers={"User-Agent": "InversionesBursatiles/1.0"})
        if r.status_code == 200:
            d = r.json()
            if d:
                data["reservas"] = {"valor": float(d[-1].get("v", 0)), "fecha": d[-1].get("d", "")}
            else:
                data["reservas"] = {"valor": 26500, "fecha": "fallback"}
        else:
            data["reservas"] = {"valor": 26500, "fecha": "fallback"}
    except Exception:
        data["reservas"] = {"valor": 26500, "fecha": "fallback"}
 
    # Riesgo país — Ámbito
    try:
        r = requests.get("https://mercados.ambito.com/riesgo-pais/datos", timeout=10, headers={"User-Agent": "InversionesBursatiles/1.0"})
        if r.status_code == 200:
            rp_data = r.json()
            if isinstance(rp_data, dict):
                val = float(str(rp_data.get("ultimo", "0")).replace(".", "").replace(",", "."))
                data["riesgo_pais"] = {"valor": val, "fecha": datetime.now().strftime("%Y-%m-%d")}
            else:
                data["riesgo_pais"] = {"valor": None, "fecha": ""}
        else:
            data["riesgo_pais"] = {"valor": None, "fecha": ""}
    except Exception as e:
        logger.warning(f"Riesgo país ARG error: {e}")
        data["riesgo_pais"] = {"valor": None, "fecha": ""}
 
    # Brecha cambiaria + CCL (bug real encontrado y confirmado con Bruno,
    # auditoría 28/07/2026, ver adendum de sesión: este mismo request a
    # Ámbito ya calculaba `ccl` para sacar la brecha, pero el valor de CCL
    # en sí se descartaba -- nunca se guardaba en ningún lado. Como
    # data/ccl_cache.json nunca se creaba, pricing_engine.get_ccl() y
    # trailing_stop._get_ccl() siempre caían a un fallback fijo (1487.0),
    # disfrazado de dato de mercado, incluso para trailing stops sobre
    # posiciones reales. Fuente de CCL confirmada explícitamente con Bruno:
    # Ámbito (mercados.ambito.com/dolar/cl/variacion) -- la misma que ya
    # se pedía acá, ahora persistida en vez de tirada.
    try:
        r = requests.get("https://mercados.ambito.com/dolar/cl/variacion", timeout=10, headers={"User-Agent": "InversionesBursatiles/1.0"})
        if r.status_code == 200:
            ccl_data = r.json()
            ccl = float(str(ccl_data.get("compra", "0")).replace(".", "").replace(",", "."))
            oficial = data.get("tipo_cambio", {}).get("valor") or 1
            if oficial > 0 and ccl > 0:
                brecha = round(((ccl / oficial) - 1) * 100, 1)
                data["brecha"] = {"valor": brecha, "fecha": datetime.now().strftime("%Y-%m-%d")}
            else:
                data["brecha"] = {"valor": None, "fecha": ""}
            if ccl > 0:
                data["ccl"] = {"valor": ccl, "fecha": datetime.now().strftime("%Y-%m-%d")}
                _persist_ccl_cache(ccl, fuente="ambito")
            else:
                data["ccl"] = {"valor": None, "fecha": ""}
        else:
            data["brecha"] = {"valor": None, "fecha": ""}
            data["ccl"] = {"valor": None, "fecha": ""}
    except Exception as e:
        logger.warning(f"Brecha/CCL error: {e}")
        data["brecha"] = {"valor": None, "fecha": ""}
        data["ccl"] = {"valor": None, "fecha": ""}
 
    # IPC mensual — datos.gob.ar (INDEC)
    def _fetch_ipc():
        url_ipc = (
            "https://apis.datos.gob.ar/series/api/series/"
            "?ids=" + DATOS_GOB_SERIES["ipc"] + "&last=3&format=json"
        )
        r_ipc = requests.get(url_ipc, timeout=15, headers={"User-Agent": "InversionesBursatiles/1.0"})
        r_ipc.raise_for_status()
        pts = r_ipc.json().get("data", [])
        if len(pts) >= 2:
            ipc_var = round((float(pts[-1][1]) / float(pts[-2][1]) - 1) * 100, 2)
            return ipc_var, str(pts[-1][0])
        return None, None

    val, dt = _with_last_known_fallback("ipc", _fetch_ipc)
    data["ipc"] = {"valor": val, "fecha": dt or ""}

    # Desempleo — datos.gob.ar (INDEC/EPH)
    # Corregido jul-2026: el ID opaco "41.1_DESO_TOTAL_D_L_29" quedó deprecado
    # (INDEC renumeró el dataset a 45.x) y devolvía 400 en cada corrida.
    # Se pasa a descarga CSV directa (más robusto: sobrevive renumeraciones
    # del ID opaco porque usa la URL de distribución + nombre de columna).
    # Redundancia PRIMARY->SECONDARY (jul-2026): si igual falla, cae al
    # último valor conocido en vez de None -- ver _with_last_known_fallback().
    _cfg = DATOS_GOB_CSV["desempleo"]
    val, dt = _with_last_known_fallback(
        "desempleo", lambda: _datos_gob_csv_latest(_cfg["url"], _cfg["column"], _cfg["multiplier"])
    )
    data["desempleo"] = {"valor": val, "fecha": dt or ""}

    # Balanza comercial — datos.gob.ar (INDEC)
    # Corregido jul-2026, mismo motivo que desempleo arriba. Mismo fallback.
    _cfg = DATOS_GOB_CSV["balanza"]
    val, dt = _with_last_known_fallback(
        "balanza_comercial", lambda: _datos_gob_csv_latest(_cfg["url"], _cfg["column"], _cfg["multiplier"])
    )
    data["balanza_comercial"] = {"valor": val, "fecha": dt or ""}

    # Resultado fiscal primario, como % del PBI (trailing-12m sobre PBI real)
    # Corregido jul-2026: el ID opaco de Mecon estaba deprecado (mismo bug que
    # desempleo/balanza) Y el reemplazo directo venía en ARS nominal, no %PBI.
    # Ver docstring de _resultado_fiscal_pct_pbi() para la metodología y su
    # limitación conocida frente al %PBI que publica Hacienda. Mismo
    # fallback: si el cálculo falla (ej. una de las 2 fuentes caída), cae al
    # último %PBI conocido en vez de perder la variable.
    val, dt = _with_last_known_fallback("resultado_fiscal", _resultado_fiscal_pct_pbi)
    data["resultado_fiscal"] = {"valor": val, "fecha": dt or ""}
 
    logger.info(f"ARG macro: {sum(1 for v in data.values() if v['valor'] is not None)}/{len(data)} variables obtenidas")
    return data
 
 
# ─────────────────────────────────────────────
# datos.gob.ar (Argentina) — API pública sin key
# ─────────────────────────────────────────────

DATOS_GOB_SERIES = {
    "ipc":       "148.3_INIVELNAL_DICI_M_26",
    "desempleo": "41.1_DESO_TOTAL_D_L_29",   # DEPRECADO jul-2026, ver DATOS_GOB_CSV
    "balanza":   "185.1_EXPOIM_TOTAL_D_M_26",  # DEPRECADO jul-2026, ver DATOS_GOB_CSV
    "fiscal":    "28.3_RFPFSPN_D_0_M_36",    # DEPRECADO jul-2026, reemplazado por cálculo %PBI (ver _resultado_fiscal_pct_pbi)
}

# Reemplazos vía descarga CSV directa (infra.datos.gob.ar) — más robustos que
# el endpoint de series por ID opaco, que INDEC/Mecon deprecaron sin aviso
# para varias series (hallazgo jul-2026). Confirmados en vivo, con datos
# frescos, al momento de este fix.
FALLBACK_HISTORY_PATH = "data/fallback_usage_history.json"


def _log_fallback_usage(key: str, fuente: str):
    """
    Observabilidad del fallback (roadmap externo #4, jul-2026): registra,
    por variable y por día, si la corrida de hoy usó la fuente primaria,
    cayó al último valor conocido (secundaria), o se quedó sin dato.

    Por qué importa: si una variable usa el fallback el 40% del tiempo, la
    fuente primaria está rota de forma recurrente -- aunque el sistema
    siga funcionando sin errores visibles, porque el fallback lo está
    tapando. Sin este registro, ese patrón pasa desapercibido.

    Dedupe por día: reruns del mismo día no deberían sumar entradas
    extra -- se pisa la entrada de hoy, no se acumula.
    """
    from src.github_persistence import load_json, save_json

    hoy = datetime.now().strftime("%Y-%m-%d")
    hist = load_json(FALLBACK_HISTORY_PATH, default={})
    hist.setdefault(key, [])
    hist[key] = [e for e in hist[key] if e.get("date") != hoy]
    hist[key].append({"date": hoy, "fuente": fuente})
    hist[key] = hist[key][-90:]  # ~3 meses de historial
    save_json(FALLBACK_HISTORY_PATH, hist, message=f"auto: fallback_usage {key} {hoy}")


def fallback_usage_stats(key: str, dias: int = 30) -> dict:
    """
    % de corridas que usaron fuente primaria vs. secundaria vs. sin dato,
    en los últimos `dias` días, para una variable. Pensado para un futuro
    panel ("IPC ARG: primary 92% / fallback 8%") -- hoy expone el cálculo,
    el dashboard puede consumirlo cuando se decida mostrarlo.
    """
    from src.github_persistence import load_json

    hist = load_json(FALLBACK_HISTORY_PATH, default={}).get(key, [])
    cutoff = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
    recientes = [e for e in hist if e.get("date", "") >= cutoff]
    n = len(recientes)
    if n == 0:
        return {"samples": 0}
    primary = sum(1 for e in recientes if e.get("fuente") == "primary")
    secondary = sum(1 for e in recientes if e.get("fuente") == "secondary")
    sin_dato = sum(1 for e in recientes if e.get("fuente") == "sin_dato")
    return {
        "samples": n,
        "primary_pct": round(primary / n * 100, 1),
        "secondary_pct": round(secondary / n * 100, 1),
        "sin_dato_pct": round(sin_dato / n * 100, 1),
    }


def _with_last_known_fallback(key, fetch_fn):
    """
    Redundancia PRIMARY -> SECONDARY (roadmap externo #3, jul-2026): envuelve
    cualquier fetch_fn() de una variable macro que puede fallar. Si fetch_fn
    devuelve un valor real, lo cachea en data/macro_last_known.json y lo
    retorna tal cual. Si falla (None), cae al último valor cacheado -- con
    una fecha de "stale_desde" para que se sepa que no es un dato fresco --
    en vez de perder la variable por completo (None) como pasaba antes.

    Generaliza el patrón que ya existía ad-hoc solo para TAMAR y reservas
    (fallback hardcodeado a un valor fijo) para las 4 variables que ya
    demostraron romperse este mes (desempleo, balanza, IPC, resultado
    fiscal) -- sin necesidad de investigar y mantener un proveedor
    alternativo real por variable, que sería mucho más esfuerzo para un
    beneficio similar en la práctica: seguir teniendo ALGÚN dato razonable
    en vez de None cuando la fuente primaria falla.

    IMPORTANTE: esto es una red de seguridad para cortes de 1-2 corridas, no
    una solución si la fuente está rota por semanas -- un valor de hace 2
    meses usado como si fuera de hoy puede ser peor que excluir la variable,
    así que hay un límite de antigüedad (ver MAX_STALE_DAYS) después del
    cual se deja de usar el fallback y se vuelve a None con un warning.
    """
    from src.github_persistence import load_json, save_json

    cache = load_json(LAST_KNOWN_PATH, default={})
    MAX_STALE_DAYS = 45  # más viejo que esto, mejor excluir que usar un dato añejo

    try:
        val, fecha = fetch_fn()
    except Exception as e:
        logger.warning(f"[fallback] {key}: fetch primario lanzó excepción ({e})")
        val, fecha = None, None

    if val is not None:
        cache[key] = {"valor": val, "fecha": fecha, "guardado_en": datetime.now().strftime("%Y-%m-%d")}
        save_json(LAST_KNOWN_PATH, cache, message=f"auto: macro_last_known {key} {datetime.now().strftime('%Y-%m-%d')}")
        _log_fallback_usage(key, "primary")
        return val, fecha

    # Fetch primario falló -- probar el último valor conocido
    cached = cache.get(key)
    if not cached:
        logger.warning(f"[fallback] {key}: fetch primario falló y no hay valor cacheado -- queda en None")
        _log_fallback_usage(key, "sin_dato")
        return None, None

    try:
        guardado = datetime.strptime(cached["guardado_en"], "%Y-%m-%d")
        antiguedad_dias = (datetime.now() - guardado).days
    except Exception:
        antiguedad_dias = 9999

    if antiguedad_dias > MAX_STALE_DAYS:
        logger.warning(f"[fallback] {key}: fetch primario falló y el último valor cacheado tiene "
                        f"{antiguedad_dias} días (> {MAX_STALE_DAYS}) -- demasiado viejo, queda en None")
        _log_fallback_usage(key, "sin_dato")
        return None, None

    logger.warning(f"[fallback] {key}: fetch primario falló, usando último valor conocido "
                    f"({cached['valor']}, cacheado hace {antiguedad_dias} días)")
    _log_fallback_usage(key, "secondary")
    return cached["valor"], cached.get("fecha", "")


DATOS_GOB_CSV = {
    "desempleo": {
        "url": ("https://infra.datos.gob.ar/catalog/sspm/dataset/45/distribution/"
                "45.2/download/tasa-desempleo-valores-trimestrales.csv"),
        "column": "eph_continua_tasa_desempleo_total",
        "multiplier": 100.0,  # la serie viene como fracción (0.078 = 7.8%)
    },
    "balanza": {
        "url": ("https://infra.datos.gob.ar/catalog/sspm/dataset/74/distribution/"
                "74.3/download/intercambio-comercial-argentino-mensual.csv"),
        "column": "ica_saldo_comercial",
        "multiplier": 1.0,  # ya viene en millones de USD, sin conversión
    },
    "resultado_primario": {
        "url": ("https://infra.datos.gob.ar/catalog/sspm/dataset/452/distribution/"
                "452.3/download/imig-mensual.csv"),
        "column": "resultado_primario",
        "multiplier": 1.0,  # millones de $ ARS nominales, mensual
    },
    "pbi_trimestral": {
        "url": ("https://infra.datos.gob.ar/catalog/sspm/dataset/8/distribution/"
                "8.2/download/producto-interno-bruto-precios-corrientes-valores-trimestrales-base-2004.csv"),
        "column": "producto_interno_bruto_precios_mercado",
        "multiplier": 1.0,  # millones de $ ARS nominales, trimestral
    },
}


def _datos_gob_csv_series(url, column):
    """Descarga la distribución completa de datos.gob.ar y devuelve un
    DataFrame [indice_tiempo, column] ordenado cronológicamente, sin nulos.
    Variante de _datos_gob_csv_latest() que devuelve toda la serie en vez de
    solo el último valor -- la necesitan los cálculos que suman una ventana
    (ej. trailing-12m del resultado fiscal como %PBI, ver más abajo).
    """
    try:
        r = requests.get(url, timeout=20, headers={"User-Agent": "InversionesBursatiles/1.0"})
        r.raise_for_status()
        import io
        df = pd.read_csv(io.StringIO(r.text))
        if column not in df.columns:
            logger.warning(f"datos.gob.ar CSV (serie): columna '{column}' no encontrada en {url}")
            return None
        df["indice_tiempo"] = pd.to_datetime(df["indice_tiempo"])
        out = df[["indice_tiempo", column]].dropna().sort_values("indice_tiempo").reset_index(drop=True)
        return out if not out.empty else None
    except Exception as e:
        logger.warning(f"datos.gob.ar CSV (serie) error [{column}]: {e}")
        return None


def _resultado_fiscal_pct_pbi():
    """
    Resultado fiscal primario como % del PBI, método trailing-12m:

        %PBI = (suma resultado_primario, últimos 12 meses)
             / (suma PBI nominal, últimos 4 trimestres disponibles) × 100

    LIMITACIÓN CONOCIDA (documentada jul-2026, validada empíricamente contra
    cifras oficiales antes de shippear esto):
    Este % NO va a coincidir con el que publica Hacienda en su informe mensual
    (IMIG). Hacienda calcula su %PBI contra el PBI NOMINAL PROYECTADO en la Ley
    de Presupuesto de ese año -- un supuesto de política fijado una vez al año
    y publicado en PDF, no como serie abierta y machine-readable -- en vez del
    PBI real ex-post que publica INDEC. Con inflación alta, esos dos números
    divergen fuerte y la brecha crece durante el año.

    Se validó cruzando contra dos cifras oficiales conocidas:
      - H1-2022: Hacienda publicó "0,99% del PBI" para un déficit de
        $755.975,7M. Usando el PBI REAL trailing-12m a jun-2022 da 0.31% --
        ~3.2x más chico.
      - 2020 (año pandemia): déficit primario oficial ~6.5% PBI. Usando PBI
        real anual 2020 da ~1.6% -- factor similar (~4x).
    La brecha es sistemática (mismo orden de magnitud en ambos chequeos), pero
    no hay manera de reconstruir el supuesto de Presupuesto desde datos
    abiertos -- no es una serie publicada, es un número de política que
    cambia por Ley cada año. Por eso esta función usa PBI real: da un %
    distinto al que sale en los diarios, pero es 100% automático, reproducible
    y consistente en el tiempo -- mide "resultado fiscal sobre el tamaño REAL
    de la economía", no sobre el supuesto de Presupuesto. RANGES["arg_fiscal"]
    fue recalibrado para esta definición (ver comentario ahí), no para la
    anterior.

    Devuelve (pct, fecha) o (None, None) si falta alguna de las dos series.
    """
    _cfg_rp = DATOS_GOB_CSV["resultado_primario"]
    _cfg_pbi = DATOS_GOB_CSV["pbi_trimestral"]

    df_rp = _datos_gob_csv_series(_cfg_rp["url"], _cfg_rp["column"])
    df_pbi = _datos_gob_csv_series(_cfg_pbi["url"], _cfg_pbi["column"])
    if df_rp is None or df_pbi is None or len(df_rp) < 12 or len(df_pbi) < 4:
        return None, None

    try:
        rp_12m = float(df_rp.tail(12)[_cfg_rp["column"]].sum())
        pbi_12m = float(df_pbi.tail(4)[_cfg_pbi["column"]].sum())
        if not pbi_12m:
            return None, None
        pct = round(rp_12m / pbi_12m * 100.0, 2)
        fecha = str(df_rp.iloc[-1]["indice_tiempo"].date())
        return pct, fecha
    except Exception as e:
        logger.warning(f"resultado_fiscal_pct_pbi error: {e}")
        return None, None


def _datos_gob_csv_latest(url, column, multiplier=1.0):
    """Obtiene el último valor no nulo de una columna en un CSV de datos.gob.ar,
    descargado directo de infra.datos.gob.ar (distribución completa) en vez de
    pedir una serie puntual por ID opaco al endpoint /series/api/series/.
    Ese endpoint por ID dejó de resolver varias series sin aviso (ver nota
    jul-2026); la descarga por distribución es más estable porque solo
    depende del nombre de columna, no de un identificador interno versionado.
    """
    try:
        r = requests.get(url, timeout=20, headers={"User-Agent": "InversionesBursatiles/1.0"})
        r.raise_for_status()
        import io
        df = pd.read_csv(io.StringIO(r.text))
        if column not in df.columns:
            logger.warning(f"datos.gob.ar CSV: columna '{column}' no encontrada en {url}")
            return None, None
        serie = df[["indice_tiempo", column]].dropna()
        if serie.empty:
            return None, None
        last = serie.iloc[-1]
        val = float(last[column]) * multiplier
        fecha = str(last["indice_tiempo"])
        return round(val, 2), fecha
    except Exception as e:
        logger.warning(f"datos.gob.ar CSV error [{column}]: {e}")
        return None, None


def _datos_gob_latest(series_id, n=3):
    """Obtiene el ultimo valor de una serie de datos.gob.ar (INDEC/Mecon)."""
    try:
        url = (
            "https://apis.datos.gob.ar/series/api/series/"
            "?ids=" + series_id + "&last=" + str(n) + "&format=json"
        )
        r = requests.get(url, timeout=15, headers={"User-Agent": "InversionesBursatiles/1.0"})
        r.raise_for_status()
        pts = r.json().get("data", [])
        if pts:
            return float(pts[-1][1]), str(pts[-1][0])
        return None, None
    except Exception as e:
        logger.warning("datos.gob.ar error [" + series_id + "]: " + str(e))
        return None, None


# ─────────────────────────────────────────────
# BCB/SGS (Brasil) — API pública
# ─────────────────────────────────────────────
 
BCB_SERIES = {
    "selic":        432,    # Meta SELIC
    "ipca":         433,    # IPCA acumulado 12m
    "desempleo":    24369,  # Taxa de desocupação
    "reservas":     13621,  # Reservas internacionales
    "brl_usd":    1,      # Dólar comercial (venda)
    "deuda_pib":  4503,   # Divida Liquida Setor Publico % PIB
}
 
 
def _bcb_latest(series_id):
    """Obtiene el último valor de una serie BCB/SGS."""
    try:
        end = datetime.now().strftime("%d/%m/%Y")
        start = (datetime.now() - timedelta(days=60)).strftime("%d/%m/%Y")
        url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{series_id}/dados?formato=json&dataInicial={start}&dataFinal={end}"
        headers = {"User-Agent": "InversionesBursatiles/1.0"}
        r = requests.get(url, timeout=15, headers=headers)
        r.raise_for_status()
        data = r.json()
        if data:
            latest = data[-1]
            return float(latest.get("valor", "0").replace(",", ".")), latest.get("data", "")
        return None, None
    except Exception as e:
        logger.warning(f"BCB error [serie {series_id}]: {e}")
        return None, None
 
 
def fetch_brasil_macro():
    """Obtiene variables macro de Brasil."""
    data = {}
 
    for name, series_id in BCB_SERIES.items():
        val, dt = _bcb_latest(series_id)
        data[name] = {"valor": val, "fecha": dt}
 
    # Riesgo país Brasil (EMBI) — scraping
    try:
        r = requests.get("https://mercados.ambito.com/riesgo-pais-historico/brasil", timeout=10, headers={"User-Agent": "InversionesBursatiles/1.0"})
        if r.status_code == 200:
            rp_data = r.json()
            if isinstance(rp_data, list) and len(rp_data) > 0:
                val = float(str(rp_data[-1].get("valor", "0")).replace(".", "").replace(",", "."))
                data["riesgo_pais"] = {"valor": val, "fecha": ""}
            else:
                data["riesgo_pais"] = {"valor": None, "fecha": ""}
        else:
            data["riesgo_pais"] = {"valor": None, "fecha": ""}
    except Exception as e:
        logger.warning(f"Riesgo país BRA error: {e}")
        data["riesgo_pais"] = {"valor": None, "fecha": ""}
 
    # PMI Manufacturero Brasil — FRED (BRAPMIMANMISMEI)
    try:
        fred_key = os.environ.get("FRED_API_KEY", "")
        if fred_key:
            url_pmi = (
                "https://api.stlouisfed.org/fred/series/observations"
                "?series_id=BRAPMIMANMISMEI&api_key=" + fred_key +
                "&file_type=json&limit=3&sort_order=desc"
            )
            r_pmi = requests.get(url_pmi, timeout=15, headers={"User-Agent": "InversionesBursatiles/1.0"})
            r_pmi.raise_for_status()
            obs = [o for o in r_pmi.json().get("observations", []) if o.get("value", ".") != "."]
            if obs:
                data["pmi"] = {"valor": float(obs[0]["value"]), "fecha": obs[0]["date"]}
            else:
                data["pmi"] = {"valor": None, "fecha": ""}
        else:
            data["pmi"] = {"valor": None, "fecha": "sin_fred_key"}
    except Exception as e:
        logger.warning("PMI BRA FRED error: " + str(e))
        data["pmi"] = {"valor": None, "fecha": ""}

    logger.info(f"BRA macro: {sum(1 for v in data.values() if v['valor'] is not None)}/{len(data)} variables obtenidas")
    return data
 
 
# ─────────────────────────────────────────────
# Normalización y cálculo de scores
# ─────────────────────────────────────────────
 
# Rangos (peor, mejor) para normalización
RANGES = {
    # Argentina
    "arg_tasa":     (60.0,   5.0),
    "arg_riesgo":   (2000,   100),
    "arg_reservas": (5000,   60000),
    "arg_tc":       (150.0,  70.0),
    "arg_brecha":   (80.0,   0.0),
    # Brasil
    "bra_selic":    (20.0,   5.0),
    "bra_riesgo":   (800,    80),
    "bra_ipca":     (10.0,   0.0),
    "bra_desempleo":(15.0,   3.0),
    "bra_reservas": (100000, 500000),
    "bra_brl":      (7.0,    4.0),
    # Argentina — ahora auto via datos.gob.ar
    "arg_ipc":       (6.0,    0.3),   # var% mensual
    "arg_desempleo": (15.0,   4.0),   # % desocupacion
    "arg_balanza":   (-2000,  4000),  # M USD
    "arg_fiscal":    (-2.0,   1.0),   # % PBI trailing-12m sobre PBI REAL (no el
                                       # supuesto de Presupuesto que usa Hacienda
                                       # en sus informes) -- recalibrado jul-2026,
                                       # ver _resultado_fiscal_pct_pbi()
    # Brasil — ahora auto via BCB/FRED
    "bra_deuda_pib": (100.0,  30.0),  # % PIB
    "bra_pmi":       (40.0,   58.0),  # PMI < 50 contraccion
    # USA
    "usa_fed":       (8.0,    0.5),
    "usa_cpi":      (9.0,    0.0),
    "usa_unemp":    (10.0,   2.0),
    "usa_gdp":      (-3.0,   5.0),
    "usa_conf":     (40.0,   140.0),
    "usa_pce":      (6.0,    1.0),
    "usa_hy":       (800,    200),
    "usa_dxy":      (115.0,  90.0),
    "usa_ism":      (40.0,   62.0),
}
 
 
def _normalize(valor, peor, mejor):
    """Normaliza a 0-100. Menor es mejor para tasas, mayor es mejor para reservas."""
    if mejor == peor or valor is None:
        return 50.0
    return max(0.0, min(100.0, (valor - peor) / (mejor - peor) * 100.0))


RAW_HISTORY_PATH = "data/macro_raw_history.json"
MIN_OBS_FOR_PERCENTILE = 24  # ~2 años de observaciones REALES (no duplicadas) — ver nota abajo


def _normalize_adaptive(range_key, valor, peor, mejor, raw_history):
    """
    Mejora 2.1: normalización macro con percentil rolling sobre historia real,
    en vez de depender solo de rangos fijos hardcodeados en RANGES (que quedan
    ciegos a regímenes nuevos — ej. si el EMBI supera el techo calibrado, hoy
    se clampea a 0/100 en vez de seguir reflejando que está empeorando/mejorando).

    Fallback automático a _normalize() (rango fijo) si hay menos de
    MIN_OBS_FOR_PERCENTILE observaciones históricas para esta variable — el
    historial (data/macro_raw_history.json) recién arranca a acumularse desde
    que se shippeó esto, así que en la práctica usa el fallback por ~2 meses.

    Preserva la MISMA direccionalidad que _normalize(): si mejor >= peor
    (orden "normal" de la tupla), percentil ascendente; si mejor < peor
    (orden "invertido"), percentil descendente — para no pisar la lógica de
    `invert` que cada variable aplica después en compute_macro_scores().
    """
    values = [v for v in raw_history.get(range_key, []) if v is not None]
    if len(values) < MIN_OBS_FOR_PERCENTILE:
        return _normalize(valor, peor, mejor)

    all_values = values + [valor]
    n = len(all_values)
    if mejor >= peor:
        rank = sum(1 for v in all_values if v <= valor)
    else:
        rank = sum(1 for v in all_values if v >= valor)
    return max(0.0, min(100.0, (rank / n) * 100.0))


def feature_health_score(range_key: str, ventana_reciente: int = 10) -> dict:
    """
    Feature drift monitoring (roadmap externo #5, jul-2026): compara la
    distribución RECIENTE de una variable macro contra su distribución
    HISTÓRICA, para detectar corrimientos estructurales que el sistema no
    vería de otra forma -- sin errores, sin None, sin warning, pero con la
    distribución de base cambiada. Ejemplo del caso que motiva esto:
    arg_fiscal podría pasar de moverse entre -1/+1 durante 2 años a
    moverse entre +5/+8 de golpe -- el pipeline sigue corriendo sin
    problemas, _normalize_adaptive() sigue devolviendo un score 0-100
    válido, pero la variable ya no significa lo mismo que antes.

    Métrica: z-score de la media reciente respecto a media/desvío
    históricos -- simple y estándar, más confiable que un test estadístico
    más sofisticado (KS-test, etc.) dado el poco historial real disponible
    hoy (ver limitación abajo).

    LIMITACIÓN CONOCIDA: data/macro_raw_history.json guarda una lista plana
    deduplicada por valor (ver _update_raw_history()), SIN timestamp por
    observación. Esto significa que "ventana reciente" es "las últimas N
    observaciones distintas", no "los últimos N días" -- para una variable
    que cambia a diario (ej. tipo_cambio) eso es ~N días, pero para una
    mensual (ej. IPC) son ~N meses. No hay forma de tener ventanas de
    calendario comparables entre variables sin agregar timestamps al
    historial (tarea aparte, no incluida acá).
    """
    raw_history = _load_raw_history()
    valores = [v for v in raw_history.get(range_key, []) if v is not None]
    n = len(valores)

    if n < MIN_OBS_FOR_PERCENTILE:
        return {
            "status": "insuficiente_historia",
            "samples": n,
            "nota": f"Necesita ≥{MIN_OBS_FOR_PERCENTILE} observaciones para un chequeo de drift confiable (hoy: {n})",
        }

    ventana = max(1, min(ventana_reciente, n // 3))  # nunca más de 1/3 de la historia como "reciente"
    reciente = valores[-ventana:]
    historico = valores[:-ventana]

    media_hist = float(np.mean(historico))
    std_hist = float(np.std(historico))
    media_reciente = float(np.mean(reciente))

    z = 0.0 if std_hist == 0 else (media_reciente - media_hist) / std_hist

    if abs(z) >= 2.5:
        status = "drift_fuerte"
    elif abs(z) >= 1.5:
        status = "drift_moderado"
    else:
        status = "estable"

    return {
        "status": status,
        "z_score": round(z, 2),
        "media_historica": round(media_hist, 3),
        "std_historica": round(std_hist, 3),
        "media_reciente": round(media_reciente, 3),
        "samples_historicos": len(historico),
        "samples_recientes": len(reciente),
    }


def feature_health_report() -> dict:
    """Corre feature_health_score() para todas las variables con historial
    acumulado -- pensado para un futuro panel de diagnóstico o para
    revisarlo puntualmente; hoy expone el cálculo, no se muestra todavía
    en ningún dashboard."""
    raw_history = _load_raw_history()
    return {k: feature_health_score(k) for k in raw_history.keys()}


def _load_raw_history() -> dict:
    from src.github_persistence import load_json
    return load_json(RAW_HISTORY_PATH, default={})


def _update_raw_history(raw_today: dict):
    """Agrega el valor crudo de hoy de cada variable macro al historial, para
    que _normalize_adaptive tenga con qué calcular percentiles reales.

    Deduplica valores consecutivos repetidos: variables mensuales/trimestrales
    (CPI, desempleo, GDP, etc.) reportan el MISMO valor en las 4 corridas
    diarias del pipeline hasta que sale el próximo dato real — sin esto, el
    historial se llena de copias idénticas (ej. ~120 copias del mismo CPI
    mensual) en vez de observaciones distintas, sesgando el percentil hacia
    cualquier valor que haya persistido más tiempo entre publicaciones."""
    from src.github_persistence import load_json, save_json
    if not raw_today:
        return
    history = load_json(RAW_HISTORY_PATH, default={})
    for range_key, val in raw_today.items():
        history.setdefault(range_key, [])
        if history[range_key] and history[range_key][-1] == val:
            continue  # mismo valor que la última observación — no es un dato nuevo
        history[range_key].append(val)
        history[range_key] = history[range_key][-1500:]  # cap generoso — con dedupe, esto ya no se llena de copias
    save_json(RAW_HISTORY_PATH, history, message=f"auto: macro_raw_history {datetime.now().strftime('%Y-%m-%d')}")


def bootstrap_fred_history(years: int = 3, api_key: str = None):
    """
    Mejora 2.1 (bootstrap): pre-carga `data/macro_raw_history.json` con `years`
    años de historia real de FRED para las 9 variables de USA (las únicas
    100% vía API con historia larga disponible — ARG depende de scraping de
    Ámbito y BRA es mayormente manual, sin backfill histórico fácil acá).

    Sin esto, _normalize_adaptive() recién empieza a usar percentil real
    después de MIN_OBS_FOR_PERCENTILE (60) corridas diarias — con esto, las
    9 variables de USA arrancan con percentil real desde el día 1.

    NO se llama automáticamente desde el pipeline — es un comando manual de
    una sola vez (correr una vez en Railway, o localmente con FRED_API_KEY).
    Usa observaciones de FRED como están publicadas (mensuales para la
    mayoría de estas series — no hace falta resolución diaria para un
    percentil sobre 3 años).
    """
    import requests
    from src.github_persistence import load_json, save_json

    api_key = api_key or os.getenv("FRED_API_KEY", "")
    if not api_key:
        logger.warning("[bootstrap_fred_history] Sin FRED_API_KEY, no se puede bootstrapear")
        return {}

    fred_series_usa = {
        "usa_fed":  "FEDFUNDS",
        "usa_cpi":  "CPIAUCSL",
        "usa_unemp": "UNRATE",
        "usa_gdp":  "A191RL1Q225SBEA",
        "usa_conf": "UMCSENT",
        "usa_pce":  "PCEPILFE",
        "usa_hy":   "BAMLH0A0HYM2",
        "usa_dxy":  "DTWEXBGS",
        # usa_ism (NAPMPI) deliberadamente excluido: su disponibilidad histórica
        # en FRED es irregular (ver §8.4 de la arquitectura) — mejor no asumir.
    }

    history = load_json(RAW_HISTORY_PATH, default={})
    start_date = (datetime.now() - timedelta(days=365 * years)).strftime("%Y-%m-%d")

    for range_key, series_id in fred_series_usa.items():
        try:
            url = "https://api.stlouisfed.org/fred/series/observations"
            params = {
                "series_id": series_id, "api_key": api_key, "file_type": "json",
                "observation_start": start_date, "sort_order": "asc",
            }
            r = requests.get(url, params=params, timeout=15)
            if r.status_code != 200:
                logger.warning(f"[bootstrap_fred_history] {series_id}: HTTP {r.status_code}")
                continue
            obs = r.json().get("observations", [])
            values = [float(o["value"]) for o in obs if o.get("value") not in (None, ".", "")]
            if values:
                history[range_key] = values[-1100:]
                logger.info(f"[bootstrap_fred_history] {range_key} ({series_id}): {len(values)} observaciones cargadas")
        except Exception as e:
            logger.warning(f"[bootstrap_fred_history] {range_key} falló: {e}")

    save_json(RAW_HISTORY_PATH, history, message=f"bootstrap: macro_raw_history FRED {years}y")
    return history
 
 
def _compute_market_confidence(range_keys_used, raw_history, n_expected=9):
    """
    Mejora 2.4: confidence del macro_score por mercado.

    El score macro de ARG/BRA y el de USA no son comparables en confiabilidad:
    USA tiene bootstrap de 3 años desde FRED (bootstrap_fred_history) y arranca
    usando percentil real desde el día 1, mientras que ARG depende de scraping
    de Ámbito y BRA es mayormente manual (xlsx) — sin historia previa cargada,
    ambos dependen del fallback de rango fijo en _normalize_adaptive hasta
    acumular MIN_OBS_FOR_PERCENTILE observaciones reales propias.

    Combina dos señales (50/50):
      - cobertura: variables que devolvieron valor hoy / variables esperadas (9)
      - calidad de normalización: de las que devolvieron valor, cuántas ya
        tienen suficiente historia real para usar percentil rolling en vez
        de rango fijo hardcodeado

    No modifica ningún score macro existente — es un campo adicional de
    transparencia, pensado para alimentar confidence_score.py (global) y,
    más adelante, un indicador en el dashboard.
    """
    n_obtenidas = len(range_keys_used)
    coverage = (n_obtenidas / n_expected) if n_expected else 0.0

    n_adaptive = sum(
        1 for rk in range_keys_used
        if len([v for v in raw_history.get(rk, []) if v is not None]) >= MIN_OBS_FOR_PERCENTILE
    )
    adaptive_ratio = (n_adaptive / n_obtenidas) if n_obtenidas else 0.0

    score = round((coverage * 0.5 + adaptive_ratio * 0.5) * 100, 1)
    if score >= 75:
        label = "ALTA"
    elif score >= 45:
        label = "MEDIA"
    else:
        label = "BAJA"

    return {
        "score": score,
        "label": label,
        "variables_esperadas": n_expected,
        "variables_obtenidas": n_obtenidas,
        "variables_con_percentil_real": n_adaptive,
        "variables_con_rango_fijo": n_obtenidas - n_adaptive,
    }


def _check_rango_sano(var_name, mercado, valor, peor, mejor, raw_history=None, range_key=None, margen=1.0):
    """
    Data Quality Layer (roadmap externo #2): flaguea un valor crudo que llegó
    pero está sospechosamente fuera de cualquier rango plausible -- no el
    rango de normalización (peor/mejor), sino un margen más ancho alrededor.

    IMPORTANTE (encontrado 23/07/2026 al probar esto con datos reales): varias
    variables (empezando por arg_tc) ya superaron MIN_OBS_FOR_PERCENTILE y
    _normalize_adaptive() las normaliza por percentil sobre el historial real,
    NO contra el rango fijo (peor/mejor) -- que puede quedar stale sin que el
    score se vea afectado (ej. arg_tc: rango fijo calibrado para 70-150,
    dólar real ~1480 desde hace rato, pero el score sigue siendo un percentil
    sano ~28 porque compara contra sí mismo, no contra 70-150). Si este check
    siguiera comparando solo contra peor/mejor, marcaría arg_tc como anomalía
    TODOS los días -- una falsa alarma permanente que le resta valor a toda
    la capa (si todo es anomalía, nada lo es). Por eso, si hay suficiente
    historial real para esta variable, el rango sano se deriva del propio
    historial (min/max observado ± margen) en vez del rango fijo -- mismo
    criterio de "confío en el dato reciente, no en la calibración original"
    que ya usa _normalize_adaptive() para el score.

    Devuelve None si está OK, o un dict describiendo la anomalía.
    """
    valores_hist = [v for v in (raw_history or {}).get(range_key, []) if v is not None] if range_key else []
    if len(valores_hist) >= MIN_OBS_FOR_PERCENTILE:
        lo, hi = min(valores_hist), max(valores_hist)
        modo = "adaptativo (historial real)"
    else:
        lo, hi = (peor, mejor) if peor <= mejor else (mejor, peor)
        modo = "fijo (RANGES)"
    ancho = abs(hi - lo) or 1.0
    limite_bajo = lo - ancho * margen
    limite_alto = hi + ancho * margen
    if valor < limite_bajo or valor > limite_alto:
        return {
            "mercado": mercado,
            "variable": var_name,
            "valor": valor,
            "modo_comparado": modo,
            "rango_esperado": [round(lo, 2), round(hi, 2)],
            "limite_sano": [round(limite_bajo, 2), round(limite_alto, 2)],
        }
    return None


def compute_macro_scores(arg_data, bra_data, usa_data):
    """
    Calcula scores macro normalizados por país.
    Retorna: {"MERVAL": score, "BOVESPA": score, "SP500": score, "timestamp": ..., "detalles": ...}
    """
    timestamp = datetime.now().isoformat()
    detalles = {"ARG": {}, "BRA": {}, "USA": {}}
    # Data Quality Layer (roadmap externo #2, jul-2026): detectar valores
    # crudos que llegaron pero están sospechosamente fuera de cualquier rango
    # plausible (ej. desempleo=45%, o un cambio de unidad no avisado tipo
    # USD -> miles USD). No reemplaza el clampeo de _normalize()/_normalize_
    # adaptive() -- eso sigue igual, el score no se rompe -- pero deja un
    # registro explícito para que no pase desapercibido, a diferencia de un
    # clamp silencioso a 0/100.
    anomalias_macro = []
    raw_history = _load_raw_history()
    raw_today = {}
 
    # ── Argentina ──
    arg_scores = []
    arg_range_keys = []
    vars_arg = [
        ("tasa_tamar",  "arg_tasa",     True),   # menor es mejor
        ("riesgo_pais", "arg_riesgo",   True),
        ("reservas",    "arg_reservas", False),  # mayor es mejor
        ("tipo_cambio", "arg_tc",       True),
        ("brecha",           "arg_brecha",    True),
        ("ipc",              "arg_ipc",       True),
        ("desempleo",        "arg_desempleo", True),
        ("balanza_comercial","arg_balanza",   False),
        ("resultado_fiscal", "arg_fiscal",    False),
    ]
    for var_name, range_key, invert in vars_arg:
        val = arg_data.get(var_name, {}).get("valor")
        if val is not None:
            peor, mejor = RANGES[range_key]
            anomalia = _check_rango_sano(var_name, "ARG", val, peor, mejor, raw_history=raw_history, range_key=range_key)
            if anomalia:
                anomalias_macro.append(anomalia)
                logger.warning(f"[data_quality] ARG.{var_name}={val} fuera de rango sano {anomalia['limite_sano']}")
            s = _normalize_adaptive(range_key, val, peor, mejor, raw_history)
            if invert:
                s = 100.0 - s
            arg_scores.append(s)
            arg_range_keys.append(range_key)
            detalles["ARG"][var_name] = {"valor": val, "score": round(s, 1)}
            raw_today[range_key] = val
 
    arg_macro = round(np.mean(arg_scores), 1) if arg_scores else 42.0
 
    # ── Brasil ──
    bra_scores = []
    bra_range_keys = []
    vars_bra = [
        ("selic",       "bra_selic",     True),
        ("riesgo_pais", "bra_riesgo",    True),
        ("ipca",        "bra_ipca",      True),
        ("desempleo",   "bra_desempleo", True),
        ("reservas",    "bra_reservas",  False),
        ("brl_usd",    "bra_brl",      True),
        ("deuda_pib",  "bra_deuda_pib", True),
        ("pmi",        "bra_pmi",       False),
    ]
    for var_name, range_key, invert in vars_bra:
        val = bra_data.get(var_name, {}).get("valor")
        if val is not None:
            peor, mejor = RANGES[range_key]
            anomalia = _check_rango_sano(var_name, "BRA", val, peor, mejor, raw_history=raw_history, range_key=range_key)
            if anomalia:
                anomalias_macro.append(anomalia)
                logger.warning(f"[data_quality] BRA.{var_name}={val} fuera de rango sano {anomalia['limite_sano']}")
            s = _normalize_adaptive(range_key, val, peor, mejor, raw_history)
            if invert:
                s = 100.0 - s
            bra_scores.append(s)
            bra_range_keys.append(range_key)
            detalles["BRA"][var_name] = {"valor": val, "score": round(s, 1)}
            raw_today[range_key] = val
 
    bra_macro = round(np.mean(bra_scores), 1) if bra_scores else 52.0
 
    # ── USA ──
    usa_scores = []
    usa_range_keys = []
    vars_usa = [
        ("fed_funds",    "usa_fed",    True),
        ("cpi",          "usa_cpi",    True),
        ("unemployment", "usa_unemp",  True),
        ("gdp_growth",   "usa_gdp",    False),
        ("consumer_conf","usa_conf",   False),
        ("pce_core",     "usa_pce",    True),
        ("hy_spread",    "usa_hy",     True),
        ("dxy",          "usa_dxy",    True),
        ("ism_mfg",      "usa_ism",    False),
    ]
    for var_name, range_key, invert in vars_usa:
        val = usa_data.get(var_name, {}).get("valor")
        if val is not None:
            peor, mejor = RANGES[range_key]
            anomalia = _check_rango_sano(var_name, "USA", val, peor, mejor, raw_history=raw_history, range_key=range_key)
            if anomalia:
                anomalias_macro.append(anomalia)
                logger.warning(f"[data_quality] USA.{var_name}={val} fuera de rango sano {anomalia['limite_sano']}")
            s = _normalize_adaptive(range_key, val, peor, mejor, raw_history)
            if invert:
                s = 100.0 - s
            usa_scores.append(s)
            usa_range_keys.append(range_key)
            detalles["USA"][var_name] = {"valor": val, "score": round(s, 1)}
            raw_today[range_key] = val
 
    usa_macro = round(np.mean(usa_scores), 1) if usa_scores else 45.0

    macro_confidence = {
        "MERVAL": _compute_market_confidence(arg_range_keys, raw_history),
        "BOVESPA": _compute_market_confidence(bra_range_keys, raw_history),
        "SP500": _compute_market_confidence(usa_range_keys, raw_history),
    }

    result = {
        "macro_scores": {
            "MERVAL": arg_macro,
            "BOVESPA": bra_macro,
            "SP500": usa_macro,
        },
        "macro_confidence": macro_confidence,
        "timestamp": timestamp,
        "variables_obtenidas": {
            "ARG": len(arg_scores),
            "BRA": len(bra_scores),
            "USA": len(usa_scores),
        },
        "detalles": detalles,
        "anomalias_macro": anomalias_macro,
    }

    # Data Quality Layer (roadmap externo #2, corregido 24/07/2026): antes
    # esto escribía data/data_quality.json local con open() directo -- no
    # persistía entre redeploys de Railway (a diferencia de
    # macro_last_known.json, que sí usa github_persistence) y solo
    # guardaba el snapshot de la última corrida, no un historial. Una
    # revisión externa lo marcó como "estás descartando la capa de calidad
    # de datos más valiosa que construiste" -- correcto: sin historial no
    # se puede medir frecuencia de anomalías, series más problemáticas, ni
    # degradación de proveedores en el tiempo. Ahora se persiste como
    # historial (data/data_quality_history.json), un registro por día
    # (dedupe si el pipeline corre varias veces el mismo día), últimos ~90
    # días.
    try:
        from src.github_persistence import load_json as _dq_load, save_json as _dq_save
        hoy = datetime.now().strftime("%Y-%m-%d")
        dq_hist = _dq_load(DATA_QUALITY_HISTORY_PATH, default=[])
        dq_hist = [e for e in dq_hist if e.get("date") != hoy]
        dq_hist.append({
            "date": hoy,
            "timestamp": timestamp,
            "variables_obtenidas": result["variables_obtenidas"],
            "anomalias": anomalias_macro,
        })
        dq_hist = dq_hist[-90:]
        _dq_save(DATA_QUALITY_HISTORY_PATH, dq_hist, message=f"auto: data_quality {hoy}")
    except Exception as e:
        logger.warning(f"No se pudo guardar {DATA_QUALITY_HISTORY_PATH}: {e}")
 
    # Cache para debugging y fallback
    try:
        os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
        with open(CACHE_PATH, "w") as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
    except Exception:
        pass
 
    logger.info(
        f"Macro scores auto: ARG={arg_macro} (conf={macro_confidence['MERVAL']['label']}) "
        f"BRA={bra_macro} (conf={macro_confidence['BOVESPA']['label']}) "
        f"USA={usa_macro} (conf={macro_confidence['SP500']['label']})"
    )
    _update_raw_history(raw_today)
    return result
 
 
# ─────────────────────────────────────────────
# Historial de scores macro — para delta semanal en el dashboard
# (persiste a GitHub porque el filesystem de Railway es efímero)
# ─────────────────────────────────────────────

HISTORY_PATH = "data/macro_score_history.json"
GH_REPO_FULL = "Brunogatti79/inversiones-bursatiles"


def sync_macro_history_from_github():
    """Descarga macro_score_history.json fresco de GitHub. Llamar al arrancar Railway."""
    from src.github_persistence import pull_file
    pull_file(HISTORY_PATH)


def _update_macro_history(macro_scores, macro_confidence=None):
    """Agrega el score macro de hoy al historial (dedupe por fecha) y lo pushea a GitHub.

    macro_confidence es opcional (compatibilidad hacia atrás con llamadas viejas)
    para no romper si algún caller no lo pasa."""
    from src.github_persistence import load_json, save_json
    today = datetime.now().strftime("%Y-%m-%d")
    history = load_json(HISTORY_PATH, default=[])
    macro_confidence = macro_confidence or {}

    history = [h for h in history if h.get("date") != today]
    entry = {
        "date": today,
        "MERVAL": macro_scores.get("MERVAL"),
        "BOVESPA": macro_scores.get("BOVESPA"),
        "SP500": macro_scores.get("SP500"),
    }
    if macro_confidence:
        entry["confidence"] = {
            mkt: macro_confidence[mkt].get("label")
            for mkt in ("MERVAL", "BOVESPA", "SP500") if mkt in macro_confidence
        }
    history.append(entry)
    history.sort(key=lambda h: h.get("date", ""))
    history = history[-90:]  # ~3 meses de historial

    save_json(HISTORY_PATH, history, message=f"auto: macro_score_history {today}")


# ─────────────────────────────────────────────
# Función principal — llamada desde pipeline
# ─────────────────────────────────────────────
 
def update_xlsx_macro(detalles, xlsx_path="data/modelo_macro_micro_señales.xlsx"):
    """
    Actualiza la hoja 'Macro Variables' del xlsx con los valores frescos
    obtenidos por macro_auto, para que queden persistentes.
    """
    import openpyxl
 
    if not os.path.exists(xlsx_path):
        logger.warning(f"[macro_auto] xlsx no encontrado para update: {xlsx_path}")
        return False
 
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Macro Variables"]
 
        # Mapeo: variable_key de detalles → substring que aparece en columna Variable del xlsx
        MAPPING_ARG = {
            "tasa_tamar":  "Tasa de interés TAMAR",
            "riesgo_pais": "Riesgo País",
            "reservas":    "Reservas BCRA",
            "tipo_cambio": "Tipo de cambio real",
            "brecha":      "Brecha cambiaria",
        }
        MAPPING_BRA = {
            "selic":       "Tasa SELIC",
            "riesgo_pais": "Riesgo País",
            "ipca":        "Inflación IPCA",
            "desempleo":   "Desempleo",
            "reservas":    "Reservas BCB",
            "brl_usd":     "Tipo de cambio BRL",
        }
        MAPPING_USA = {
            "fed_funds":     "Fed Funds Rate",
            "cpi":           "Inflación CPI",
            "unemployment":  "Desempleo",
            "gdp_growth":    "GDP Growth",
            "consumer_conf": "Confianza consumidor",
            "pce_core":      "PCE Core",
            "hy_spread":     "Spread High Yield",
            "dxy":           "Índice DXY",
            "ism_mfg":       "ISM Manufacturero",
        }
 
        pais_map = {
            "Argentina": ("ARG", MAPPING_ARG),
            "Brasil":    ("BRA", MAPPING_BRA),
            "EE.UU.":    ("USA", MAPPING_USA),
        }
 
        updated_count = 0
        for row in range(2, ws.max_row + 1):
            pais_cell = str(ws.cell(row, 1).value or "")
            var_cell  = str(ws.cell(row, 2).value or "")
 
            for pais_nombre, (pais_key, mapping) in pais_map.items():
                if pais_cell != pais_nombre:
                    continue
                det_pais = detalles.get(pais_key, {})
                for var_key, var_xlsx_name in mapping.items():
                    if var_xlsx_name in var_cell and var_key in det_pais:
                        new_val = det_pais[var_key].get("valor")
                        new_score = det_pais[var_key].get("score")
                        if new_val is not None:
                            ws.cell(row, 3).value = new_val  # Columna Valor
                            updated_count += 1
                        if new_score is not None:
                            ws.cell(row, 4).value = new_score  # Columna Score
                        # Actualizar fuente con timestamp
                        old_fuente = str(ws.cell(row, 6).value or "")
                        ws.cell(row, 6).value = f"Auto {datetime.now().strftime('%d/%m/%Y')} — {old_fuente[:60]}"
                        break
 
        wb.save(xlsx_path)
        logger.info(f"[macro_auto] xlsx actualizado: {updated_count} variables en {xlsx_path}")
        return updated_count > 0
 
    except Exception as e:
        logger.error(f"[macro_auto] Error actualizando xlsx: {e}")
        return False
 
 
def fetch_all_macro():
    """
    Descarga datos macro de los 3 países y calcula scores.
    Si falla alguna API, usa fallback del cache o hardcoded.
    Actualiza el xlsx para persistencia.
    Retorna dict compatible con xlsx_signals format.
    """
    logger.info("Descargando datos macro automáticos...")
 
    arg_data = fetch_argentina_macro()
    bra_data = fetch_brasil_macro()
    usa_data = fetch_usa_macro()
 
    result = compute_macro_scores(arg_data, bra_data, usa_data)
 
    # Persistir en xlsx para que la próxima ejecución tenga datos frescos
    vars_ok = result.get("variables_obtenidas", {})
    total_ok = sum(vars_ok.values())
    if total_ok >= 8:  # Al menos 8 de 27 variables obtenidas
        try:
            update_xlsx_macro(result.get("detalles", {}))
        except Exception as e:
            logger.warning(f"[macro_auto] No se pudo actualizar xlsx: {e}")

    # Acumular historial de scores macro (para delta semanal en el dashboard)
    try:
        _update_macro_history(result["macro_scores"], result.get("macro_confidence"))
    except Exception as e:
        logger.warning(f"[macro_auto] No se pudo actualizar macro_score_history: {e}")
 
    # Formato compatible con lo que espera pipeline/analyzer
    return {
        "macro_scores": result["macro_scores"],
        "macro_confidence": result.get("macro_confidence", {}),
        "macro_timestamp": result["timestamp"],
        "macro_auto": True,
        "detalles": result["detalles"],
    }
 
 
def _persist_ccl_cache(ccl: float, fuente: str = "ambito"):
    """
    Persiste el CCL en data/ccl_cache.json en el formato que ya esperan
    los lectores existentes -- pricing_engine.get_ccl() y
    trailing_stop._get_ccl() leen la clave "compra", contrato que NO se
    modifica acá (bug real, auditoría 28/07/2026, ver adendum de sesión:
    el archivo nunca se creaba, así que esos dos lectores siempre caían a
    fallback). Fuente confirmada explícitamente con Bruno: Ámbito.

    Solo escribe local -- el push a GitHub para persistencia entre
    redeploys de Railway lo sigue haciendo pipeline.py con
    github_persistence.push_file(), mismo patrón que el resto de los
    archivos persistidos (no una excepción bypasseando ese mecanismo).
    """
    try:
        os.makedirs(os.path.dirname(CCL_CACHE_PATH), exist_ok=True)
        payload = {
            "compra":    ccl,
            "fuente":    fuente,
            "fecha":     datetime.now().strftime("%Y-%m-%d"),
            "timestamp": datetime.now().isoformat(),
        }
        with open(CCL_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"[macro_auto] No se pudo persistir ccl_cache.json: {e}")


def get_ccl_data(max_age_hours: float = 4.0) -> dict:
    """
    Fuente de datos para GET /api/ccl (start_server.py). Antes de esta
    función, el endpoint importaba `get_ccl_data` de este módulo -- una
    función que nunca existió en el repo, por lo que /api/ccl siempre
    devolvía HTTP 500 (bug real encontrado en auditoría 28/07/2026, ver
    adendum de sesión, no relacionado con el bug de persistencia de
    ccl_cache.json pero descubierto en la misma revisión).

    Prioridad:
      1. Cache local (data/ccl_cache.json) si tiene menos de
         max_age_hours -- evita pegarle a Ámbito en cada request al
         endpoint cuando el pipeline ya lo actualizó hace poco (mismo
         concepto de TTL ~4h ya documentado en versiones previas de la
         arquitectura para este archivo).
      2. Si el cache no existe o está vencido: fetch en vivo a Ámbito,
         mismo request y parseo que fetch_argentina_macro(), y lo persiste
         para que el próximo request (o el próximo pipeline) lo reuse.
      3. Si todo falla: devuelve compra=None explícito -- no inventa un
         valor. El caller (start_server._handle_get_ccl) decide qué hacer
         con eso, no se enmascara con un fallback silencioso acá.
    """
    try:
        if os.path.exists(CCL_CACHE_PATH):
            with open(CCL_CACHE_PATH, encoding="utf-8") as f:
                cached = json.load(f)
            ts = cached.get("timestamp")
            if ts:
                edad_horas = (datetime.now() - datetime.fromisoformat(ts)).total_seconds() / 3600
                if edad_horas <= max_age_hours and float(cached.get("compra", 0) or 0) > 0:
                    return cached
    except Exception as e:
        logger.warning(f"[macro_auto] get_ccl_data: cache local inválido, reintentando fetch: {e}")

    try:
        r = requests.get("https://mercados.ambito.com/dolar/cl/variacion", timeout=10,
                          headers={"User-Agent": "InversionesBursatiles/1.0"})
        if r.status_code == 200:
            ccl_data = r.json()
            ccl = float(str(ccl_data.get("compra", "0")).replace(".", "").replace(",", "."))
            if ccl > 0:
                _persist_ccl_cache(ccl, fuente="ambito")
                return {
                    "compra":    ccl,
                    "fuente":    "ambito",
                    "fecha":     datetime.now().strftime("%Y-%m-%d"),
                    "timestamp": datetime.now().isoformat(),
                }
    except Exception as e:
        logger.warning(f"[macro_auto] get_ccl_data: fetch en vivo a Ámbito falló: {e}")

    return {
        "compra": None,
        "fuente": None,
        "fecha":  "",
        "error":  "CCL no disponible (cache vencido y fetch en vivo a Ámbito falló)",
    }


def get_cached_macro():
    """Lee el último cache de macro como fallback."""
    try:
        if os.path.exists(CACHE_PATH):
            with open(CACHE_PATH) as f:
                cached = json.load(f)
            return {
                "macro_scores": cached["macro_scores"],
                "macro_confidence": cached.get("macro_confidence", {}),
                "macro_timestamp": cached["timestamp"],
                "macro_auto": True,
            }
    except Exception:
        pass
    return None
 
