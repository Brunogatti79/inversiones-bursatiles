#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════╗
║  MODELO MACRO-MICRO-SEÑALES v2.0                                    ║
║  Inversiones Bursátiles — MERVAL · BOVESPA · S&P 500               ║
║  Fecha: mayo 2026                                                   ║
║                                                                      ║
║  CAMBIOS v2 (respecto a v1):                                        ║
║  1. ASSET_QUALITY vs ENTRY_SCORE (scores separados)                 ║
║  2. R/R normalizado en ranking accionable                           ║
║  3. Distancia al máximo 52s como variable técnica                   ║
║  4. Compatibilidad total con pipeline Railway + dashboard HTML      ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
import json
import os
import warnings
warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════
# 1. DATOS MACRO POR PAÍS (9 variables cada uno)
#    Formato: (valor_actual, peor, mejor, impacto, descripcion)
#    impacto: +1 = mayor es mejor, -1 = menor es mejor
# ═══════════════════════════════════════════════════════════════

MACRO_ARG = {
    # ── Actualizado: abril 2026 ──
    "Tasa de interés TAMAR"     : (29.0,   60.0,    5.0,   -1, "TNA plazo fijo 30d - abr-2026"),
    "Riesgo País (pb)"          : (730,    2000,    100,   -1, "EMBI Argentina abr-2026"),
    "Inflación mensual (%)"     : (2.9,    10.0,    0.0,   -1, "IPC feb-2026 INDEC"),
    "Desempleo (%)"             : (6.5,    20.0,    2.0,   -1, "Estimado Q1-2026"),
    "Reservas BCRA (kUSD)"      : (26500,  5000,    60000,  1, "Reservas netas abr-2026"),
    "Tipo de cambio real (TCR)" : (108.0,  150.0,   70.0,  -1, "TCR abr-2026 - apreciación moderada"),
    "Brecha cambiaria (%)"      : (12.0,   80.0,    0.0,   -1, "CCL vs oficial abr-2026"),
    "Balanza comercial (M USD)" : (1200,   -2000,   3000,   1, "Superávit feb-2026 INDEC"),
    "Resultado fiscal (%PBI)"   : (0.3,    -5.0,    2.0,    1, "Superávit primario feb-2026"),
}

MACRO_BRA = {
    "Tasa SELIC (%)"            : (14.75,  20.0,    5.0,   -1, "SELIC post-COPOM mar-2026"),
    "Riesgo País (pb)"          : (235,    800,     80,    -1, "EMBI Brasil abr-2026"),
    "Inflación IPCA (%)"        : (3.81,   10.0,    0.0,   -1, "IPCA feb-2026"),
    "Desempleo (%)"             : (6.4,    15.0,    3.0,   -1, "IBGE abr-2026"),
    "Reservas BCB (kUSD)"       : (350000, 100000,  500000,  1, "Reservas internacionales BCB"),
    "BRL/USD"                   : (5.75,   7.0,     4.0,   -1, "Tipo de cambio abr-2026"),
    "Diferencial tasa real (pp)": (5.2,    0.0,     8.0,    1, "SELIC real vs inflación"),
    "Deuda/PIB (%)"             : (78.5,   100.0,   40.0,  -1, "Deuda bruta gob. central"),
    "PMI Manufacturero"         : (52.1,   42.0,    60.0,   1, "S&P Global PMI mar-2026"),
}

MACRO_USA = {
    "Fed Funds Rate (%)"        : (4.375,  8.0,     0.5,   -1, "Fed en pausa abr-2026"),
    "Inflación CPI (%)"         : (3.1,    9.0,     0.0,   -1, "CPI mar-2026"),
    "Desempleo (%)"             : (4.3,    10.0,    2.0,   -1, "BLS abr-2026"),
    "GDP Growth YoY (%)"        : (2.1,   -3.0,     5.0,    1, "Revisado a la baja"),
    "Confianza consumidor"      : (88.0,   40.0,    140.0,  1, "Conference Board abr-2026"),
    "PCE Core (%)"              : (2.9,    6.0,     1.0,   -1, "PCE subyacente feb-2026"),
    "Spread High Yield (pb)"    : (420,    800,     200,   -1, "ICE BofA HY mar-2026"),
    "Índice DXY"                : (104.5,  115.0,   90.0,  -1, "Dólar index abr-2026"),
    "ISM Manufacturero"         : (50.3,   40.0,    62.0,   1, "ISM mar-2026"),
}

# ═══════════════════════════════════════════════════════════════
# 2. TICKERS Y CONFIGURACIÓN POR MERCADO
# ═══════════════════════════════════════════════════════════════

MERVAL_TICKERS = {
    "GGAL.BA": ("Grupo Financiero Galicia", "Financiero", 45),
    "YPFD.BA": ("YPF", "Energía", 50),
    "PAMP.BA": ("Pampa Energía", "Utilities", 48),
    "BMA.BA":  ("Banco Macro", "Financiero", 44),
    "TECO2.BA":("Telecom Argentina", "Telecom", 42),
    "TRAN.BA": ("Transener", "Utilities", 40),
    "TGSU2.BA":("TGS", "Energía", 46),
    "EDN.BA":  ("Edenor", "Utilities", 43),
    "CRES.BA": ("Cresud", "Agro", 42),
    "IRSA.BA": ("IRSA", "Real Estate", 40),
    "LOMA.BA": ("Loma Negra", "Materiales", 41),
    "TXAR.BA": ("Ternium Argentina", "Materiales", 43),
    "CEPU.BA": ("Central Puerto", "Utilities", 45),
    "MIRG.BA": ("Mirgor", "Consumo", 38),
    "SUPV.BA": ("Supervielle", "Financiero", 42),
    "BBAR.BA": ("BBVA Argentina", "Financiero", 44),
    "COME.BA": ("Sociedad Comercial del Plata", "Holding", 35),
    "ALUA.BA": ("Aluar", "Materiales", 40),
    "VALO.BA": ("Grupo Valores", "Financiero", 38),
    "BYMA.BA": ("BYMA", "Financiero", 42),
}

BOVESPA_TICKERS = {
    "PETR4.SA": ("Petrobras PN", "Energía", 52),
    "VALE3.SA": ("Vale ON", "Minería", 48),
    "ITUB4.SA": ("Itaú Unibanco PN", "Financiero", 50),
    "BBAS3.SA": ("Banco do Brasil ON", "Financiero", 48),
    "BBDC4.SA": ("Bradesco PN", "Financiero", 44),
    "WEGE3.SA": ("WEG ON", "Industrial", 50),
    "RENT3.SA": ("Localiza ON", "Consumo", 42),
    "EQTL3.SA": ("Equatorial ON", "Utilities", 46),
    "MGLU3.SA": ("Magazine Luiza ON", "Retail", 35),
    "EMBR3.SA": ("Embraer ON", "Industrial", 48),
    "B3SA3.SA": ("B3 ON", "Financiero", 45),
    "SUZB3.SA": ("Suzano ON", "Papel/Celulosa", 46),
    "ABEV3.SA": ("Ambev ON", "Consumo", 46),
    "CYRE3.SA": ("Cyrela ON", "Real Estate", 45),
    "EGIE3.SA": ("Engie Brasil ON", "Utilities", 50),
    "BPAC11.SA":("BTG Pactual units", "Financiero", 48),
    "HAPV3.SA": ("Hapvida ON", "Salud", 40),
    "CMIG4.SA": ("Cemig PN", "Utilities", 44),
    "ITSA4.SA": ("Itaúsa PN", "Holding", 46),
}

SP500_TICKERS = {
    "AAPL": ("Apple", "Tecnología", 50),
    "MSFT": ("Microsoft", "Tecnología", 52),
    "NVDA": ("NVIDIA", "Tecnología", 50),
    "AMZN": ("Amazon", "Consumo", 48),
    "GOOGL":("Alphabet", "Tecnología", 50),
    "META": ("Meta Platforms", "Tecnología", 48),
    "JPM":  ("JPMorgan Chase", "Financiero", 50),
    "JNJ":  ("Johnson & Johnson", "Salud", 55),
    "XOM":  ("Exxon Mobil", "Energía", 52),
    "CVX":  ("Chevron", "Energía", 50),
    "UNH":  ("UnitedHealth", "Salud", 52),
    "PG":   ("Procter & Gamble", "Consumo", 54),
    "KO":   ("Coca-Cola", "Consumo", 54),
    "MCD":  ("McDonald's", "Consumo", 52),
    "CAT":  ("Caterpillar", "Industrial", 48),
    "BA":   ("Boeing", "Industrial", 42),
    "DIS":  ("Disney", "Comunicación", 44),
    "NFLX": ("Netflix", "Comunicación", 48),
    "AMD":  ("AMD", "Tecnología", 46),
    "TSLA": ("Tesla", "Consumo", 42),
    "V":    ("Visa", "Financiero", 52),
    "MA":   ("Mastercard", "Financiero", 52),
    "HD":   ("Home Depot", "Consumo", 50),
}


# ═══════════════════════════════════════════════════════════════
# 3. FUNCIONES CORE DEL MODELO
# ═══════════════════════════════════════════════════════════════

def normalizar(valor, peor, mejor):
    """Normaliza un valor al rango 0-100 dados los extremos."""
    if mejor == peor:
        return 50.0
    return max(0.0, min(100.0, (valor - peor) / (mejor - peor) * 100.0))


def calcular_score_macro(macro_dict):
    """Calcula score macro como promedio simple de variables normalizadas."""
    scores = []
    detalles = {}
    for var, (valor, peor, mejor, impacto, desc) in macro_dict.items():
        s = normalizar(valor, peor, mejor)
        if impacto == -1:
            s = 100.0 - s
        scores.append(s)
        detalles[var] = round(s, 1)
    return round(np.mean(scores), 1), detalles


def calcular_score_tecnico(precio_actual, precios_historicos):
    """
    Calcula score técnico basado en RSI(14), Momentum 21d, cruce MA20/MA50.
    Retorna: (score_tecnico, rsi, momentum_21d, ma20_gt_ma50)
    """
    if precios_historicos is None or len(precios_historicos) < 50:
        return 40.0, 50.0, 0.0, False
    
    closes = precios_historicos['Close'].values
    
    # RSI(14)
    deltas = np.diff(closes[-15:])
    gains = np.where(deltas > 0, deltas, 0)
    losses = np.where(deltas < 0, -deltas, 0)
    avg_gain = np.mean(gains) if len(gains) > 0 else 0.001
    avg_loss = np.mean(losses) if len(losses) > 0 else 0.001
    rs = avg_gain / max(avg_loss, 0.001)
    rsi = 100.0 - (100.0 / (1.0 + rs))
    
    # RSI score: oversold (30) = 80pts, neutral (50) = 50pts, overbought (70) = 20pts
    if rsi <= 30:
        rsi_score = 80.0 + (30.0 - rsi) * (20.0 / 30.0)
    elif rsi <= 50:
        rsi_score = 50.0 + (50.0 - rsi) * (30.0 / 20.0)
    elif rsi <= 70:
        rsi_score = 20.0 + (70.0 - rsi) * (30.0 / 20.0)
    else:
        rsi_score = max(0.0, 20.0 - (rsi - 70.0) * (20.0 / 30.0))
    
    # Momentum 21d
    if len(closes) >= 22:
        momentum = ((closes[-1] / closes[-22]) - 1.0) * 100.0
    else:
        momentum = 0.0
    mom_score = min(100.0, max(0.0, 50.0 + momentum * 2.0))
    
    # MA20 vs MA50
    ma20 = np.mean(closes[-20:]) if len(closes) >= 20 else closes[-1]
    ma50 = np.mean(closes[-50:]) if len(closes) >= 50 else closes[-1]
    ma_cross = ma20 > ma50
    ma_score = 70.0 if ma_cross else 30.0
    
    # Score técnico ponderado
    score = 0.40 * rsi_score + 0.35 * mom_score + 0.25 * ma_score
    
    return round(score, 1), round(rsi, 1), round(momentum, 2), ma_cross


# ═══════════════════════════════════════════════════════════════
# 4. MEJORA 1: ASSET_QUALITY vs ENTRY_SCORE
# ═══════════════════════════════════════════════════════════════

def calcular_asset_quality(score_macro, score_fundamental, score_sectorial):
    """
    Calidad intrínseca del activo.
    ASSET_QUALITY = 50% Macro + 30% Fundamental + 20% Sectorial
    """
    return round(0.50 * score_macro + 0.30 * score_fundamental + 0.20 * score_sectorial, 1)


# ═══════════════════════════════════════════════════════════════
# 5. MEJORA 2: RATIO RIESGO/RETORNO
# ═══════════════════════════════════════════════════════════════

def calcular_rr(precio_actual, max_52s, min_52s):
    """
    R/R = Upside potencial / Downside potencial
    Normalizado a 0-100 con rango [0, 3]
    """
    if precio_actual <= 0 or max_52s <= 0 or min_52s <= 0:
        return 0.0, 0.0
    
    upside = (max_52s - precio_actual) / precio_actual
    downside = (precio_actual - min_52s) / precio_actual
    
    if downside <= 0.001:
        rr_ratio = min(upside / 0.001, 5.0)
    else:
        rr_ratio = upside / downside
    
    rr_normalizado = min(100.0, max(0.0, (rr_ratio / 3.0) * 100.0))
    return round(rr_ratio, 2), round(rr_normalizado, 1)


# ═══════════════════════════════════════════════════════════════
# 6. MEJORA 3: DISTANCIA AL MÁXIMO 52s NORMALIZADA
# ═══════════════════════════════════════════════════════════════

def normalizar_dist_max(dist_max_pct):
    """
    dist_max_pct viene negativo (ej: -22.5).
    0% (en el max) = score 0, -40%+ (lejos) = score 100
    """
    distancia = abs(dist_max_pct)
    return round(min(100.0, max(0.0, (distancia / 40.0) * 100.0)), 1)


def calcular_entry_score(score_tecnico, rr_normalizado, dist_max_norm):
    """
    Calidad del punto de entrada.
    ENTRY_SCORE = 60% Técnico + 25% R/R + 15% Dist. Max
    """
    return round(0.60 * score_tecnico + 0.25 * rr_normalizado + 0.15 * dist_max_norm, 1)


# ═══════════════════════════════════════════════════════════════
# 7. SCORE FINAL V2 Y SEÑALES
# ═══════════════════════════════════════════════════════════════

def calcular_score_final_v2(asset_quality, entry_score):
    """
    SCORE_FINAL_V2 = 50% ASSET_QUALITY + 50% ENTRY_SCORE
    """
    score = round(0.50 * asset_quality + 0.50 * entry_score, 1)
    
    if score >= 70:
        señal = "⭐ Compra Fuerte"
    elif score >= 60:
        señal = "🟢 Compra"
    elif score >= 45:
        señal = "🟡 Neutral/Esperar"
    elif score >= 35:
        señal = "🟠 Venta Parcial"
    else:
        señal = "🔴 Venta"
    
    return score, señal


def calcular_ranking_accionable(score_final_v2, rr_norm):
    """
    RANKING_ACCIONABLE = 60% Score_Final_V2 + 40% R/R_Norm
    """
    return round(0.60 * score_final_v2 + 0.40 * rr_norm, 1)


# ═══════════════════════════════════════════════════════════════
# 8. COMPATIBILIDAD V1 (para dashboard existente)
# ═══════════════════════════════════════════════════════════════

def calcular_score_v1(score_macro, score_tecnico, score_sectorial):
    """Score V1 original: 40% Macro + 40% Técnico + 20% Sectorial"""
    score = round(0.40 * score_macro + 0.40 * score_tecnico + 0.20 * score_sectorial, 1)
    
    if score >= 70:
        señal = "⭐ Compra Fuerte"
    elif score >= 60:
        señal = "🟢 Compra"
    elif score >= 45:
        señal = "🟡 Neutral/Esperar"
    elif score >= 35:
        señal = "🟠 Venta Parcial"
    else:
        señal = "🔴 Venta"
    
    return score, señal


# ═══════════════════════════════════════════════════════════════
# 9. PIPELINE PRINCIPAL
# ═══════════════════════════════════════════════════════════════

def descargar_precios(ticker, periodo="1y"):
    """Descarga precios de Yahoo Finance con manejo de errores."""
    try:
        data = yf.download(ticker, period=periodo, progress=False, auto_adjust=True)
        if data.empty:
            return None
        # Flatten multi-index columns if present
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)
        return data
    except Exception as e:
        print(f"  ⚠ Error descargando {ticker}: {e}")
        return None


def procesar_mercado(tickers_dict, macro_dict, nombre_mercado):
    """
    Procesa un mercado completo: descarga precios, calcula todos los scores.
    Retorna DataFrame con todas las columnas V1 + V2.
    """
    score_macro, detalles_macro = calcular_score_macro(macro_dict)
    print(f"\n{'='*60}")
    print(f"  {nombre_mercado} — Score Macro: {score_macro}")
    print(f"{'='*60}")
    
    resultados = []
    
    for ticker, (nombre, sector, score_sect) in tickers_dict.items():
        print(f"  Procesando {ticker}...", end=" ")
        
        # Descargar precios
        precios = descargar_precios(ticker)
        
        if precios is None or len(precios) < 10:
            print("⚠ Sin datos")
            resultados.append({
                'Ticker': ticker, 'Empresa': nombre, 'Sector': sector,
                'Mercado': nombre_mercado,
                'Precio actual': None, 'SEÑAL_V1': '⚠ Sin datos',
                'SEÑAL_V2': '⚠ Sin datos',
            })
            continue
        
        precio_actual = float(precios['Close'].iloc[-1])
        max_52s = float(precios['High'].max())
        min_52s = float(precios['Low'].min())
        dist_max_pct = ((precio_actual - max_52s) / max_52s) * 100.0
        
        # Variaciones
        var_1d = ((precio_actual / float(precios['Close'].iloc[-2])) - 1) * 100 if len(precios) >= 2 else 0
        var_5d = ((precio_actual / float(precios['Close'].iloc[-6])) - 1) * 100 if len(precios) >= 6 else 0
        var_1m = ((precio_actual / float(precios['Close'].iloc[-22])) - 1) * 100 if len(precios) >= 22 else 0
        
        # Score técnico
        score_tec, rsi, momentum, ma_cross = calcular_score_tecnico(precio_actual, precios)
        
        # R/R (Mejora 2)
        rr_ratio, rr_norm = calcular_rr(precio_actual, max_52s, min_52s)
        
        # Dist max normalizada (Mejora 3)
        dist_max_norm = normalizar_dist_max(dist_max_pct)
        
        # Score fundamental (placeholder — se reemplaza con ratios_consolidado_quant.csv)
        score_fund = 50.0  # Default; se sobreescribe al mergear con ratios
        
        # Asset Quality (Mejora 1)
        asset_quality = calcular_asset_quality(score_macro, score_fund, score_sect)
        
        # Entry Score (Mejora 1)
        entry_score = calcular_entry_score(score_tec, rr_norm, dist_max_norm)
        
        # Score Final V2
        score_v2, señal_v2 = calcular_score_final_v2(asset_quality, entry_score)
        
        # Score V1 (compatibilidad)
        score_v1, señal_v1 = calcular_score_v1(score_macro, score_tec, score_sect)
        
        # Ranking accionable
        ranking_acc = calcular_ranking_accionable(score_v2, rr_norm)
        
        resultado = {
            'Ticker': ticker,
            'Empresa': nombre,
            'Sector': sector,
            'Mercado': nombre_mercado,
            'Precio actual': round(precio_actual, 2),
            'Var 1d (%)': round(var_1d, 2),
            'Var 5d (%)': round(var_5d, 2),
            'Var 1m (%)': round(var_1m, 2),
            'Máximo 52s': round(max_52s, 2),
            'Mínimo 52s': round(min_52s, 2),
            'Dist. Máx 52s (%)': round(dist_max_pct, 1),
            'RSI(14)': rsi,
            'Momentum 21d (%)': momentum,
            'MA20>MA50': ma_cross,
            'Score Macro': score_macro,
            'Score Técnico': score_tec,
            'Score Sectorial': score_sect,
            'Score Fundamental': score_fund,
            # ── V1 (compatibilidad) ──
            'SCORE FINAL': score_v1,
            'SEÑAL': señal_v1,
            # ── V2 (mejoras) ──
            'ASSET_QUALITY': asset_quality,
            'ENTRY_SCORE': entry_score,
            'R/R Ratio': rr_ratio,
            'R/R Norm (0-100)': rr_norm,
            'Dist Max Norm (0-100)': dist_max_norm,
            'SCORE_FINAL_V2': score_v2,
            'SEÑAL_V2': señal_v2,
            'RANKING_ACCIONABLE': ranking_acc,
        }
        
        resultados.append(resultado)
        print(f"V1={score_v1} ({señal_v1})  V2={score_v2} ({señal_v2})  R/R={rr_ratio}  Rank={ranking_acc}")
    
    return pd.DataFrame(resultados), score_macro, detalles_macro


def integrar_ratios(df, ratios_path="ratios_consolidado_quant.csv"):
    """
    Mergea con ratios_consolidado_quant.csv para obtener Score Cuantitativo real.
    Recalcula Asset Quality y cascada descendente.
    """
    if not os.path.exists(ratios_path):
        print(f"  ⚠ {ratios_path} no encontrado, usando Score Fundamental = 50")
        return df
    
    try:
        df_rat = pd.read_csv(ratios_path, sep=';', decimal=',', encoding='utf-8-sig')
        df_rat.columns = [c.strip() for c in df_rat.columns]
        
        # Limpiar Score Cuantitativo
        if 'Score Cuantitativo' in df_rat.columns:
            df_rat['Score Cuantitativo'] = pd.to_numeric(
                df_rat['Score Cuantitativo'].astype(str).str.replace(' ','').str.replace(',','.'),
                errors='coerce'
            )
            
            # Mergear por Ticker
            score_map = dict(zip(df_rat['Ticker'], df_rat['Score Cuantitativo']))
            
            for idx, row in df.iterrows():
                sc = score_map.get(row['Ticker'])
                if sc is not None and not np.isnan(sc):
                    df.at[idx, 'Score Fundamental'] = sc
                    # Recalcular cascada
                    aq = calcular_asset_quality(row['Score Macro'], sc, row['Score Sectorial'])
                    es = row['ENTRY_SCORE']  # no cambia
                    sf2, sn2 = calcular_score_final_v2(aq, es)
                    ra = calcular_ranking_accionable(sf2, row['R/R Norm (0-100)'])
                    df.at[idx, 'ASSET_QUALITY'] = aq
                    df.at[idx, 'SCORE_FINAL_V2'] = sf2
                    df.at[idx, 'SEÑAL_V2'] = sn2
                    df.at[idx, 'RANKING_ACCIONABLE'] = ra
            
            print(f"  ✅ Ratios integrados: {len(score_map)} especies con Score Cuantitativo")
    except Exception as e:
        print(f"  ⚠ Error leyendo ratios: {e}")
    
    return df


# ═══════════════════════════════════════════════════════════════
# 10. EXPORTACIÓN A XLSX (compatible con pipeline dashboard)
# ═══════════════════════════════════════════════════════════════

def exportar_xlsx(df_all, macro_detalles, output_path="modelo_macro_micro_señales.xlsx"):
    """
    Exporta a Excel con sheets compatibles con el dashboard HTML.
    Agrega columnas V2 al final de cada sheet de señales.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # --- Sheet: Macro Variables ---
    ws_m = wb.active
    ws_m.title = "Macro Variables"
    
    headers_macro = ["País", "Variable", "Valor", "Score (0-100)", "Impacto"]
    for c, h in enumerate(headers_macro, 1):
        ws_m.cell(1, c, h).font = Font(bold=True)
    
    row = 2
    for pais, macro in [("Argentina", MACRO_ARG), ("Brasil", MACRO_BRA), ("EE.UU.", MACRO_USA)]:
        score_total, detalles = calcular_score_macro(macro)
        for var, (val, peor, mejor, imp, desc) in macro.items():
            ws_m.cell(row, 1, pais)
            ws_m.cell(row, 2, var)
            ws_m.cell(row, 3, val)
            ws_m.cell(row, 4, detalles[var])
            ws_m.cell(row, 5, "↑ mejor" if imp == 1 else "↓ mejor")
            row += 1
        # Score total
        ws_m.cell(row, 1, pais)
        ws_m.cell(row, 2, "SCORE MACRO TOTAL")
        ws_m.cell(row, 4, score_total)
        ws_m.cell(row, 4).font = Font(bold=True)
        row += 1
    
    # --- Sheets por mercado ---
    for mercado in ["MERVAL", "BOVESPA", "S&P500"]:
        m_label = mercado if mercado != "S&P500" else "S&P 500"
        df_m = df_all[df_all['Mercado'] == m_label].copy()
        
        if df_m.empty:
            continue
        
        ws = wb.create_sheet(f"{mercado} Señales")
        
        cols_export = [
            'Ticker', 'Empresa', 'Sector', 'Precio actual',
            'Var 1d (%)', 'Var 5d (%)', 'Var 1m (%)',
            'RSI(14)', 'Momentum 21d (%)', 'MA20>MA50',
            'Máximo 52s', 'Mínimo 52s', 'Dist. Máx 52s (%)',
            'Score Macro', 'Score Técnico', 'Score Sectorial',
            'SCORE FINAL', 'SEÑAL',
            # V2 columns
            'Score Fundamental', 'ASSET_QUALITY', 'ENTRY_SCORE',
            'R/R Ratio', 'R/R Norm (0-100)', 'Dist Max Norm (0-100)',
            'SCORE_FINAL_V2', 'SEÑAL_V2', 'RANKING_ACCIONABLE',
        ]
        
        for c, h in enumerate(cols_export, 1):
            ws.cell(1, c, h).font = Font(bold=True, size=9)
        
        for ri, (_, row_data) in enumerate(df_m.iterrows(), 2):
            for ci, col in enumerate(cols_export, 1):
                val = row_data.get(col)
                ws.cell(ri, ci, val)
    
    # --- Sheet: Ranking Combinado V2 ---
    ws_r = wb.create_sheet("Ranking Combinado")
    
    df_ranked = df_all.sort_values('RANKING_ACCIONABLE', ascending=False).reset_index(drop=True)
    
    rank_cols = ['Ticker', 'Mercado', 'Empresa', 'ASSET_QUALITY', 'ENTRY_SCORE',
                 'SCORE_FINAL_V2', 'SEÑAL_V2', 'R/R Ratio', 'RANKING_ACCIONABLE',
                 'SCORE FINAL', 'SEÑAL']
    
    for c, h in enumerate(rank_cols, 1):
        ws_r.cell(1, c, h).font = Font(bold=True, size=9)
    
    for ri, (_, row_data) in enumerate(df_ranked.iterrows(), 2):
        ws_r.cell(ri, 1, ri - 1)  # Rank number
        for ci, col in enumerate(rank_cols, 1):
            ws_r.cell(ri, ci, row_data.get(col))
    
    wb.save(output_path)
    print(f"\n✅ Exportado: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════
# 11. MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  MODELO INVERSIONES BURSÁTILES v2.0")
    print("  3 mejoras: Asset Quality | Entry Score | R/R Ranking")
    print("=" * 60)
    
    # Procesar mercados
    df_merval, sm_arg, dm_arg = procesar_mercado(MERVAL_TICKERS, MACRO_ARG, "MERVAL")
    df_bovespa, sm_bra, dm_bra = procesar_mercado(BOVESPA_TICKERS, MACRO_BRA, "BOVESPA")
    df_sp500, sm_usa, dm_usa = procesar_mercado(SP500_TICKERS, MACRO_USA, "S&P 500")
    
    # Combinar
    df_all = pd.concat([df_merval, df_bovespa, df_sp500], ignore_index=True)
    
    # Integrar ratios fundamentales
    ratios_paths = [
        "ratios_consolidado_quant.csv",
        "data/ratios_consolidado_quant.csv",
        "../ratios_consolidado_quant.csv",
    ]
    for rp in ratios_paths:
        if os.path.exists(rp):
            df_all = integrar_ratios(df_all, rp)
            break
    
    # Exportar
    macro_detalles = {"Argentina": dm_arg, "Brasil": dm_bra, "EE.UU.": dm_usa}
    xlsx_path = exportar_xlsx(df_all, macro_detalles)
    
    # CSV de respaldo
    csv_path = xlsx_path.replace('.xlsx', '.csv')
    df_all.to_csv(csv_path, index=False, sep=';', decimal=',')
    print(f"✅ CSV respaldo: {csv_path}")
    
    # Status JSON para webhook
    status = {
        "timestamp": datetime.now().isoformat(),
        "version": "2.0",
        "scores_macro": {"ARG": sm_arg, "BRA": sm_bra, "USA": sm_usa},
        "total_especies": len(df_all),
        "señales_v2": df_all['SEÑAL_V2'].value_counts().to_dict(),
        "top5_ranking": df_all.nlargest(5, 'RANKING_ACCIONABLE')[['Ticker','RANKING_ACCIONABLE','SEÑAL_V2']].to_dict('records'),
    }
    
    with open("run_status.json", "w") as f:
        json.dump(status, f, indent=2, ensure_ascii=False, default=str)
    print(f"✅ Status: run_status.json")
    
    # Resumen
    print(f"\n{'='*60}")
    print(f"  RESUMEN EJECUCIÓN")
    print(f"{'='*60}")
    print(f"  Scores Macro: ARG={sm_arg} | BRA={sm_bra} | USA={sm_usa}")
    print(f"  Especies procesadas: {len(df_all)}")
    print(f"\n  Señales V2:")
    for s, c in df_all['SEÑAL_V2'].value_counts().items():
        print(f"    {s}: {c}")
    print(f"\n  TOP 5 RANKING ACCIONABLE:")
    top5 = df_all.nlargest(5, 'RANKING_ACCIONABLE')
    for _, r in top5.iterrows():
        print(f"    {r['Ticker']:<12} {r['Mercado']:<10} Rank={r['RANKING_ACCIONABLE']}  {r['SEÑAL_V2']}  R/R={r['R/R Ratio']}")
    
    return df_all


if __name__ == "__main__":
    df = main()
