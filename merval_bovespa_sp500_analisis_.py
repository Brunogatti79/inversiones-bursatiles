# ==============================================================================
#  ANÁLISIS DE CIERRES DIARIOS - MERVAL (Argentina) + BOVESPA (Brasil) + S&P 500 (EE.UU.)
#  Yahoo Finance API via yfinance
#
#  Instalación previa:
#      pip install yfinance pandas openpyxl
#
#  Uso:
#      python merval_bovespa_sp500_analisis.py
#
#  Genera 4 archivos:
#      1. merval_cierres.csv           → Cierres diarios acciones MERVAL (ARS)
#      2. bovespa_cierres.csv          → Cierres diarios acciones BOVESPA (BRL)
#      3. sp500_cierres.csv            → Cierres diarios acciones S&P 500 (USD)
#      4. analisis_variaciones.xlsx    → Variaciones % por período + señales (6 hojas)
# ==============================================================================

import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta

# ── PARÁMETROS ─────────────────────────────────────────────────────────────────
START_DATE = "2025-01-01"
END_DATE   = datetime.today().strftime("%Y-%m-%d")
INTERVAL   = "1d"

# ── TICKERS MERVAL (Bolsa de Buenos Aires) ─────────────────────────────────────
MERVAL = {
    "GGAL.BA":  "Grupo Financiero Galicia",
    "BMA.BA":   "Banco Macro",
    "PAMP.BA":  "Pampa Energía",
    "YPF.BA":   "YPF",
    "TXAR.BA":  "Ternium Argentina",
    "ALUA.BA":  "Aluar",
    "CRES.BA":  "Cresud",
    "SUPV.BA":  "Supervielle",
    "CEPU.BA":  "Central Puerto",
    "LOMA.BA":  "Loma Negra",
    "MIRG.BA":  "Mirgor",
    "TECO2.BA": "Telecom Argentina",
    "TGSU2.BA": "Transportadora Gas del Sur",
    "VALO.BA":  "Grupo Supervielle (VALO)",
    "COME.BA":  "Soc. Comercial del Plata",
    "EDN.BA":   "Edenor",
    "HARG.BA":  "Holcim Argentina",
    "VIST.BA":  "Vista Energy",
    "TRAN.BA":  "Transener",
    "MOLI.BA":  "Molinos Río de la Plata",
    "BYMA.BA":  "BYMA",
    "IRSA.BA":  "IRSA",
    "^MERV":    "ÍNDICE MERVAL",
}

# ── TICKERS BOVESPA (Bolsa de São Paulo) ───────────────────────────────────────
BOVESPA = {
    "PETR4.SA": "Petrobras PN",
    "VALE3.SA": "Vale",
    "ITUB4.SA": "Itaú Unibanco",
    "BBDC4.SA": "Bradesco",
    "ABEV3.SA": "Ambev",
    "WEGE3.SA": "WEG",
    "RENT3.SA": "Localiza",
    "RDOR3.SA": "Rede D'Or",
    "BBAS3.SA": "Banco do Brasil",
    "MGLU3.SA": "Magazine Luiza",
    "SUZB3.SA": "Suzano",
    "JBSS3.SA": "JBS",
    "EQTL3.SA": "Equatorial",
    "RAIZ4.SA": "Raízen",
    "EMBR3.SA": "Embraer",
    "HAPV3.SA": "Hapvida",
    "LREN3.SA": "Lojas Renner",
    "CSNA3.SA": "CSN",
    "CYRE3.SA": "Cyrela",
    "EGIE3.SA": "Engie Brasil",
    "BPAC11.SA":"BTG Pactual",
    "CPLE6.SA": "Copel",
    "^BVSP":    "ÍNDICE BOVESPA",
}

# ── TICKERS S&P 500 (NYSE / NASDAQ) ───────────────────────────────────────────
# Principales componentes por sector + índice completo
SP500 = {
    # Tecnología
    "AAPL":   "Apple",
    "MSFT":   "Microsoft",
    "NVDA":   "NVIDIA",
    "GOOGL":  "Alphabet (Google)",
    "META":   "Meta Platforms",
    "AMZN":   "Amazon",
    # Finanzas
    "JPM":    "JPMorgan Chase",
    "BAC":    "Bank of America",
    "GS":     "Goldman Sachs",
    "V":      "Visa",
    # Energía
    "XOM":    "ExxonMobil",
    "CVX":    "Chevron",
    # Salud
    "JNJ":    "Johnson & Johnson",
    "UNH":    "UnitedHealth",
    "LLY":    "Eli Lilly",
    # Consumo
    "WMT":    "Walmart",
    "PG":     "Procter & Gamble",
    "KO":     "Coca-Cola",
    "MCD":    "McDonald's",
    # Industria
    "CAT":    "Caterpillar",
    "BA":     "Boeing",
    "GE":     "GE Aerospace",
    # Telecomunicaciones
    "TSLA":   "Tesla",
    "^GSPC":  "ÍNDICE S&P 500",         # Índice completo
}


# ── FUNCIONES ──────────────────────────────────────────────────────────────────

def descargar_cierres(tickers_dict, mercado, start, end, interval):
    """Descarga cierres diarios para un diccionario de tickers."""
    print(f"\n{'='*60}")
    print(f"  Descargando {mercado} ({len(tickers_dict)} tickers)")
    print(f"  Período: {start} → {end}")
    print(f"{'='*60}")

    closes = {}
    for ticker, nombre in tickers_dict.items():
        try:
            df = yf.download(
                ticker,
                start=start,
                end=end,
                interval=interval,
                progress=False,
                auto_adjust=True
            )
            if not df.empty:
                # Aplanar MultiIndex si lo hay
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = [col[0] for col in df.columns]
                closes[f"{ticker}|{nombre}"] = df["Close"].squeeze()
                print(f"  ✓ {ticker:12s} {nombre}")
            else:
                print(f"  ✗ {ticker:12s} sin datos")
        except Exception as e:
            print(f"  ✗ {ticker:12s} ERROR: {e}")

    result = pd.DataFrame(closes)
    result.index = pd.to_datetime(result.index).normalize()
    result.index.name = "Fecha"
    result = result.sort_index()
    return result


def calcular_variaciones(df):
    """Calcula variaciones % para 1d, 5d, 1m, 3m, YTD y agrega señales."""
    hoy = df.index[-1]
    resumen = []

    for col in df.columns:
        ticker, nombre = col.split("|", 1)
        serie = df[col].dropna()
        if len(serie) < 2:
            continue

        precio_actual = serie.iloc[-1]

        def var_pct(dias=None, desde=None):
            if desde is not None:
                sub = serie[serie.index >= desde]
                if len(sub) < 2:
                    return None
                return (sub.iloc[-1] / sub.iloc[0] - 1) * 100
            sub = serie.iloc[-dias-1:-1] if dias else serie
            if len(sub) < 1:
                return None
            return (precio_actual / sub.iloc[0] - 1) * 100

        v1d  = var_pct(1)
        v5d  = var_pct(5)
        v1m  = var_pct(21)
        v3m  = var_pct(63)
        ytd  = var_pct(desde=pd.Timestamp(f"{hoy.year}-01-01"))

        señal = "NEUTRAL"
        if v1m is not None and v3m is not None:
            if v1m > 5 and v3m > 10:
                señal = "🟢 COMPRA"
            elif v1m < -5 and v3m < -10:
                señal = "🔴 VENTA"
            elif v1m > 3:
                señal = "🟡 ALZA RECIENTE"
            elif v1m < -3:
                señal = "🟠 BAJA RECIENTE"

        resumen.append({
            "Ticker":          ticker,
            "Nombre":          nombre,
            "Precio actual":   round(precio_actual, 2),
            "Var 1 día (%)":   round(v1d, 2) if v1d  is not None else None,
            "Var 5 días (%)":  round(v5d, 2) if v5d  is not None else None,
            "Var 1 mes (%)":   round(v1m, 2) if v1m  is not None else None,
            "Var 3 meses (%)": round(v3m, 2) if v3m  is not None else None,
            "Var YTD (%)":     round(ytd, 2) if ytd  is not None else None,
            "Señal":           señal,
        })

    return pd.DataFrame(resumen)


def guardar_csv(df, filename, mercado):
    """Guarda DataFrame de cierres en CSV con formato argentino."""
    out = df.copy()
    out.index = out.index.strftime("%d/%m/%Y")
    out.columns = [c.split("|")[1] for c in out.columns]
    out.to_csv(filename, sep=";", decimal=",", encoding="utf-8-sig")
    print(f"\n  ✅ CSV guardado: {filename}  ({out.shape[0]} días x {out.shape[1]} activos)")


def aplicar_formato_excel(ws, df_shape, titulo):
    """Aplica formato visual consistente a una hoja de Excel."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    fill_par     = PatternFill("solid", fgColor="DCE6F1")
    fill_impar   = PatternFill("solid", fgColor="FFFFFF")
    border_thin  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # Insertar fila de título
    ws.insert_rows(1)
    ws["A1"] = titulo
    ws["A1"].font      = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")

    # Encabezado (fila 2 tras insertar título)
    for cell in ws[2]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border_thin
    ws.row_dimensions[2].height = 28

    # Filas de datos
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=1):
        fill = fill_par if i % 2 == 0 else fill_impar
        for cell in row:
            cell.fill      = fill
            cell.border    = border_thin
            cell.alignment = Alignment(horizontal="right")

    # Primera columna centrada
    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="center")

    # Ancho automático aproximado por columna
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "B3"


def guardar_excel(df_merval, df_bovespa, df_sp500,
                  var_merval, var_bovespa, var_sp500, filename):
    """Genera Excel con 6 hojas: cierres y variaciones de los 3 mercados."""

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        # ── Cierres ──────────────────────────────────────────────────────────
        for df, sheet in [
            (df_merval,  "MERVAL - Cierres"),
            (df_bovespa, "BOVESPA - Cierres"),
            (df_sp500,   "S&P 500 - Cierres"),
        ]:
            out = df.copy()
            out.index   = out.index.strftime("%d/%m/%Y")
            out.columns = [c.split("|")[1] for c in out.columns]
            out.to_excel(writer, sheet_name=sheet)

        # ── Variaciones ───────────────────────────────────────────────────────
        var_merval.to_excel(writer,  sheet_name="MERVAL - Variaciones",  index=False)
        var_bovespa.to_excel(writer, sheet_name="BOVESPA - Variaciones", index=False)
        var_sp500.to_excel(writer,   sheet_name="S&P 500 - Variaciones", index=False)

    # ── Aplicar formato post-escritura ────────────────────────────────────────
    from openpyxl import load_workbook
    wb = load_workbook(filename)

    titulos = {
        "MERVAL - Cierres":      f"MERVAL — Cierres Diarios  |  {START_DATE} → {END_DATE}",
        "BOVESPA - Cierres":     f"BOVESPA — Cierres Diarios  |  {START_DATE} → {END_DATE}",
        "S&P 500 - Cierres":     f"S&P 500 — Cierres Diarios  |  {START_DATE} → {END_DATE}",
        "MERVAL - Variaciones":  f"MERVAL — Variaciones %  |  {END_DATE}",
        "BOVESPA - Variaciones": f"BOVESPA — Variaciones %  |  {END_DATE}",
        "S&P 500 - Variaciones": f"S&P 500 — Variaciones %  |  {END_DATE}",
    }

    for sheet_name, titulo in titulos.items():
        ws = wb[sheet_name]
        aplicar_formato_excel(ws, None, titulo)

    wb.save(filename)
    print(f"\n  ✅ Excel guardado: {filename}  (6 hojas)")


# ── EJECUCIÓN PRINCIPAL ────────────────────────────────────────────────────────

if __name__ == "__main__":

    print("\n" + "="*60)
    print("  MERVAL + BOVESPA + S&P 500 — Análisis de Cierres Diarios")
    print(f"  Ejecutado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("="*60)

    # 1. Descargar datos
    df_merval  = descargar_cierres(MERVAL,  "MERVAL",  START_DATE, END_DATE, INTERVAL)
    df_bovespa = descargar_cierres(BOVESPA, "BOVESPA", START_DATE, END_DATE, INTERVAL)
    df_sp500   = descargar_cierres(SP500,   "S&P 500", START_DATE, END_DATE, INTERVAL)

    # 2. Exportar CSVs de cierres
    guardar_csv(df_merval,  "merval_cierres.csv",  "MERVAL")
    guardar_csv(df_bovespa, "bovespa_cierres.csv", "BOVESPA")
    guardar_csv(df_sp500,   "sp500_cierres.csv",   "S&P 500")

    # 3. Calcular variaciones
    var_merval  = calcular_variaciones(df_merval)
    var_bovespa = calcular_variaciones(df_bovespa)
    var_sp500   = calcular_variaciones(df_sp500)

    # 4. Resumen en consola
    print("\n── TOP 5 MERVAL por var. 1 mes ──────────────────────────")
    print(var_merval.sort_values("Var 1 mes (%)", ascending=False)
          [["Ticker","Nombre","Precio actual","Var 1 mes (%)","Señal"]]
          .head(5).to_string(index=False))

    print("\n── TOP 5 BOVESPA por var. 1 mes ─────────────────────────")
    print(var_bovespa.sort_values("Var 1 mes (%)", ascending=False)
          [["Ticker","Nombre","Precio actual","Var 1 mes (%)","Señal"]]
          .head(5).to_string(index=False))

    print("\n── TOP 5 S&P 500 por var. 1 mes ─────────────────────────")
    print(var_sp500.sort_values("Var 1 mes (%)", ascending=False)
          [["Ticker","Nombre","Precio actual","Var 1 mes (%)","Señal"]]
          .head(5).to_string(index=False))

    # 5. Exportar Excel consolidado
    guardar_excel(
        df_merval, df_bovespa, df_sp500,
        var_merval, var_bovespa, var_sp500,
        "analisis_variaciones.xlsx"
    )

    print("\n✅ Proceso completado. Archivos generados:")
    print("   → merval_cierres.csv")
    print("   → bovespa_cierres.csv")
    print("   → sp500_cierres.csv")
    print("   → analisis_variaciones.xlsx  (6 hojas)")
    print()
